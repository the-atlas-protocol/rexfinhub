"""Quantify NAV-vs-price impact with walk-back logic for both."""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
import psycopg2
import psycopg2.extras
from datetime import date, timedelta

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")

OUT = open("_nav_impact.txt", "w", encoding="utf-8")
def L(*a): print(*a); print(*a, file=OUT); OUT.flush()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

def index_sheet(sheet_name):
    ws = wb[sheet_name]
    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cols = {}
    for i, c in enumerate(header[1:], 1):
        if c and " Equity" in str(c):
            parts = str(c).split()
            key = f"{parts[0]}_LN" if parts[1] == "LN" else parts[0]
            cols[key] = i
    rows_by_date = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d and hasattr(d, 'year'):
            rows_by_date[d.date() if hasattr(d, 'date') else d] = row
    return cols, rows_by_date

L("Loading data_price, data_nav, microsector...")
p_cols, p_rows = index_sheet("data_price")
n_cols, n_rows = index_sheet("data_nav")

ws_ms = wb["microsector"]
ms_tickers = set(str(t).strip() for t in next(ws_ms.iter_rows(min_row=4, max_row=4, values_only=True))[1:] if t and isinstance(t, str))
L(f"  data_price: {len(p_cols)} tickers, {len(p_rows)} dates")
L(f"  data_nav:   {len(n_cols)} tickers, {len(n_rows)} dates")
L(f"  ETN tickers: {len(ms_tickers)}")

def walk_back_value(rows_by_date, cols, ticker, target):
    """Find latest non-null value for ticker on or before target."""
    if ticker not in cols: return None, None
    col = cols[ticker]
    cur_d = target
    for _ in range(10):
        row = rows_by_date.get(cur_d)
        if row:
            v = row[col] if col < len(row) else None
            if isinstance(v, (int, float)) and v > 0:
                return float(v), cur_d
        cur_d -= timedelta(days=1)
    return None, None

# ── DB pull: Feb (prior) positions at frozen/quarterly vendors ─────
conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
FROZEN_EX = [9, 10, 11, 13, 15, 16, 17]
cur.execute("""
    SELECT e.ticker, ex.name ex_name, c.name country,
           m.exchange_aum_usd feb_aum, mf.price_usd feb_price_db
    FROM etp_exchange_monthly_aum m JOIN etp e USING (etp_id)
    JOIN exchange ex USING (exchange_id) JOIN country c USING (country_id)
    JOIN etp_monthly_fund mf ON mf.etp_id = m.etp_id AND mf.month_id = m.month_id
    WHERE m.exchange_id = ANY(%s) AND m.month_id = 13 AND m.exchange_aum_usd > 0 AND mf.price_usd > 0
""", (FROZEN_EX,))
positions = cur.fetchall()
L(f"\nStale/frozen Feb positions to reprice for Mar: {len(positions)}")

# Reprice each both ways
PRIOR_END = date(2026, 2, 28)
CUR_END = date(2026, 3, 31)

L(f"\n=== Method comparison ===")
totals = {"old_all_price": 0, "new_etf_nav_etn_price": 0,
          "etf_old": 0, "etf_new": 0, "etn_old": 0, "etn_new": 0}
nav_missing_etfs = []
big_movers = []

for r in positions:
    t = r["ticker"]
    feb_aum = float(r["feb_aum"])
    is_etn = t in ms_tickers

    # OLD: shares = feb_aum / prior_price; new_aum = shares * cur_price
    prior_price, _ = walk_back_value(p_rows, p_cols, t, PRIOR_END)
    cur_price, _ = walk_back_value(p_rows, p_cols, t, CUR_END)
    if not (prior_price and cur_price):
        continue
    old_aum = feb_aum * (cur_price / prior_price)

    # NEW: ETN uses price (same), ETF uses NAV
    if is_etn:
        new_aum = old_aum
        totals["etn_old"] += old_aum; totals["etn_new"] += new_aum
    else:
        prior_nav, _ = walk_back_value(n_rows, n_cols, t, PRIOR_END)
        cur_nav, _ = walk_back_value(n_rows, n_cols, t, CUR_END)
        if prior_nav and cur_nav:
            new_aum = feb_aum * (cur_nav / prior_nav)
        else:
            new_aum = old_aum  # fallback
            nav_missing_etfs.append(t)
        totals["etf_old"] += old_aum; totals["etf_new"] += new_aum

    totals["old_all_price"] += old_aum
    totals["new_etf_nav_etn_price"] += new_aum
    diff = new_aum - old_aum
    if abs(diff) > 100_000:
        big_movers.append((t, r["country"], r["ex_name"], old_aum, new_aum, diff, is_etn))

L(f"OLD method (all use market price walk-back):")
L(f"  Total stale-repriced: ${totals['old_all_price']/1e6:.2f}M")
L(f"NEW method (ETNs price, ETFs NAV walk-back):")
L(f"  Total stale-repriced: ${totals['new_etf_nav_etn_price']/1e6:.2f}M")
L(f"  Δ vs old:             ${(totals['new_etf_nav_etn_price'] - totals['old_all_price'])/1e6:+.2f}M")
L("")
L(f"By type:")
L(f"  ETFs OLD: ${totals['etf_old']/1e6:.2f}M  NEW: ${totals['etf_new']/1e6:.2f}M  Δ ${(totals['etf_new'] - totals['etf_old'])/1e6:+.2f}M")
L(f"  ETNs OLD: ${totals['etn_old']/1e6:.2f}M  NEW: ${totals['etn_new']/1e6:.2f}M  Δ ${(totals['etn_new'] - totals['etn_old'])/1e6:+.2f}M")

L(f"\nETFs with no NAV data (fell back to price): {len(set(nav_missing_etfs))}")
if nav_missing_etfs:
    L(f"  Tickers: {sorted(set(nav_missing_etfs))[:20]}")

L(f"\nLargest position-level shifts (|Δ| > $100K):")
for t, country, ex, old, new, diff, is_etn in sorted(big_movers, key=lambda x: -abs(x[5]))[:25]:
    tag = "ETN" if is_etn else "ETF"
    L(f"  {t:<8} {tag:<3} {country:<11} {ex[:25]:<25}  ${old/1e6:>7.2f}M → ${new/1e6:>7.2f}M  ({diff/1e6:+.3f}M)")

# Coverage check: which RTREX tickers have NAV data
L(f"\n=== NAV COVERAGE ===")
cur.execute("SELECT ticker FROM etp")
all_rex = set(r["ticker"] for r in cur.fetchall())
nav_present = []
nav_absent = []
for t in all_rex:
    val, when = walk_back_value(n_rows, n_cols, t, CUR_END)
    if val:
        nav_present.append(t)
    else:
        nav_absent.append(t)
L(f"  REX tickers with NAV data (walked back from Mar 31): {len(nav_present)}/{len(all_rex)}")
L(f"  Tickers missing NAV: {sorted(nav_absent)}")

# Among ETNs specifically, which have NAV?
etn_with_nav = [t for t in ms_tickers if t in all_rex and t in nav_present]
etn_without = [t for t in ms_tickers if t in all_rex and t not in nav_present]
L(f"\n  ETNs with NAV: {len(etn_with_nav)}/{len(ms_tickers & all_rex)}")
L(f"  ETN tickers in NAV: {sorted(etn_with_nav)}")
L(f"  ETNs without NAV: {sorted(etn_without)}")

wb.close(); conn.close(); OUT.close()
