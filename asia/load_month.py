"""
Generic monthly data loader — parameterized by --month YYYY-MM.

Reads config/vendor_status.yaml to determine per-vendor behavior:
  - active vendors: dispatch to the specific parser, expect a file in grace_data/{month}/
  - frozen_permanent / waiting*: apply shares_invariant_reprice from prior-month shares
  - zeroed: skip (no rows inserted)

Usage:
    python load_month.py --month 2026-04

Lifecycle (inception/delist) is handled separately by refresh_all_months.py which reads W1.

Safety:
    - Backs up DB before writing (pg_dump)
    - Single transaction for all inserts/updates
    - Sanity checks against Grace's summary file (when present)
"""
import os, sys, io, argparse, subprocess, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
from datetime import date, timedelta
from decimal import Decimal

import yaml
import openpyxl
import psycopg2
import psycopg2.extras

BBG_PATH = r"C:\Users\RyuEl-Asmar\REX Financial LLC\REX Financial LLC - MasterFiles\MASTER Data\bloomberg_daily_file.xlsm"
CONFIG_PATH = Path("config/vendor_status.yaml")

# ── CLI ────────────────────────────────────────────────────────────────
ap = argparse.ArgumentParser()
ap.add_argument("--month", required=True, help="Target month YYYY-MM")
ap.add_argument("--data-dir", default=None, help="Override data dir (default: grace_data/{YYYY-MM})")
ap.add_argument("--apply", action="store_true", help="Actually write to DB (default: dry-run)")
args = ap.parse_args()

YEAR, MONTH = map(int, args.month.split("-"))
DATA_DIR = Path(args.data_dir) if args.data_dir else Path(f"grace_data/{args.month}")
DRY_RUN = not args.apply

print(f"{'='*70}\nload_month — {args.month}   mode={'DRY RUN' if DRY_RUN else 'APPLY'}\n{'='*70}")
print(f"Data dir: {DATA_DIR}")
if not DATA_DIR.exists():
    print(f"  (does not exist — frozen/waiting vendors will still be repriced)")

# ── Load config ────────────────────────────────────────────────────────
with open(CONFIG_PATH) as f:
    config = yaml.safe_load(f)
vendors = config["vendors"]
print(f"Loaded vendor config: {len(vendors)} vendors")

# ── DB connection ──────────────────────────────────────────────────────
conn = psycopg2.connect(host="localhost", port=int(os.environ.get("REX_ASIA_PORT","5433")), user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()

# Resolve month_id
cur.execute("""SELECT month_id, month_end FROM calendar_month
               WHERE EXTRACT(YEAR FROM month_end)=%s AND EXTRACT(MONTH FROM month_end)=%s""",
            (YEAR, MONTH))
row = cur.fetchone()
if not row:
    print(f"ERROR: no calendar_month for {args.month}. Insert one first:")
    print(f"  INSERT INTO calendar_month (month_id, month_end) VALUES (?, '{YEAR}-{MONTH:02d}-??');")
    sys.exit(1)
MONTH_ID = row["month_id"]
MONTH_END = row["month_end"]
print(f"month_id={MONTH_ID}  month_end={MONTH_END}")

cur.execute("SELECT ticker, etp_id FROM etp")
etp_map = {r["ticker"]: r["etp_id"] for r in cur.fetchall()}
cur.execute("""SELECT ex.name ex_name, c.name country, ex.exchange_id
               FROM exchange ex JOIN country c USING (country_id)""")
exchange_map = {(r["country"], r["ex_name"]): r["exchange_id"] for r in cur.fetchall()}

PRIOR_MONTH_ID = MONTH_ID - 1

# Delisted funds for this month — read W1 directly
print("Loading lifecycle from W1...")
import pandas as pd
w1 = pd.read_excel(BBG_PATH, sheet_name="w1", header=0)
w1["base"] = w1["Ticker"].astype(str).str.split().str[0]
w1["listing"] = w1["Ticker"].astype(str).str.split().str[1]
w1["db_key"] = w1.apply(lambda r: f"{r['base']}_LN" if r.get("listing") == "LN" else r["base"], axis=1)
delisted_before_or_on = set()
for _, r in w1.iterrows():
    if r["db_key"] in etp_map and pd.notna(r.get("Delist Date")):
        delist = pd.Timestamp(r["Delist Date"]).date()
        if delist <= MONTH_END:
            delisted_before_or_on.add(r["db_key"])
print(f"  {len(delisted_before_or_on)} tickers already delisted by {MONTH_END}")

# ── Helpers ────────────────────────────────────────────────────────────
def find_file(pattern):
    matches = list(DATA_DIR.glob(pattern)) if DATA_DIR.exists() else []
    if not matches: return None
    return matches[0]

# ── PARSERS ────────────────────────────────────────────────────────────
def parse_ksd(path):
    """KSD parser. Format changed in May 2026: Shares column inserted between Name and AUM.
    Now finds AUM column dynamically from row 4 headers."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # find AUM column from row 4 headers
    aum_col = 4
    for c in range(1, ws.max_column + 1):
        h = ws.cell(4, c).value
        if isinstance(h, str) and "AUM" in h.upper():
            aum_col = c; break
    rows = []
    for r in range(5, ws.max_row + 1):
        ticker = ws.cell(row=r, column=2).value
        aum = ws.cell(row=r, column=aum_col).value
        if ticker and isinstance(ticker, str) and ticker.isupper() and len(ticker) <= 6:
            if isinstance(aum, (int, float)) and aum > 0 and ticker not in delisted_before_or_on:
                rows.append((ticker.strip(), aum))
    wb.close()
    return rows

def parse_sbi(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        aum = ws.cell(row=r, column=4).value
        if isinstance(t, str) and t.isupper() and isinstance(aum, (int, float)) and aum > 0:
            if t not in delisted_before_or_on:
                rows.append((t.strip(), aum))
    wb.close(); return rows

def parse_rakuten(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        aum = ws.cell(row=r, column=5).value
        if isinstance(t, str) and t.isupper() and isinstance(aum, (int, float)) and aum > 0:
            if t not in delisted_before_or_on:
                rows.append((t.strip(), aum))
    wb.close(); return rows

def parse_monex(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        aum = ws.cell(row=r, column=4).value
        if isinstance(t, str) and t.isupper() and isinstance(aum, (int, float)) and aum > 0:
            if t not in delisted_before_or_on:
                rows.append((t.strip(), aum))
    wb.close(); return rows

def parse_matsui(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb.active
    rows = []
    for r in range(3, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        aum = ws.cell(row=r, column=3).value
        if isinstance(t, str) and t.isupper() and isinstance(aum, (int, float)) and aum > 0:
            if t not in delisted_before_or_on:
                rows.append((t.strip(), aum))
    wb.close(); return rows

def parse_moomoo_japan_monthly(path):
    """Oct 2025 shares (last per-ticker snapshot) × current Mar-end price,
    scaled to match Grace's current-month aggregate from 'by month&category'."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["by each etf"]
    # Find latest date column
    date_cols = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v and hasattr(v, "year"):
            date_cols.append((v.date(), c))
    if not date_cols:
        wb.close(); return []
    date_cols.sort()
    _, latest_col = date_cols[-1]
    shares_col = latest_col + 1

    oct_positions = {}
    for r in range(3, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        sh = ws.cell(row=r, column=shares_col).value
        if isinstance(t, str) and t.isupper() and isinstance(sh, (int, float)) and sh > 0:
            if t.strip() not in delisted_before_or_on:
                oct_positions[t.strip()] = float(sh)

    # Grace aggregate
    ws_cat = wb["by month&category"]
    grace_total = 0.0
    for r in range(2, ws_cat.max_row + 1):
        mo = ws_cat.cell(row=r, column=1).value
        usd = ws_cat.cell(row=r, column=5).value
        if mo and hasattr(mo, "year") and mo.year == YEAR and mo.month == MONTH:
            if isinstance(usd, (int, float)):
                grace_total += float(usd)
    wb.close()

    # Current-month prices
    cur.execute("""SELECT e.ticker, mf.price_usd FROM etp_monthly_fund mf JOIN etp e USING (etp_id)
                   WHERE mf.month_id = %s""", (MONTH_ID,))
    prices = {r["ticker"]: float(r["price_usd"]) for r in cur.fetchall()}

    raw_total = sum(oct_positions[t] * prices.get(t, 0) for t in oct_positions if t in prices)
    scale = grace_total / raw_total if raw_total > 0 else 1.0

    rows = []
    for t, sh in oct_positions.items():
        p = prices.get(t)
        if not p: continue
        aum = sh * p * scale
        rows.append((t, aum))
    return rows

def _parse_viewtrade_for_country(path, country):
    """ViewTrade monthly file format (2026-04+): single sheet 'April' (or
    most-recent month name) with columns A=Country, B=Client, C=Symbol,
    D=Qty, E=NAV, F=Value (USD). Groups by ticker for the given country,
    summing across all client brokers. Skips rows for delisted tickers."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active  # most-recent monthly tab is the active one
    by_ticker: dict[str, float] = {}
    for r in range(2, ws.max_row + 1):
        c = ws.cell(row=r, column=1).value
        t = ws.cell(row=r, column=3).value
        v = ws.cell(row=r, column=6).value
        if c == country and isinstance(t, str) and isinstance(v, (int, float)) and v > 0:
            tk = t.strip().upper()
            if tk and tk not in delisted_before_or_on:
                by_ticker[tk] = by_ticker.get(tk, 0) + v
    wb.close()
    return list(by_ticker.items())

def parse_viewtrade_hk(path): return _parse_viewtrade_for_country(path, "Hong Kong")
def parse_viewtrade_sg(path): return _parse_viewtrade_for_country(path, "Singapore")
def parse_viewtrade_tw(path): return _parse_viewtrade_for_country(path, "Taiwan")



def parse_hk_webull(path):
    """HK Webull: Ticker | Naming | ISIN | Remark | Market value USD. ISIN US* -> US ticker, IE* -> _LN UCITS variant."""
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        t = ws.cell(r, 1).value
        isin = ws.cell(r, 3).value
        aum = ws.cell(r, 5).value
        if not (isinstance(t, str) and t.strip()): continue
        if not isinstance(aum, (int, float)) or aum <= 0: continue
        tk = t.strip().upper()
        if isinstance(isin, str) and isin.startswith("IE"):
            tk = tk + "_LN"
        if tk in delisted_before_or_on: continue
        rows.append((tk, float(aum)))
    wb.close()
    return rows

def parse_ace_tesla(path):
    """ACE TESLA Value Chain ETF — single TSLT position from KSD file's side table."""
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb.active
    for r in range(5, 10):
        v = ws.cell(row=r, column=10).value
        if isinstance(v, (int, float)) and v > 0:
            wb.close()
            return [("TSLT", v)]
    wb.close()
    return []

def parse_syfe_verbal(_=None):
    """SYFE — verbal from Grace's summary. Read Mar column from summary file."""
    summary = find_file("Asset report summary*.xlsx")
    if not summary: return []
    wb = openpyxl.load_workbook(summary, data_only=True); ws = wb.active
    # Find the row with 'SYFE' in col B
    mar_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v and hasattr(v, "year") and v.year == YEAR and v.month == MONTH:
            mar_col = c; break
    if not mar_col: wb.close(); return []
    for r in range(1, ws.max_row + 1):
        if (ws.cell(row=r, column=2).value or "").strip() == "SYFE":
            v = ws.cell(row=r, column=mar_col).value
            wb.close()
            if isinstance(v, (int, float)) and v > 0:
                return [("FEPI", float(v) * 1_000_000)]  # summary is in $M
            return []
    wb.close(); return []

PARSERS = {
    "ksd": parse_ksd,
    "sbi": parse_sbi,
    "rakuten": parse_rakuten,
    "monex": parse_monex,
    "matsui": parse_matsui,
    "hk_webull": parse_hk_webull,
    "viewtrade_hk": parse_viewtrade_hk,
    "viewtrade_sg": parse_viewtrade_sg,
    "viewtrade_tw": parse_viewtrade_tw,
    "moomoo_japan_monthly": parse_moomoo_japan_monthly,
    "ace_tesla": parse_ace_tesla,
    "syfe_verbal": parse_syfe_verbal,
}

# ── Read Grace's aggregate value for a vendor from the summary file ──
def grace_summary_value(country, grace_source_name):
    """Return (value_in_usd, or None) for (country, source) in Grace's summary, current month."""
    summary = find_file("Asset report summary*.xlsx")
    if not summary: return None
    wb = openpyxl.load_workbook(summary, data_only=True); ws = wb.active
    month_col = None
    for r in range(1, 5):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and hasattr(v, "year") and v.year == YEAR and v.month == MONTH:
                month_col = c; break
        if month_col: break
    if not month_col: wb.close(); return None
    for r in range(3, ws.max_row + 1):
        c_country = (ws.cell(row=r, column=1).value or "")
        c_source = (ws.cell(row=r, column=2).value or "")
        c_country = str(c_country).strip().replace("\xa0", "").strip()
        c_source = str(c_source).strip()
        if c_country == country and c_source == grace_source_name:
            val = ws.cell(row=r, column=month_col).value
            wb.close()
            if isinstance(val, (int, float)) and val > 0:
                return float(val) * 1_000_000  # summary values in $M -> raw $
            return None
    wb.close(); return None

def scaled_to_grace_aggregate(exchange_id, country, grace_source_name):
    """Use prior-month per-fund shares, reprice at current price, then scale total to match Grace's aggregate."""
    grace_total = grace_summary_value(country, grace_source_name)
    if not grace_total or grace_total <= 0:
        print(f"    WARN: no Grace aggregate for ({country}, {grace_source_name}) — falling back to shares_invariant")
        return shares_invariant_reprice(exchange_id)

    # Prior shares (from prior_aum / prior_price)
    cur.execute("""
        SELECT e.ticker, m.exchange_aum_usd prior_aum, mf.price_usd prior_price
        FROM etp_exchange_monthly_aum m JOIN etp e USING (etp_id)
        JOIN etp_monthly_fund mf ON mf.etp_id = m.etp_id AND mf.month_id = m.month_id
        WHERE m.exchange_id = %s AND m.month_id = %s AND m.exchange_aum_usd > 0 AND mf.price_usd > 0
    """, (exchange_id, PRIOR_MONTH_ID))
    prior = cur.fetchall()
    cur.execute("""SELECT e.ticker, mf.price_usd FROM etp_monthly_fund mf JOIN etp e USING (etp_id)
                   WHERE mf.month_id = %s""", (MONTH_ID,))
    cur_prices = {r["ticker"]: float(r["price_usd"]) for r in cur.fetchall()}

    raw_rows = []
    raw_total = 0.0
    for r in prior:
        t = r["ticker"]
        if t in delisted_before_or_on: continue
        cp = cur_prices.get(t)
        if not cp: continue
        shares = float(r["prior_aum"]) / float(r["prior_price"])
        reprice = shares * cp
        raw_rows.append((t, shares, reprice))
        raw_total += reprice

    scale = grace_total / raw_total if raw_total > 0 else 1.0
    print(f"    Grace aggregate ${grace_total/1e6:.2f}M, raw reprice ${raw_total/1e6:.2f}M, scale={scale:.4f}")
    return [(t, reprice * scale) for (t, _, reprice) in raw_rows]

# ── Shares-invariant reprice for waiting/frozen vendors ────────────────
def shares_invariant_reprice(exchange_id):
    """Derive prior-month implied shares (prior AUM / prior price) and reprice at current price."""
    cur.execute("""
        SELECT e.ticker, m.exchange_aum_usd prior_aum, mf.price_usd prior_price
        FROM etp_exchange_monthly_aum m JOIN etp e USING (etp_id)
        JOIN etp_monthly_fund mf ON mf.etp_id = m.etp_id AND mf.month_id = m.month_id
        WHERE m.exchange_id = %s AND m.month_id = %s AND m.exchange_aum_usd > 0 AND mf.price_usd > 0
    """, (exchange_id, PRIOR_MONTH_ID))
    prior = cur.fetchall()
    cur.execute("""SELECT e.ticker, mf.price_usd FROM etp_monthly_fund mf JOIN etp e USING (etp_id)
                   WHERE mf.month_id = %s""", (MONTH_ID,))
    cur_prices = {r["ticker"]: float(r["price_usd"]) for r in cur.fetchall()}

    rows = []
    for r in prior:
        t = r["ticker"]
        if t in delisted_before_or_on: continue
        cp = cur_prices.get(t)
        if not cp: continue
        shares = float(r["prior_aum"]) / float(r["prior_price"])
        rows.append((t, shares * cp))
    return rows

# ── Dispatch per-vendor ────────────────────────────────────────────────
all_inserts = []  # (etp_id, exchange_id, aum_usd, shares, source)

for vendor_key, vcfg in vendors.items():
    country, exchange = vendor_key.split("/", 1)
    ex_id = exchange_map.get((country, exchange))
    if not ex_id:
        print(f"  [{vendor_key}]: EXCHANGE NOT IN DB — skipping")
        continue

    status = vcfg["status"]
    parser_name = vcfg.get("parser")
    rows_raw = []
    source = ""

    if status == "active":
        # Need file
        pattern = vcfg.get("file_pattern")
        file_path = find_file(pattern) if pattern else None
        if parser_name in ("ace_tesla", "syfe_verbal") and not file_path:
            # ACE Tesla uses KSD file; SYFE uses summary file
            if parser_name == "ace_tesla":
                file_path = find_file("*Korea REX report KSD*.xlsx")
        if parser_name == "syfe_verbal":
            pass  # syfe_verbal reads summary itself
        elif not file_path:
            print(f"  [{vendor_key}]: STATUS=active but no file matching {pattern} — ERROR")
            continue
        parser = PARSERS.get(parser_name)
        if not parser:
            print(f"  [{vendor_key}]: unknown parser '{parser_name}' — skipping")
            continue
        if parser_name == "syfe_verbal":
            rows_raw = parser(None)
        else:
            rows_raw = parser(file_path)
        source = "reported"
        print(f"  [{vendor_key}]: {len(rows_raw)} rows, ${sum(v for _,v in rows_raw)/1e6:.2f}M  [reported]")
    elif status == "zeroed":
        print(f"  [{vendor_key}]: ZEROED ({vcfg.get('reason','')[:60]}...) — no rows")
        continue
    elif status in ("frozen_permanent", "waiting_13f", "waiting_hardcopy", "waiting"):
        methodology = vcfg.get("methodology", "shares_invariant_reprice")
        if methodology == "shares_invariant_reprice":
            rows_raw = shares_invariant_reprice(ex_id)
            source = "repriced"
            print(f"  [{vendor_key}]: {len(rows_raw)} repriced, ${sum(v for _,v in rows_raw)/1e6:.2f}M  [{status}]")
        else:
            print(f"  [{vendor_key}]: unknown methodology '{methodology}' — skipping")
            continue
    elif status == "active_aggregate_only":
        # Grace gives aggregate total only; allocate via prior-month shares scaled to Grace's total
        grace_src = vcfg.get("grace_source_name", exchange)
        rows_raw = scaled_to_grace_aggregate(ex_id, country, grace_src)
        source = "scaled_aggregate"
        print(f"  [{vendor_key}]: {len(rows_raw)} scaled to Grace aggregate, ${sum(v for _,v in rows_raw)/1e6:.2f}M  [{status}]")
    else:
        print(f"  [{vendor_key}]: unknown status '{status}' — skipping")
        continue

    # Collect inserts
    for ticker, aum in rows_raw:
        etp_id = etp_map.get(ticker)
        if not etp_id:
            print(f"    skip unknown ticker {ticker}")
            continue
        # Infer shares
        cur.execute("SELECT price_usd FROM etp_monthly_fund WHERE etp_id=%s AND month_id=%s", (etp_id, MONTH_ID))
        row = cur.fetchone()
        shares = (float(aum) / float(row["price_usd"])) if (row and row["price_usd"]) else 0
        all_inserts.append((etp_id, ex_id, float(aum), shares, source))

total_new = sum(x[2] for x in all_inserts) / 1e6
print(f"\nTotal Asia to insert: ${total_new:.2f}M across {len(all_inserts)} (fund, exchange) positions")

if DRY_RUN:
    print(f"\nDRY RUN — no DB writes performed.")
    print(f"To apply: python load_month.py --month {args.month} --apply")
else:
    # Backup first
    backup_file = f"rex_asia_pre_load_{args.month.replace('-','')}.backup"
    print(f"\nBacking up to {backup_file}...")
    os.environ["PGPASSWORD"] = ""
    r = subprocess.run(["./pgsql/bin/pg_dump.exe", "-h", "localhost", "-p", os.environ.get("REX_ASIA_PORT","5433"),
                        "-U", "postgres", "-F", "c", "-f", backup_file, "rex_asia"],
                       capture_output=True, text=True)
    if r.returncode != 0:
        print(f"BACKUP FAILED: {r.stderr}")
        sys.exit(1)
    print(f"  Backup: {Path(backup_file).stat().st_size // 1024} KB")

    cur2 = conn.cursor()
    cur2.execute("DELETE FROM etp_exchange_monthly_aum WHERE month_id = %s", (MONTH_ID,))
    print(f"  Cleared {cur2.rowcount} existing month_id={MONTH_ID} rows")
    for etp_id, ex_id, aum, shares, source in all_inserts:
        cur2.execute("""
            INSERT INTO etp_exchange_monthly_aum (etp_id, exchange_id, month_id, exchange_aum_usd, shares_outstanding, source_type)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (etp_id, ex_id, MONTH_ID, Decimal(str(round(aum, 4))), Decimal(str(round(shares, 6))), source))
    conn.commit()
    print(f"  COMMITTED {len(all_inserts)} rows")
    print(f"\nRollback (if needed): pg_restore -h localhost -p 5433 -U postgres -c -d rex_asia {backup_file}")

conn.close()
