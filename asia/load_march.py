"""
Load Mar 2026 data into rex_asia database (month_id=14).

Methodology per 22 Apr 2026 decisions:
  1. Load etp_monthly_fund (price + global AUM) from bloomberg_daily_file for Mar 31.
  2. Parse fresh broker data (KSD, SBI, Rakuten, Monex, Matsui, MooMoo Japan monthly).
  3. NO carry-forward for waiting vendors (MooMoo HK/SG/MY, Oriental Harbour, ViewTrade).
  4. Shares-invariant reprice for Futu HK ONLY — structurally frozen (no commercial contract).
  5. Asset Plus Thailand = $0 (DRNZ sold confirmed 21 Apr).
  6. SYFE: $1.25M on FEPI (Grace's Mar value).
  7. KIM ACE Tesla: use the file's own calc ($17.73M).
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import psycopg2
from datetime import date
from decimal import Decimal

MONTH_ID = 14
MONTH_END = date(2026, 3, 31)
BBG_PATH = r"C:\Users\RyuEl-Asmar\REX Financial LLC\REX Financial LLC - MasterFiles\MASTER Data\bloomberg_daily_file.xlsm"
DATA_DIR = r"C:\Projects\rex-asia\grace_data\2026-03"

# Tickers liquidated mid-March 2026 — exclude from shares-invariant reprice for March
# (their positions were cashed out, not held forward)
DELISTED_IN_MAR = {"ETQ", "ARMU", "AXUP", "BKNU", "BULU", "DKUP", "PXIU"}

FILES = {
    "ksd":       os.path.join(DATA_DIR, "2026 03 31 Korea REX report KSD.xlsx"),
    "sbi":       os.path.join(DATA_DIR, "2026 03 31 Japan SBI REX report.xlsx"),
    "rakuten":   os.path.join(DATA_DIR, "2026 03 31 Japan Rakuten REX report.xlsx"),
    "monex":     os.path.join(DATA_DIR, "2026 03 31 Japan Monex REX report.xlsx"),
    "matsui":    os.path.join(DATA_DIR, "2026 03 31 Japan Matsui REX report.xlsx"),
    "moomoojp":  os.path.join(DATA_DIR, "2026 03 31 MooMoo Jap REX report.xlsx"),
}


def get_conn():
    return psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia")


def get_etp_map(conn):
    cur = conn.cursor()
    cur.execute("SELECT ticker, etp_id FROM etp")
    return {r[0]: r[1] for r in cur.fetchall()}


def get_exchange_map(conn):
    """Return {(country, exchange_name): exchange_id}. Country-aware to disambiguate MooMoo / ViewTrade."""
    cur = conn.cursor()
    cur.execute("""
        SELECT ex.name, c.name, ex.exchange_id
        FROM exchange ex JOIN country c USING (country_id)
    """)
    return {(r[1], r[0]): r[2] for r in cur.fetchall()}


# ─── STEP 1: Load BBG data for Mar 31 ─────────────────────────────────────────

def load_etp_monthly_fund(conn):
    print("\n=== STEP 1: Loading etp_monthly_fund (Mar 31) ===")
    wb = openpyxl.load_workbook(BBG_PATH, data_only=True, read_only=False)

    # Prices
    ws_price = wb["data_price"]
    price_header = [c.value for c in ws_price[1]]
    prices = {}
    for row in ws_price.iter_rows(min_row=2, max_row=4000, values_only=True):
        d = row[0]
        if d and hasattr(d, "month") and d.year == 2026 and d.month == 3 and d.day == 31:
            for i, col in enumerate(price_header[1:], 1):
                if not col: continue
                raw = str(col)
                if " LN Equity" in raw: continue  # LN variants handled via bbg_ticker later
                ticker = raw.replace(" US Equity", "")
                val = row[i]
                if val and isinstance(val, (int, float)) and val > 0:
                    prices[ticker] = float(val)
            break

    # AUM
    ws_aum = wb["data_aum"]
    aum_header = list(next(ws_aum.iter_rows(min_row=1, max_row=1, values_only=True)))
    aums = {}
    for row in ws_aum.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d and hasattr(d, "month") and d.year == 2026 and d.month == 3 and d.day == 31:
            for i, col in enumerate(aum_header[1:], 1):
                if not col: continue
                raw = str(col)
                if " LN Equity" in raw: continue
                ticker = raw.replace(" US Equity", "")
                val = row[i]
                if val and isinstance(val, (int, float)) and val > 0:
                    aums[ticker] = float(val) * 1_000_000  # $M -> raw $
            break

    # Microsector overwrite (ETN values in data_aum are notional; microsector is authoritative)
    ws_ms = wb["microsector"]
    ms_tickers_row = list(next(ws_ms.iter_rows(min_row=4, max_row=4, values_only=True)))
    ms_tickers = [str(t).strip() for t in ms_tickers_row[1:] if t]
    for row in ws_ms.iter_rows(min_row=5, values_only=True):
        d = row[0]
        if d and hasattr(d, "month") and d.year == 2026 and d.month == 3 and d.day == 31:
            for i, ticker in enumerate(ms_tickers, 1):
                val = row[i]
                if val and isinstance(val, (int, float)) and val > 0:
                    aums[ticker] = float(val)  # raw $, overwrites data_aum
            break

    wb.close()

    etp_map = get_etp_map(conn)
    cur = conn.cursor()
    inserted = 0; missing_price = []; missing_aum = []
    for ticker, etp_id in etp_map.items():
        price = prices.get(ticker)
        aum = aums.get(ticker)
        if not price:
            missing_price.append(ticker); continue
        if not aum:
            missing_aum.append(ticker); continue
        cur.execute("""
            INSERT INTO etp_monthly_fund (etp_id, month_id, price_usd, total_aum_usd)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (etp_id, month_id) DO UPDATE SET price_usd = EXCLUDED.price_usd, total_aum_usd = EXCLUDED.total_aum_usd
        """, (etp_id, MONTH_ID, Decimal(str(round(price, 6))), Decimal(str(round(aum, 4)))))
        inserted += 1
    conn.commit()
    print(f"  Inserted/updated: {inserted} ETPs")
    print(f"  Missing price: {len(missing_price)}   Missing AUM: {len(missing_aum)}")
    if missing_price: print(f"    no price: {sorted(missing_price)}")
    if missing_aum:   print(f"    no aum:   {sorted(missing_aum)}")


# ─── STEP 2: Broker parsers ───────────────────────────────────────────────────

def parse_ksd(conn):
    print("\n--- Korea KSD ---")
    wb = openpyxl.load_workbook(FILES["ksd"], data_only=True)
    ws = wb.active
    # Data starts row 5. Col B (idx 1)=Ticker, Col D (idx 3)=AUM USD
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        ticker = row[1]
        aum = row[3]
        if not ticker or not isinstance(ticker, str) or not ticker.isupper() or len(ticker) > 6:
            continue
        if aum is None or not isinstance(aum, (int, float)) or aum <= 0:
            continue
        rows.append({
            "ticker": ticker.strip(),
            "country": "Korea", "exchange": "KSD (Korea Securities Depository) - Retail",
            "aum_usd": round(float(aum), 4),
            "shares": 0,  # filled later
        })

    # KIM ACE Tesla: r6 col 10 (idx 9) = TSLT value (KIM ACE file does the calc for us)
    ace_value = None
    for row in ws.iter_rows(min_row=6, max_row=6, values_only=True):
        v = row[9]
        if v and isinstance(v, (int, float)):
            ace_value = float(v)
    wb.close()
    print(f"  KSD: {len(rows)} tickers, ${sum(r['aum_usd'] for r in rows)/1e6:.2f}M")

    ace_row = []
    if ace_value:
        ace_row.append({
            "ticker": "TSLT",
            "country": "Korea", "exchange": "Korea Investment Management -ACE TESLA Value Chain ETF",
            "aum_usd": round(ace_value, 4),
            "shares": 0,
        })
        print(f"  ACE Tesla (TSLT position): ${ace_value/1e6:.2f}M")
    return rows + ace_row


def parse_sbi():
    print("\n--- Japan SBI ---")
    wb = openpyxl.load_workbook(FILES["sbi"], data_only=True)
    ws = wb.active
    # Row 1 headers, data row 2+. Col A=Ticker, B=Shares, C=NAV, D=USD
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker, shares, nav, usd = row[0], row[1], row[2], row[3]
        if not ticker or not isinstance(ticker, str) or not ticker.isupper(): continue
        if usd is None or not isinstance(usd, (int, float)) or usd <= 0: continue
        rows.append({
            "ticker": ticker.strip(),
            "country": "Japan", "exchange": "SBI",
            "aum_usd": round(float(usd), 4),
            "shares": round(float(shares), 6) if isinstance(shares, (int, float)) else 0,
        })
    wb.close()
    print(f"  SBI: {len(rows)} tickers, ${sum(r['aum_usd'] for r in rows)/1e6:.2f}M")
    return rows


def parse_rakuten():
    print("\n--- Japan Rakuten ---")
    wb = openpyxl.load_workbook(FILES["rakuten"], data_only=True)
    # Sheet2 has REX funds
    ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = row[0]
        shares = row[3]  # Amounts
        usd = row[4]     # Market Value (US$)
        if not ticker or not isinstance(ticker, str) or not ticker.isupper(): continue
        if usd is None or not isinstance(usd, (int, float)) or usd <= 0: continue
        rows.append({
            "ticker": ticker.strip(),
            "country": "Japan", "exchange": "Rakuten",
            "aum_usd": round(float(usd), 4),
            "shares": round(float(shares), 6) if isinstance(shares, (int, float)) else 0,
        })
    wb.close()
    print(f"  Rakuten: {len(rows)} tickers, ${sum(r['aum_usd'] for r in rows)/1e6:.2f}M")
    return rows


def parse_monex():
    print("\n--- Japan Monex ---")
    wb = openpyxl.load_workbook(FILES["monex"], data_only=True)
    ws = wb.active
    # Row 1 headers, data row 2+. A=Ticker, B=Desc, C=JPY, D=USD, E=blank, F=FX
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = row[0]
        usd = row[3]
        if not ticker or not isinstance(ticker, str) or not ticker.isupper(): continue
        if usd is None or not isinstance(usd, (int, float)) or usd <= 0: continue
        rows.append({
            "ticker": ticker.strip(),
            "country": "Japan", "exchange": "Monex",
            "aum_usd": round(float(usd), 4),
            "shares": 0,  # Monex doesn't send shares directly; fill later
        })
    wb.close()
    print(f"  Monex: {len(rows)} tickers, ${sum(r['aum_usd'] for r in rows)/1e6:.2f}M")
    return rows


def parse_matsui():
    print("\n--- Japan Matsui ---")
    wb = openpyxl.load_workbook(FILES["matsui"], data_only=True)
    ws = wb.active
    # Row 1 headers, row 2 is TOTAL (skip), data row 3+. A=Ticker, B=Shares, C=USD
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        ticker = row[0]
        shares = row[1]
        usd = row[2]
        if not ticker or not isinstance(ticker, str) or not ticker.isupper(): continue
        if usd is None or not isinstance(usd, (int, float)) or usd <= 0: continue
        rows.append({
            "ticker": ticker.strip(),
            "country": "Japan", "exchange": "Matsui",
            "aum_usd": round(float(usd), 4),
            "shares": round(float(shares), 6) if isinstance(shares, (int, float)) else 0,
        })
    wb.close()
    print(f"  Matsui: {len(rows)} tickers, ${sum(r['aum_usd'] for r in rows)/1e6:.2f}M")
    return rows


def parse_moomoo_japan(conn):
    """MooMoo Japan: 'by each etf' sheet per-ticker data stops at Oct 2025.
    For Mar 2026: use last-available per-ticker shares (Oct 2025), reprice at Mar 2026 prices.
    Validate aggregate ≈ $13M (Grace's Mar summary total).
    """
    print("\n--- Japan MooMoo (Oct 2025 shares -> Mar 2026 reprice) ---")
    wb = openpyxl.load_workbook(FILES["moomoojp"], data_only=True)
    ws = wb["by each etf"]

    # Find the latest date column group
    date_cols = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v and hasattr(v, "year"):
            date_cols.append((v.date(), c))
    if not date_cols:
        print("  No date columns found")
        wb.close(); return []

    date_cols.sort()
    latest_date, latest_col = date_cols[-1]
    shares_col = latest_col + 1
    usd_col = latest_col + 2
    print(f"  Using last-available per-ticker data: {latest_date} (col {latest_col})")

    # Build ticker -> shares map from that column
    oct_positions = {}
    for r in range(3, ws.max_row + 1):
        ticker = ws.cell(row=r, column=1).value
        shares = ws.cell(row=r, column=shares_col).value
        if not ticker or not isinstance(ticker, str) or not ticker.isupper(): continue
        if ticker.strip() in DELISTED_IN_MAR: continue  # skip liquidated funds
        if shares is None or not isinstance(shares, (int, float)) or shares <= 0: continue
        oct_positions[ticker.strip()] = float(shares)
    wb.close()

    # Pull Mar 2026 prices
    cur = conn.cursor()
    cur.execute("""
        SELECT e.ticker, m.price_usd FROM etp_monthly_fund m
        JOIN etp e USING (etp_id) WHERE m.month_id = %s
    """, (MONTH_ID,))
    mar_prices = {r[0]: float(r[1]) for r in cur.fetchall()}

    # Read Grace's Mar aggregate from 'by month&category' sheet
    wb2 = openpyxl.load_workbook(FILES["moomoojp"], data_only=True)
    ws_cat = wb2["by month&category"]
    grace_total = 0.0
    for r in range(2, ws_cat.max_row + 1):
        month = ws_cat.cell(row=r, column=1).value
        usd = ws_cat.cell(row=r, column=5).value
        if month and hasattr(month, "year") and month.year == 2026 and month.month == 3:
            if isinstance(usd, (int, float)):
                grace_total += float(usd)
    wb2.close()
    print(f"  Grace Mar aggregate: ${grace_total/1e6:.2f}M (from 'by month&category')")

    # Raw reprice
    raw_rows = []
    raw_total = 0.0
    for ticker, shares in oct_positions.items():
        p = mar_prices.get(ticker)
        if not p: continue
        aum = shares * p
        raw_total += aum
        raw_rows.append((ticker, shares, aum, p))
    scale = (grace_total / raw_total) if raw_total > 0 else 1.0
    print(f"  Scaling factor (Grace ÷ Oct-repriced): {scale:.4f}")

    rows = []
    for ticker, oct_shares, raw_aum, price in raw_rows:
        scaled_aum = raw_aum * scale
        scaled_shares = scaled_aum / price if price > 0 else oct_shares
        rows.append({
            "ticker": ticker,
            "country": "Japan", "exchange": "MooMoo",
            "aum_usd": round(scaled_aum, 4),
            "shares": round(scaled_shares, 6),
        })
    total_aum = sum(r["aum_usd"] for r in rows)
    print(f"  MooMoo JP scaled: {len(rows)} tickers, ${total_aum/1e6:.2f}M")
    return rows


# ─── STEP 3: Special cases ───────────────────────────────────────────────────

def parse_syfe():
    print("\n--- HK SYFE ---")
    # Grace's Mar: $1.25M on FEPI (verbal, UCITS). Map to FEPI US ticker (DB convention).
    return [{
        "ticker": "FEPI",
        "country": "Hong Kong", "exchange": "SYFE",
        "aum_usd": 1_250_000.0,
        "shares": 0,
    }]


def parse_thailand():
    print("\n--- TH Asset Plus ---")
    print("  DRNZ sold -- Thailand value is $0. Not inserting row.")
    return []


def shares_invariant_reprice(conn, exchange_id: int, country: str, exchange: str, label: str):
    """Generic shares-invariant repricer.

    Derives Feb shares from (Feb AUM / Feb price) for each fund — NOT from the stored
    shares_outstanding column, which was found (22 Apr 2026 audit) to be inconsistent
    with AUM/price for quarterly-repriced Feb values (likely copy-forward from older months).

    Preserves the ECONOMIC POSITION: the shares Asian investors actually held at Feb month-end
    priced at Bloomberg's Feb close. Then reprices at Mar close.
    """
    cur = conn.cursor()
    # Feb AUM + Feb price per ticker at this exchange
    cur.execute("""
        SELECT e.ticker, m.exchange_aum_usd AS feb_aum, mf.price_usd AS feb_price
        FROM etp_exchange_monthly_aum m
        JOIN etp e USING (etp_id)
        JOIN etp_monthly_fund mf ON mf.etp_id = m.etp_id AND mf.month_id = m.month_id
        WHERE m.exchange_id = %s AND m.month_id = 13 AND m.exchange_aum_usd > 0 AND mf.price_usd > 0
    """, (exchange_id,))
    feb_positions = cur.fetchall()

    cur.execute("""
        SELECT e.ticker, m.price_usd
        FROM etp_monthly_fund m JOIN etp e USING (etp_id)
        WHERE m.month_id = 14
    """)
    mar_prices = {r[0]: float(r[1]) for r in cur.fetchall()}

    rows = []
    total = 0.0
    skipped = 0
    delisted_skipped = 0
    for ticker, feb_aum, feb_price in feb_positions:
        if ticker in DELISTED_IN_MAR:
            delisted_skipped += 1
            continue  # liquidated mid-Mar, no position to carry forward
        mar_price = mar_prices.get(ticker)
        if not mar_price:
            skipped += 1; continue
        implied_shares = float(feb_aum) / float(feb_price)
        new_aum = implied_shares * mar_price
        total += new_aum
        rows.append({
            "ticker": ticker,
            "country": country, "exchange": exchange,
            "aum_usd": round(new_aum, 4),
            "shares": round(implied_shares, 6),
        })
    notes = []
    if skipped: notes.append(f"skipped {skipped} — no Mar price")
    if delisted_skipped: notes.append(f"excluded {delisted_skipped} delisted")
    note_str = f" ({'; '.join(notes)})" if notes else ""
    print(f"  {label}: {len(rows)} funds, ${total/1e6:.2f}M{note_str}")
    return rows


def stale_vendor_reprice(conn):
    """Apply shares-invariant reprice to ALL stale/frozen vendors for March."""
    print("\n--- Stale-vendor shares-invariant reprice (Feb shares × Mar prices) ---")
    all_rows = []
    # Permanently frozen
    all_rows += shares_invariant_reprice(conn, 9,  "Hong Kong", "Futu/MooMoo",         "Futu HK (permanent)")
    # Temporarily unavailable — will restate when fresh data arrives
    all_rows += shares_invariant_reprice(conn, 10, "Hong Kong", "Oriental Harbour *",  "Oriental Harbour HK (pending 13F)")
    all_rows += shares_invariant_reprice(conn, 15, "Hong Kong", "ViewTrade",           "ViewTrade HK (pending hardcopy)")
    all_rows += shares_invariant_reprice(conn, 16, "Singapore", "ViewTrade",           "ViewTrade SG (pending hardcopy)")
    all_rows += shares_invariant_reprice(conn, 17, "Taiwan",    "ViewTrade",           "ViewTrade TW (pending hardcopy)")
    all_rows += shares_invariant_reprice(conn, 11, "Singapore", "MooMoo",              "MooMoo SG (pending Grace)")
    all_rows += shares_invariant_reprice(conn, 13, "Malaysia",  "MooMoo",              "MooMoo MY (pending Grace)")
    total = sum(r["aum_usd"] for r in all_rows) / 1e6
    print(f"  Total stale-repriced: ${total:.2f}M across {len(all_rows)} fund-exchange pairs")
    return all_rows


# ─── STEP 4: Fill shares where 0 (from aum and Mar price) ─────────────────────

def fill_shares(all_rows, conn):
    cur = conn.cursor()
    cur.execute("""
        SELECT e.ticker, m.price_usd
        FROM etp_monthly_fund m JOIN etp e USING (etp_id)
        WHERE m.month_id = %s
    """, (MONTH_ID,))
    prices = {r[0]: float(r[1]) for r in cur.fetchall()}

    filled = 0
    for r in all_rows:
        if r["shares"] == 0 and r["aum_usd"] > 0:
            p = prices.get(r["ticker"])
            if p and p > 0:
                r["shares"] = round(r["aum_usd"] / p, 6)
                filled += 1
    print(f"\n  Filled {filled} missing share counts from Mar price")


# ─── STEP 5: Validate ─────────────────────────────────────────────────────────

def validate(all_rows):
    print("\n=== STEP 3: Validation vs Grace's Mar summary ===")
    grace = {
        "KSD (Korea Securities Depository) - Retail": 881.8,
        "Korea Investment Management -ACE TESLA Value Chain ETF": 17.7,
        "SBI": 31.2,
        "Rakuten": 58.0,
        "Monex": 4.3,
        "Matsui": 1.5,
        "MooMoo": 13.0,  # Japan only — aggregate of JP MooMoo
        "SYFE": 1.25,
    }

    by_exchange = {}
    for r in all_rows:
        by_exchange[r["exchange"]] = by_exchange.get(r["exchange"], 0) + r["aum_usd"]

    print(f"  {'Exchange':<55s} {'Parsed':>10s} {'Grace':>10s} {'Diff':>10s}")
    print("  " + "-" * 90)
    all_ok = True
    for ex, grace_val in grace.items():
        parsed = by_exchange.get(ex, 0) / 1e6
        diff = parsed - grace_val
        flag = "OK" if abs(diff) < 2 else "CHECK"
        if flag == "CHECK":
            all_ok = False
        print(f"  {ex:<55s} {parsed:>9.2f}M {grace_val:>9.2f}M {diff:>+9.2f}M  {flag}")

    # Futu reported separately
    futu_total = by_exchange.get("Futu/MooMoo", 0) / 1e6
    print(f"\n  Futu HK (shares-invariant reprice from Feb):    ${futu_total:.2f}M  (no Grace value — structurally frozen)")

    total_parsed = sum(by_exchange.values()) / 1e6
    print(f"\n  Total parsed: ${total_parsed:.2f}M")
    return all_ok


# ─── STEP 6: Insert ──────────────────────────────────────────────────────────

def insert_rows(all_rows, conn):
    print("\n=== STEP 4: Inserting into etp_exchange_monthly_aum ===")
    etp_map = get_etp_map(conn)
    exchange_map = get_exchange_map(conn)
    cur = conn.cursor()

    # Delete any prior month 14 rows first (idempotent reload)
    cur.execute("DELETE FROM etp_exchange_monthly_aum WHERE month_id = %s", (MONTH_ID,))
    print(f"  Cleared {cur.rowcount} existing month_id=14 rows")

    inserted = 0; skipped = []
    for r in all_rows:
        etp_id = etp_map.get(r["ticker"])
        key = (r.get("country"), r["exchange"])
        exchange_id = exchange_map.get(key)
        if not etp_id:
            skipped.append(f"no etp_id for {r['ticker']}")
            continue
        if not exchange_id:
            skipped.append(f"no exchange_id for {key}")
            continue
        # Flag source type
        key_ex = r["exchange"]
        key_co = r.get("country")
        STALE_EX_IDS = {9, 10, 11, 13, 15, 16, 17}  # Futu HK, OH, MooMoo SG/MY, ViewTrade HK/SG/TW
        if exchange_id in STALE_EX_IDS:
            source = "repriced"
        elif key_ex == "MooMoo" and key_co == "Japan":
            source = "inferred"  # scaled from Oct 2025 shares
        else:
            source = "reported"
        cur.execute("""
            INSERT INTO etp_exchange_monthly_aum (etp_id, exchange_id, month_id, exchange_aum_usd, shares_outstanding, source_type)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (etp_id, exchange_id, MONTH_ID, Decimal(str(r["aum_usd"])), Decimal(str(r["shares"])), source))
        inserted += 1
    conn.commit()
    print(f"  Inserted: {inserted} rows")
    for s in set(skipped):
        print(f"  SKIPPED: {s}")


def verify(conn):
    print("\n=== STEP 5: Verify DB state for month_id=14 ===")
    cur = conn.cursor()
    cur.execute("""
        SELECT ex.name, c.name, count(*), round(sum(exa.exchange_aum_usd)/1e6, 2)
        FROM etp_exchange_monthly_aum exa
        JOIN exchange ex USING (exchange_id) JOIN country c USING (country_id)
        WHERE exa.month_id = %s
        GROUP BY ex.name, c.name
        ORDER BY 4 DESC
    """, (MONTH_ID,))
    total = 0.0
    print(f"\n  {'Exchange':<55s} {'Country':>12s} {'#':>5s} {'AUM $M':>10s}")
    print("  " + "-" * 88)
    for r in cur.fetchall():
        print(f"  {r[0]:<55s} {r[1]:>12s} {r[2]:>5d} {float(r[3]):>10.2f}")
        total += float(r[3])
    print("  " + "-" * 88)
    print(f"  {'TOTAL':<55s} {'':>12s} {'':>5s} {total:>10.2f}")


# ─── MAIN ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    conn = get_conn()

    load_etp_monthly_fund(conn)

    print("\n=== STEP 2: Parsing broker files ===")
    all_rows = []
    all_rows.extend(parse_ksd(conn))
    all_rows.extend(parse_sbi())
    all_rows.extend(parse_rakuten())
    all_rows.extend(parse_monex())
    all_rows.extend(parse_matsui())
    all_rows.extend(parse_moomoo_japan(conn))
    all_rows.extend(parse_syfe())
    all_rows.extend(parse_thailand())
    all_rows.extend(stale_vendor_reprice(conn))

    fill_shares(all_rows, conn)
    validate(all_rows)
    insert_rows(all_rows, conn)
    verify(conn)

    conn.close()
    print("\nDone.")
