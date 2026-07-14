"""
Read report_data.json, compute all derived values, write enriched_report_data.json.

This replaces the previously-missing enrichment step that was done ad-hoc
in a terminal session for Feb 2026. Now fully reproducible.

Computes:
  - bloomberg_total_rex_aum: Bloomberg-sourced sum of ALL REX tickers (US + LN) with microsector overwrite.
  - headlines: total Asia AUM, prior, dollar change, MoM%, flows, market move, % in Asia.
  - suites: per-family aggregates with flows/market move/KPIs.
  - funds: per-fund derived (market_move, flows, pct_of_global, mom, global_mom, dollar_change).
  - narrative: text strings used in the HTML (data-driven so no stale bullets).

Usage:
    python enrich_report_data.py --month 2026-03
         --input report_data.json --output enriched_report_data.json
"""
import argparse
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl

BBG_PATH = r"C:\Users\RyuEl-Asmar\REX Financial LLC\REX Financial LLC - MasterFiles\MASTER Data\bloomberg_daily_file.xlsm"


def bloomberg_total_rex_aum(year: int, month: int) -> tuple[float, int, list]:
    """Sum AUM for ALL REX tickers (US + LN + ETN microsector overwrite) from Bloomberg month-end.
    Returns (total_usd, tickers_counted, missing_bbg_columns).
    """
    wb = openpyxl.load_workbook(BBG_PATH, data_only=True, read_only=False)

    # data_aum (values in $M). Includes all Bloomberg-tracked tickers.
    ws_aum = wb["data_aum"]
    aum_header = list(next(ws_aum.iter_rows(min_row=1, max_row=1, values_only=True)))
    ticker_aum = {}
    for row in ws_aum.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if not (d and hasattr(d, "year") and d.year == year and d.month == month):
            continue
        # Prefer the LAST row of the month (month-end close)
        for i, col in enumerate(aum_header[1:], 1):
            if not col: continue
            raw = str(col)
            _p = raw.split()
            if len(_p) < 2 or _p[1] not in ("US","LN"): continue
            val = row[i]
            if val is None or not isinstance(val, (int, float)) or val <= 0: continue
            parts = raw.split()
            if len(parts) < 2: continue
            bbg_key = f"{parts[0]} {parts[1]}"  # e.g. "TSLT US" or "FEPI LN"
            ticker_aum[bbg_key] = float(val) * 1_000_000  # $M -> $

    # Microsector overwrite for ETNs (raw $)
    ws_ms = wb["microsector_aum"]
    ms_tickers = [str(v).strip() if v else None for v in next(ws_ms.iter_rows(min_row=4, max_row=4, values_only=True))]
    ms_tickers = ms_tickers[1:]  # skip Date
    for row in ws_ms.iter_rows(min_row=5, values_only=True):
        d = row[0]
        if not (d and hasattr(d, "year") and d.year == year and d.month == month):
            continue
        for i, t in enumerate(ms_tickers, 1):
            if not t: continue
            val = row[i]
            if val is None or not isinstance(val, (int, float)) or val <= 0: continue
            key = f"{t} US"  # MicroSectors ETNs are US-listed
            ticker_aum[key] = float(val)

    wb.close()
    # REX tickers live in DB — but for the total, we want ALL Bloomberg-tracked REX.
    # Simplest: match tickers by their product_family heuristic? No — use the etp table instead.
    # The audit here is a TOTAL across REX, so we need the REX universe list from the DB.
    import psycopg2
    conn = psycopg2.connect(host="localhost", port=int(__import__("os").environ.get("REX_ASIA_PORT","5433")), user="postgres", dbname="rex_asia")
    cur = conn.cursor()
    cur.execute("SELECT ticker FROM etp")
    rex_tickers = [r[0] for r in cur.fetchall()]
    conn.close()

    total = 0.0
    counted = 0
    missing = []
    for t in rex_tickers:
        base = t[:-3] if t.endswith("_LN") else t
        tried = [f"{base} LN" if t.endswith("_LN") else f"{base} US", f"{base} LN"]
        hit = None
        for key in tried:
            if key in ticker_aum:
                hit = ticker_aum[key]; break
        if hit is not None:
            total += hit; counted += 1
        else:
            missing.append(t)
    return total, counted, missing


def compute_market_move_and_flows(asia_aum, asia_prior, global_aum, global_prior):
    """Market move = asia_prior × (global_aum / global_prior - 1).
    Flows = (asia_aum - asia_prior) - market_move.
    """
    if asia_prior <= 0 or global_prior <= 0:
        return 0.0, (asia_aum - asia_prior)
    pct_change = (global_aum / global_prior) - 1
    market_move = asia_prior * pct_change
    dollar_change = asia_aum - asia_prior
    flows = dollar_change - market_move
    return market_move, flows


def build_narrative(headlines, suites, countries, exchanges, funds, fund_countries, month_label, month_short):
    """Computed strings for the HTML — fully data-driven, no hardcoded bullets."""
    def fmt_bn(x): return f"${x/1e9:.1f}B"
    def fmt_m(x):  return f"${x/1e6:.1f}M" if abs(x) >= 1e6 else f"${x/1e3:.0f}K"
    def fmt_signed_m(x):
        sign = "+" if x >= 0 else "-"
        return f"{sign}${abs(x)/1e6:.1f}M"
    def pct1(x): return f"{x*100:.1f}%"

    # Build per-fund per-country allocation lookup
    country_by_fund = {}
    for fc in fund_countries:
        country_by_fund.setdefault(fc["ticker"], {})[fc["country"]] = float(fc["aum"])

    def suite_country_share(suite_name):
        totals = {}
        for f in funds:
            if f.get("family_name") != suite_name: continue
            for c, v in country_by_fund.get(f["ticker"], {}).items():
                totals[c] = totals.get(c, 0) + v
        total = sum(totals.values())
        return {c: (v, v/total if total>0 else 0) for c, v in sorted(totals.items(), key=lambda x: -x[1])}, total

    # T-REX
    trex = suites.get("T-REX", {})
    trex_shares, trex_total = suite_country_share("T-REX")
    trex_funds = [f for f in funds if f.get("family_name") == "T-REX" and f.get("asia_aum",0) > 0]
    trex_top_flow = max(trex_funds, key=lambda f: f.get("flows", 0)) if trex_funds else None
    trex_top_aum  = max(trex_funds, key=lambda f: f.get("asia_aum", 0)) if trex_funds else None

    # MicroSectors
    ms_shares, ms_total = suite_country_share("MicroSectors")
    ms_funds = [f for f in funds if f.get("family_name") == "MicroSectors" and f.get("asia_aum",0) > 0]
    ms_top_mm  = max(ms_funds, key=lambda f: f.get("market_move", 0)) if ms_funds else None
    ms_worst_mm = min(ms_funds, key=lambda f: f.get("market_move", 0)) if ms_funds else None

    # Income suite — split into EPI + G&I by ticker family convention
    EPI_TICKERS = {"AIPI", "FEPI", "CEPI", "ATCL"}
    GI_TICKERS = {"TSII", "NVII", "MSII", "COII", "ULTI", "WMTI", "LLII", "PLTI",
                  "HOII", "GIF", "CWII"}
    epi_funds = [f for f in funds if f["ticker"] in EPI_TICKERS and f.get("asia_aum",0) > 0]
    gi_funds  = [f for f in funds if f["ticker"] in GI_TICKERS  and f.get("asia_aum",0) > 0]
    epi_asia = sum(f["asia_aum"] for f in epi_funds)
    gi_asia  = sum(f["asia_aum"] for f in gi_funds)

    # EPI country share
    epi_country = {}
    for f in epi_funds:
        for c, v in country_by_fund.get(f["ticker"], {}).items():
            epi_country[c] = epi_country.get(c, 0) + v
    epi_sorted = sorted(epi_country.items(), key=lambda x: -x[1])

    gi_country = {}
    for f in gi_funds:
        for c, v in country_by_fund.get(f["ticker"], {}).items():
            gi_country[c] = gi_country.get(c, 0) + v
    gi_sorted = sorted(gi_country.items(), key=lambda x: -x[1])

    # Bullets
    def top_country_bullet(shares, suite_name):
        items = list(shares.items())
        if len(items) < 2: return ""
        (c1, (v1, p1)), (c2, (v2, p2)) = items[0], items[1]
        return f"{c1} ({p1*100:.1f}%) and {c2} ({p2*100:.1f}%) dominate {suite_name} Asia AUM"

    trex_bullets = []
    if trex_shares:
        first = list(trex_shares.items())[0]
        trex_bullets.append(
            f"{first[0]} accounts for {pct1(first[1][1])} of T-REX Asia AUM ({fmt_m(first[1][0])})"
        )
    if trex_top_flow and trex_top_flow.get("flows", 0) > 1e6:
        trex_bullets.append(
            f"<strong>{trex_top_flow['ticker']}</strong> drew {fmt_signed_m(trex_top_flow['flows'])} in est. net inflows — largest T-REX fund flow"
        )
    trex_bullets.append(
        f"T-REX Asia represents {pct1(trex.get('pct_in_asia', 0))} of its global AUM; "
        f"{fmt_signed_m(trex.get('market_move', 0))} from market, {fmt_signed_m(trex.get('flows', 0))} from flows"
    )

    micro_bullets = []
    if ms_shares:
        micro_bullets.append(top_country_bullet(ms_shares, "MicroSectors"))
    if ms_top_mm and ms_top_mm.get("market_move", 0) > 1e6:
        mm = ms_top_mm["market_move"]; gm = ms_top_mm.get("global_mom", 0) or 0
        micro_bullets.append(
            f"<strong>{ms_top_mm['ticker']}</strong> rose on market momentum — {fmt_signed_m(mm)} market move ({gm*100:+.0f}% global MoM)"
        )
    if ms_worst_mm and ms_worst_mm.get("market_move", 0) < -1e6:
        mm = ms_worst_mm["market_move"]; gm = ms_worst_mm.get("global_mom", 0) or 0
        micro_bullets.append(
            f"<strong>{ms_worst_mm['ticker']}</strong> absorbed the sharpest market decline ({fmt_signed_m(mm)}, {gm*100:+.0f}% global MoM)"
        )
    micro_bullets.append(
        f"MicroSectors Asia: {fmt_signed_m(suites.get('MicroSectors',{}).get('market_move',0))} market, "
        f"{fmt_signed_m(suites.get('MicroSectors',{}).get('flows',0))} flows"
    )

    epi_bullets = []
    if epi_sorted:
        (c1, v1) = epi_sorted[0]
        epi_bullets.append(
            f"{c1} leads EPI distribution at {pct1(v1/epi_asia) if epi_asia else '0%'} ({fmt_m(v1)})"
        )
    epi_bullets.append(
        f"EPI suite total: {fmt_m(epi_asia)} across {len(epi_funds)} fund{'s' if len(epi_funds)!=1 else ''}"
    )

    gi_bullets = []
    if gi_sorted:
        (c1, v1) = gi_sorted[0]
        gi_bullets.append(
            f"{c1} dominates G&I at {pct1(v1/gi_asia) if gi_asia else '0%'} ({fmt_m(v1)}) — most geographically concentrated suite"
        )
    gi_bullets.append(
        f"G&I + IncomeMax suite: {fmt_m(gi_asia)} across {len(gi_funds)} funds"
    )

    # Month labels
    from datetime import datetime
    month_dt = datetime.strptime(month_label, "%Y-%m")
    month_long = month_dt.strftime("%B %Y")
    month_short_fmt = month_dt.strftime("%b '%y").replace("'25", "’25").replace("'26", "’26")

    return {
        "month_long": month_long,
        "month_short": month_short_fmt,
        "trex_bullets": trex_bullets,
        "micro_bullets": micro_bullets,
        "epi_bullets": epi_bullets,
        "gi_bullets": gi_bullets,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True, help="YYYY-MM")
    ap.add_argument("--input", default="report_data.json")
    ap.add_argument("--output", default="enriched_report_data.json")
    args = ap.parse_args()

    raw = json.loads(Path(args.input).read_text())
    year, month = map(int, args.month.split("-"))

    print(f"Enriching report for {args.month}...")

    # 1. Bloomberg total REX AUM (reproducible)
    bbg_total, n, missing = bloomberg_total_rex_aum(year, month)
    print(f"  bloomberg_total_rex_aum: ${bbg_total:,.0f} from {n} REX tickers (missing from BBG: {missing})")

    # 2. Per-fund derived
    db_global_total = 0.0
    asia_total = 0.0
    asia_total_prior = 0.0
    total_market_move = 0.0
    total_flows = 0.0
    for f in raw["funds"]:
        asia = float(f["asia_aum"])
        asia_prior = float(f["asia_aum_prior"])
        global_aum = float(f["global_aum"])
        global_prior = float(f["global_aum_prior"])

        mm, fl = compute_market_move_and_flows(asia, asia_prior, global_aum, global_prior)
        f["market_move"] = mm
        f["flows"] = fl
        f["dollar_change"] = asia - asia_prior
        f["mom"] = (asia - asia_prior) / asia_prior if asia_prior > 0 else None
        f["global_mom"] = (global_aum - global_prior) / global_prior if global_prior > 0 else None
        f["pct_of_global"] = (asia / global_aum) if global_aum > 0 else None

        db_global_total += global_aum
        asia_total += asia
        asia_total_prior += asia_prior
        total_market_move += mm
        total_flows += fl

    # 3. Suites
    suites = {}
    for f in raw["funds"]:
        fam = f["family_name"]
        s = suites.setdefault(fam, {
            "aum": 0.0, "aum_prior": 0.0, "global_aum": 0.0, "global_aum_prior": 0.0,
            "flows": 0.0, "market_move": 0.0, "fund_count": 0,
        })
        s["aum"] += float(f["asia_aum"])
        s["aum_prior"] += float(f["asia_aum_prior"])
        s["global_aum"] += float(f["global_aum"])
        s["global_aum_prior"] += float(f["global_aum_prior"])
        s["flows"] += f["flows"]
        s["market_move"] += f["market_move"]
        s["fund_count"] += 1

    for fam, s in suites.items():
        s["dollar_change"] = s["aum"] - s["aum_prior"]
        s["mom"] = s["dollar_change"] / s["aum_prior"] if s["aum_prior"] > 0 else None
        s["pct_in_asia"] = s["aum"] / s["global_aum"] if s["global_aum"] > 0 else None
        s["pct_in_asia_prior"] = s["aum_prior"] / s["global_aum_prior"] if s["global_aum_prior"] > 0 else None
        # Korea share within suite
        s["korea_aum"] = 0.0
    # Korea AUM per suite via fund_countries
    for fc in raw["fund_countries"]:
        if fc["country"] != "Korea": continue
        # Find the fund's family
        for f in raw["funds"]:
            if f["ticker"] == fc["ticker"]:
                suites[f["family_name"]]["korea_aum"] += float(fc["aum"])
                break
    for fam, s in suites.items():
        s["korea_share_pct"] = s["korea_aum"] / s["aum"] if s["aum"] > 0 else 0

    # 4. Headlines
    headlines = {
        "total_asia_aum": asia_total,
        "total_asia_aum_prior": asia_total_prior,
        "dollar_change": asia_total - asia_total_prior,
        "mom_pct": (asia_total - asia_total_prior) / asia_total_prior if asia_total_prior > 0 else None,
        "total_flows": total_flows,
        "total_market_move": total_market_move,
        "total_global_aum": bbg_total,
        "db_global_aum": db_global_total,
        "pct_in_asia": asia_total / bbg_total if bbg_total > 0 else None,
        "fund_count": sum(1 for f in raw["funds"] if float(f["asia_aum"]) > 0),
        "country_count": len([c for c in raw["countries"] if c["aum"] > 0]),
        "exchange_count": len(raw["exchanges"]),
    }

    # 5. Narrative strings (data-driven, no hardcoded text in HTML)
    narrative = build_narrative(headlines, suites, raw["countries"], raw["exchanges"],
                                 raw["funds"], raw["fund_countries"],
                                 args.month, None)

    # 6. Assemble
    raw["meta"]["bloomberg_total_rex_aum"] = bbg_total
    raw["meta"]["bbg_tickers_counted"] = n
    raw["meta"]["bbg_tickers_missing"] = missing
    raw["headlines"] = headlines
    raw["suites"] = suites
    raw["narrative"] = narrative

    Path(args.output).write_text(json.dumps(raw, indent=2, default=str))
    print(f"\nSaved {args.output}")
    print(f"  Asia: ${asia_total/1e6:.2f}M  BBG Total: ${bbg_total/1e9:.2f}B  % in Asia: {headlines['pct_in_asia']*100:.2f}%")
    print(f"  Market move: ${total_market_move/1e6:+.1f}M  Flows: ${total_flows/1e6:+.1f}M")


if __name__ == "__main__":
    main()
