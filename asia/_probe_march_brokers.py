"""Peek at Mar broker file structure to map columns."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from pathlib import Path

FILES = {
    "KSD(Feb)":  "attachments/2026 02 28 Korea REX report.xlsx",
    "KSD":       "grace_data/2026-03/2026 03 31 Korea REX report KSD.xlsx",
    "SBI":       "grace_data/2026-03/2026 03 31 Japan SBI REX report.xlsx",
    "Rakuten":   "grace_data/2026-03/2026 03 31 Japan Rakuten REX report.xlsx",
    "Monex":     "grace_data/2026-03/2026 03 31 Japan Monex REX report.xlsx",
    "Matsui":    "grace_data/2026-03/2026 03 31 Japan Matsui REX report.xlsx",
    "MooMooJP":  "grace_data/2026-03/2026 03 31 MooMoo Jap REX report.xlsx",
}

for label, f in FILES.items():
    print(f"\n=== {label}: {f}")
    wb = openpyxl.load_workbook(f, data_only=True, read_only=False)
    for sheet_name in wb.sheetnames[:3]:
        ws = wb[sheet_name]
        print(f"  [sheet: {sheet_name}]  {ws.max_row}r x {ws.max_column}c")
        # Print first 8 rows, first 10 cols
        for r in range(1, min(9, ws.max_row + 1)):
            vals = []
            for c in range(1, min(11, ws.max_column + 1)):
                v = ws.cell(row=r, column=c).value
                s = str(v) if v is not None else ""
                if len(s) > 20: s = s[:17] + "..."
                vals.append(s)
            print(f"    r{r}: " + " | ".join(vals))
    wb.close()
