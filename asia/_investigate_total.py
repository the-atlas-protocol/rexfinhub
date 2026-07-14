"""Investigate why our Mar 31 total ($5.84B) differs from Seamus's email ($6.0612B).
Hunt for REX tickers in Bloomberg that our DB doesn't track, or different valuation methods.
"""
import openpyxl
from pathlib import Path
import psycopg2
import psycopg2.extras

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")

OUT = open("_total_investigation.txt", "w", encoding="utf-8")
def P(*a, **k): print(*a, **k, file=OUT); OUT.flush()

conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
cur.execute("SELECT ticker, name FROM etp")
db_tickers = {r["ticker"]: r["name"] for r in cur.fetchall()}
conn.close()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

# 1) data_aum headers
ws = wb["data_aum"]
aum_header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
bbg_cols = [c for c in aum_header[1:] if c and " Equity" in str(c)]
P(f"Bloomberg data_aum has {len(bbg_cols)} equity tickers")

# 2) Find Mar 31 row
mar31_row = None
for row in ws.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year') and d.year == 2026 and d.month == 3 and d.day == 31:
        mar31_row = row
        break
P(f"Found Mar 31 row")

# 3) Sum REX tickers (both US and LN variants)
our_rex_tickers = set(db_tickers.keys())
data_aum_mar = {}
for i, col in enumerate(aum_header[1:], 1):
    if not col or " Equity" not in str(col):
        continue
    parts = str(col).split()
    base, listing = parts[0], parts[1]
    key = f"{base}_LN" if listing == "LN" else base
    val = mar31_row[i]
    if val and isinstance(val, (int, float)) and val > 0:
        data_aum_mar[key] = {"col": col, "val_musd": float(val)}

# 4) Microsector overwrite
ws_ms = wb["microsector"]
ms_tickers_row = list(next(ws_ms.iter_rows(min_row=4, max_row=4, values_only=True)))
ms_tickers = [str(t).strip() for t in ms_tickers_row[1:] if t and isinstance(t, str)]
ms_mar = None
for row in ws_ms.iter_rows(min_row=5, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year') and d.year == 2026 and d.month == 3 and d.day == 31:
        ms_mar = row
        break

microsector_mar_raw = {}
if ms_mar:
    for i, t in enumerate(ms_tickers, 1):
        val = ms_mar[i] if i < len(ms_mar) else None
        if val and isinstance(val, (int, float)):
            microsector_mar_raw[t] = float(val)

# 5) Compute totals two ways
P()
P("=== Scenario A: DB REX tickers ONLY (our current approach) ===")
total_a = 0
tickers_a = 0
for t in our_rex_tickers:
    if t in data_aum_mar:
        # Apply microsector overwrite
        if t in microsector_mar_raw:
            val = microsector_mar_raw[t]  # raw $
        else:
            val = data_aum_mar[t]["val_musd"] * 1_000_000  # $M -> $
        total_a += val
        tickers_a += 1
P(f"  Tickers included: {tickers_a}")
P(f"  Total: ${total_a:,.0f} = ${total_a/1e9:.4f}B")
P(f"  Seamus says: $6,061,200,000 ($6.0612B)")
P(f"  Gap: ${(total_a - 6_061_200_000):+,.0f} = ${(total_a - 6_061_200_000)/1e6:+.1f}M")

# 6) Check if any REX-looking tickers exist in Bloomberg that we don't have
P()
P("=== Scenario B: Look for REX tickers in Bloomberg NOT in our DB ===")
# Heuristic: Bloomberg tickers with names matching "REX", "T-REX", "MicroSectors", "FANG", etc.
# Actually just list all Bloomberg tickers with values >0 NOT in our DB
missing_from_db = []
for bbg_key, info in data_aum_mar.items():
    if bbg_key not in our_rex_tickers:
        missing_from_db.append((bbg_key, info["col"], info["val_musd"]))
# Also need to filter these to likely-REX (else we'd have all 947 tickers)
# Fund names in Bloomberg won't be in this sheet. Let me just print top 30 by AUM of missing
missing_from_db.sort(key=lambda x: -x[2])
P(f"  {len(missing_from_db)} Bloomberg tickers with AUM not in our DB (top 30 by AUM shown):")
total_missing = 0
for k, col, v in missing_from_db[:30]:
    P(f"    {col:<30}  ${v:>10,.2f}M")
    total_missing += v * 1e6

# 7) Sum WITHOUT microsector overwrite (using data_aum notional for ETNs)
P()
P("=== Scenario C: DB REX tickers, using data_aum notional (no microsector overwrite) ===")
total_c = 0
tickers_c = 0
for t in our_rex_tickers:
    if t in data_aum_mar:
        val = data_aum_mar[t]["val_musd"] * 1_000_000  # $M
        total_c += val
        tickers_c += 1
P(f"  Tickers: {tickers_c}, total: ${total_c/1e9:.4f}B")
P(f"  Gap vs Seamus: ${(total_c - 6_061_200_000)/1e6:+.1f}M")

# 8) Sum including LN variants
P()
P("=== Scenario D: tickers matching REX names with substring tests ===")
# look for "REX" or "T-REX" or "MicroSectors" in the column names — but col only has ticker, not name
# Actually columns are just "{TICKER} {LISTING} Equity", no company name. So we can't filter by name.
# Print full list of REX tickers that ARE in Bloomberg
P(f"  Our REX tickers: {len(our_rex_tickers)}")
hits_us = [t for t in our_rex_tickers if t in data_aum_mar and not t.endswith("_LN")]
hits_ln = [t for t in our_rex_tickers if t in data_aum_mar and t.endswith("_LN")]
P(f"  Matched in BBG (US): {len(hits_us)}")
P(f"  Matched in BBG (LN): {len(hits_ln)}  -> {hits_ln}")
missing = [t for t in our_rex_tickers if t not in data_aum_mar]
P(f"  NOT in Bloomberg: {len(missing)}  -> {missing}")

# 9) Top 10 REX by Mar 31 AUM from our data
P()
P("=== Top 20 REX by Mar 31 Bloomberg AUM (with microsector overwrite) ===")
rex_aums = []
for t in our_rex_tickers:
    if t in data_aum_mar:
        if t in microsector_mar_raw:
            val = microsector_mar_raw[t]
        else:
            val = data_aum_mar[t]["val_musd"] * 1_000_000
        rex_aums.append((t, val))
rex_aums.sort(key=lambda x: -x[1])
for t, v in rex_aums[:20]:
    P(f"  {t:<10}  ${v/1e6:>10,.2f}M")

wb.close()
OUT.close()
