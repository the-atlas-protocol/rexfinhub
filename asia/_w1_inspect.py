"""Inspect W1 sheet — find the 'last date' column for fund lifecycle."""
import openpyxl
from pathlib import Path
from datetime import date

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")
OUT = open("_w1.txt", "w", encoding="utf-8")
def P(*a, **k): print(*a, **k, file=OUT); OUT.flush()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)
ws = wb["w1"]
P(f"w1 sheet: {ws.max_row}r x {ws.max_column}c")
P()
P("=== Header rows 1-3 ===")
for r in range(1, 4):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            P(f"  r{r}c{c}: {v!r}")

P()
P("=== First 20 data rows (row 2 onwards) ===")
# Try to extract headers from row 1 (or 2) and data from row 2+
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
P(f"Headers (row 1): {headers}")
P()
for r in range(2, min(22, ws.max_row + 1)):
    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    P(f"  r{r}: {row}")

# Look for REX tickers (scan ticker column for startswith matches)
import psycopg2
conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia")
cur = conn.cursor()
cur.execute("SELECT ticker FROM etp")
rex_tickers = set(r[0] for r in cur.fetchall())
conn.close()

P()
P("=== All REX tickers found in W1 (ticker | mkt_status | delist_date) ===")
found = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if not isinstance(v, str): continue
    # W1 ticker format is like "BULZ US" — extract base ticker
    base = v.split()[0] if " " in v else v
    if base in rex_tickers:
        mkt = ws.cell(row=r, column=21).value
        delist = ws.cell(row=r, column=23).value
        fund_name = ws.cell(row=r, column=2).value
        found[base] = {"mkt_status": mkt, "delist_date": delist, "name": fund_name, "row": r, "w1_ticker": v}

for t in sorted(found.keys()):
    f = found[t]
    P(f"  {t:<8}  mkt={f['mkt_status']}   delist={f['delist_date']}   ({f['w1_ticker']})")

P()
P(f"REX tickers NOT in W1: {sorted(rex_tickers - set(found.keys()))}")

P()
P("=== Specifically delisted-by-Ryu funds ===")
TARGETS = {"ETQ", "ARMU", "AXUP", "BKNU", "BULU", "DKUP", "PXIU"}
for t in sorted(TARGETS):
    if t in found:
        f = found[t]
        P(f"  {t:<8}  mkt={f['mkt_status']:<10}  delist={f['delist_date']}")
    else:
        P(f"  {t:<8}  NOT FOUND in W1")

wb.close()
OUT.close()
