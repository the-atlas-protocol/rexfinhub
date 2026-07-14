"""For each month, compare:
  - DB total (from etp_monthly_fund)
  - Bloomberg last-trading-day total (data_aum + microsector overwrite) at proper date
  - Seamus's email total (where we have it)
Also check whether our DB date lines up with Bloomberg's actual last row.
"""
import openpyxl
from pathlib import Path
import psycopg2
import psycopg2.extras
from datetime import date, timedelta

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")

conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
cur.execute("SELECT ticker FROM etp")
rex_tickers = set(r["ticker"] for r in cur.fetchall())
cur.execute("""SELECT cm.month_id, cm.month_end,
                      ROUND(SUM(mf.total_aum_usd)::numeric/1e6, 2) AS db_total_mm
               FROM calendar_month cm LEFT JOIN etp_monthly_fund mf USING (month_id)
               GROUP BY cm.month_id, cm.month_end ORDER BY cm.month_id""")
db_months = cur.fetchall()
conn.close()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

# Pre-index data_aum rows by date
ws = wb["data_aum"]
aum_header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
ticker_to_col = {}
for i, col in enumerate(aum_header[1:], 1):
    if not col or " Equity" not in str(col): continue
    parts = str(col).split()
    key = f"{parts[0]}_LN" if parts[1] == "LN" else parts[0]
    ticker_to_col[key] = i
aum_by_date = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year'):
        aum_by_date[d.date() if hasattr(d, 'date') else d] = row

ws_ms = wb["microsector"]
ms_tickers_row = list(next(ws_ms.iter_rows(min_row=4, max_row=4, values_only=True)))
ms_ticker_cols = {}
for i, t in enumerate(ms_tickers_row[1:], 1):
    if t and isinstance(t, str):
        ms_ticker_cols[t.strip()] = i
ms_by_date = {}
for row in ws_ms.iter_rows(min_row=5, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year'):
        ms_by_date[d.date() if hasattr(d, 'date') else d] = row

def rex_total_for_date(target: date) -> tuple[float, int, date]:
    """Return (total_usd, n_matched, actual_date_used). Walks back to find a row <= target."""
    actual = target
    for _ in range(10):
        if actual in aum_by_date: break
        actual = actual - timedelta(days=1)
    if actual not in aum_by_date:
        return None, 0, None
    aum_row = aum_by_date[actual]
    ms_row = ms_by_date.get(actual)
    total = 0.0; n = 0
    for t in rex_tickers:
        col = ticker_to_col.get(t)
        if col is None: continue
        val = aum_row[col] if col < len(aum_row) else None
        if not (isinstance(val, (int, float)) and val > 0): continue
        if t in ms_ticker_cols and ms_row:
            mi = ms_ticker_cols[t]
            ms_val = ms_row[mi] if mi < len(ms_row) else None
            if isinstance(ms_val, (int, float)) and ms_val > 0:
                total += float(ms_val); n += 1; continue
        total += float(val) * 1_000_000; n += 1
    return total, n, actual

# Seamus's known data points
SEAMUS = {
    date(2026, 2, 27): 6_711_700_000,  # from 02/27 email
    date(2026, 3, 26): 5_821_500_000,
    date(2026, 3, 27): None,
    date(2026, 3, 30): 5_604_400_000,
    date(2026, 3, 31): 6_061_200_000,
}

print(f"{'MID':<4} {'cal_end':<11} {'dow':<4} {'DB':>11} {'BBG date':<11} {'BBG':>11} {'DB-BBG':>10} {'Seamus':>11} {'BBG-Seam':>10}")
print("-" * 100)
for r in db_months:
    mid = r["month_id"]; cal_end = r["month_end"]; db_mm = float(r["db_total_mm"] or 0)
    dow = cal_end.strftime("%a")
    bbg_total, n, actual = rex_total_for_date(cal_end)
    bbg_mm = bbg_total / 1e6 if bbg_total else 0
    db_bbg_gap = db_mm - bbg_mm
    seamus_val = SEAMUS.get(actual)
    seamus_mm = seamus_val / 1e6 if seamus_val else None
    bbg_seam_gap = (bbg_mm - seamus_mm) if seamus_mm else None

    print(f"{mid:<4} {str(cal_end):<11} {dow:<4} ${db_mm:>9.1f}M  {str(actual or '—'):<11} ${bbg_mm:>9.1f}M {db_bbg_gap:>+9.1f}M  "
          + (f"${seamus_mm:>9.1f}M {bbg_seam_gap:>+8.1f}M" if seamus_mm else f"{'—':>11} {'—':>10}"))
wb.close()
