"""
Load ViewTrade Feb 2026 data into rex_asia database.
Source: Grace email Mar 16 — '2026 VT asia datra Feb 2026.xlsx'
ViewTrade covers Taiwan, Hong Kong, Singapore brokers.
Data as of March 9, 2026 but loaded as Feb 2026 (month_id=13).
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import psycopg2
from decimal import Decimal

MONTH_ID = 13
VT_FILE = r"C:\Foundry\Rexfinhub\asia\attachments\2026 VT asia datra Feb 2026.xlsx"

# Ticker normalization: ViewTrade typos → DB tickers
TICKER_MAP = {
    'GOOOX': 'GOOX',
    'AI PI': 'AIPI',
    'AIPII': 'AIPI',
    'RDWUP': 'RDWU',
    'PXII': 'PXIU',
    'XRP': 'XRPR',  # REX XRP ETF
}

# Country normalization
COUNTRY_MAP = {
    'Taiwan': 'Taiwan',
    'Hong Kong': 'Hong Kong',
    'Singapore': 'Singapore',
}


def get_conn():
    return psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia")


def ensure_exchanges(conn):
    """Ensure ViewTrade exchange entries exist for HK and Singapore."""
    cur = conn.cursor()

    # Get country IDs
    cur.execute("SELECT name, country_id FROM country")
    countries = {r[0]: r[1] for r in cur.fetchall()}

    # Check existing ViewTrade exchanges
    cur.execute("SELECT exchange_id, name, country_id FROM exchange WHERE name LIKE 'ViewTrade%'")
    existing = {r[2]: r[0] for r in cur.fetchall()}  # country_id -> exchange_id

    needed = {
        'Hong Kong': countries['Hong Kong'],
        'Singapore': countries['Singapore'],
    }

    for country_name, country_id in needed.items():
        if country_id not in existing:
            cur.execute(
                "INSERT INTO exchange (name, country_id) VALUES (%s, %s) RETURNING exchange_id",
                (f'ViewTrade', country_id)
            )
            eid = cur.fetchone()[0]
            print(f"  Created ViewTrade exchange for {country_name} (id={eid})")
        else:
            print(f"  ViewTrade exchange for {country_name} already exists (id={existing[country_id]})")

    conn.commit()

    # Return full exchange map
    cur.execute("SELECT exchange_id, name, country_id FROM exchange")
    return {(r[1], r[2]): r[0] for r in cur.fetchall()}  # (name, country_id) -> exchange_id


def parse_viewtrade():
    """Parse ViewTrade Excel into rows: {ticker, country, shares, aum_usd}."""
    print("\n=== Parsing ViewTrade data ===")
    wb = openpyxl.load_workbook(VT_FILE, data_only=True)
    ws = wb['March ']

    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        broker = row[0]
        country = str(row[1] or '').strip()
        ticker = str(row[2] or '').strip()
        qty = row[3]
        price = row[4]
        value = row[5]

        # Skip header/empty rows
        if not broker or not country or not ticker:
            continue
        if ticker == 'Symbol' or country not in COUNTRY_MAP:
            continue

        # Normalize ticker
        ticker = TICKER_MAP.get(ticker, ticker)

        # Compute value if formula
        if isinstance(value, str) and value.startswith('='):
            if qty and price and isinstance(qty, (int, float)) and isinstance(price, (int, float)):
                value = qty * price
            else:
                continue
        if not isinstance(value, (int, float)) or value <= 0:
            continue

        shares = float(qty) if isinstance(qty, (int, float)) else 0

        rows.append({
            'ticker': ticker,
            'country': COUNTRY_MAP[country],
            'broker': str(broker).strip(),
            'shares': shares,
            'aum_usd': round(float(value), 4),
        })

    wb.close()

    # Aggregate by ticker × country
    agg = {}
    for r in rows:
        key = (r['ticker'], r['country'])
        if key not in agg:
            agg[key] = {'shares': 0, 'aum_usd': 0}
        agg[key]['shares'] += r['shares']
        agg[key]['aum_usd'] += r['aum_usd']

    # Summary
    by_country = {}
    for (ticker, country), vals in agg.items():
        by_country[country] = by_country.get(country, 0) + vals['aum_usd']

    print(f"  Parsed {len(rows)} position rows -> {len(agg)} ticker-country combinations")
    for c, v in sorted(by_country.items(), key=lambda x: -x[1]):
        print(f"    {c:<15s} ${v/1e6:.2f}M")
    print(f"    {'TOTAL':<15s} ${sum(by_country.values())/1e6:.2f}M")
    print(f"  Grace's summary: $25.4M")

    return agg


def load_viewtrade(agg, exchange_map, conn):
    """Insert aggregated ViewTrade data into fact table."""
    print("\n=== Loading into database ===")
    cur = conn.cursor()

    # Get etp and country maps
    cur.execute("SELECT ticker, etp_id FROM etp")
    etp_map = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT name, country_id FROM country")
    country_ids = {r[0]: r[1] for r in cur.fetchall()}

    inserted = 0
    skipped = []
    for (ticker, country), vals in agg.items():
        etp_id = etp_map.get(ticker)
        if not etp_id:
            skipped.append(f"no etp_id for {ticker}")
            continue

        country_id = country_ids.get(country)
        exchange_id = exchange_map.get(('ViewTrade', country_id))
        if not exchange_id:
            skipped.append(f"no ViewTrade exchange for {country} (country_id={country_id})")
            continue

        cur.execute("""
            INSERT INTO etp_exchange_monthly_aum
                (etp_id, exchange_id, month_id, exchange_aum_usd, shares_outstanding, source_type)
            VALUES (%s, %s, %s, %s, %s, 'reported')
            ON CONFLICT (etp_id, exchange_id, month_id)
            DO UPDATE SET exchange_aum_usd = EXCLUDED.exchange_aum_usd,
                          shares_outstanding = EXCLUDED.shares_outstanding,
                          source_type = 'reported'
        """, (
            etp_id, exchange_id, MONTH_ID,
            Decimal(str(round(vals['aum_usd'], 4))),
            Decimal(str(round(vals['shares'], 6))),
        ))
        inserted += 1

    conn.commit()
    print(f"  Inserted/updated: {inserted} rows")
    if skipped:
        for s in sorted(set(skipped)):
            print(f"  SKIPPED: {s}")


def verify(conn):
    """Show updated totals after ViewTrade load."""
    print("\n=== Verification ===")
    cur = conn.cursor()

    cur.execute("""
        SELECT e.name, c.name as country,
               count(*) as etps,
               round(sum(exa.exchange_aum_usd)/1000000, 2) as aum_mil
        FROM etp_exchange_monthly_aum exa
        JOIN exchange e ON e.exchange_id = exa.exchange_id
        JOIN country c ON c.country_id = e.country_id
        WHERE exa.month_id = %s
        GROUP BY e.name, c.name
        ORDER BY aum_mil DESC
    """, (MONTH_ID,))

    total = 0
    print(f"\n  {'Exchange':<50s} {'Country':>12s} {'ETPs':>5s} {'AUM($M)':>10s}")
    print(f"  {'-'*50} {'-'*12} {'-'*5} {'-'*10}")
    for row in cur.fetchall():
        print(f"  {row[0]:<50s} {row[1]:>12s} {row[2]:>5d} {float(row[3]):>10.2f}")
        total += float(row[3])
    print(f"\n  TOTAL: ${total:.2f}M")

    # ViewTrade subtotal
    cur.execute("""
        SELECT c.name, round(sum(exa.exchange_aum_usd)/1000000, 2)
        FROM etp_exchange_monthly_aum exa
        JOIN exchange e ON e.exchange_id = exa.exchange_id
        JOIN country c ON c.country_id = e.country_id
        WHERE exa.month_id = %s AND e.name = 'ViewTrade'
        GROUP BY c.name
        ORDER BY sum(exa.exchange_aum_usd) DESC
    """, (MONTH_ID,))
    print(f"\n  ViewTrade breakdown:")
    vt_total = 0
    for row in cur.fetchall():
        print(f"    {row[0]:<15s} ${float(row[1]):>8.2f}M")
        vt_total += float(row[1])
    print(f"    {'TOTAL':<15s} ${vt_total:>8.2f}M  (Grace: $25.4M)")


if __name__ == "__main__":
    conn = get_conn()

    # Step 1: Ensure exchange entries
    print("=== Ensuring ViewTrade exchanges ===")
    exchange_map = ensure_exchanges(conn)

    # Step 2: Parse
    agg = parse_viewtrade()

    # Step 3: Load
    load_viewtrade(agg, exchange_map, conn)

    # Step 4: Verify
    verify(conn)

    conn.close()
    print("\nDone.")
