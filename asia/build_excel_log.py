"""
Build REX_Asia_Monthly_Log.xlsx — month-by-month, per-fund, per-vendor log.

V1 scope: produce Feb26 sheet from current DB + Bloomberg, side-by-side with Grace's summary.
Structure: Funds reference | Vendors reference | FX reference | one sheet per month.
"""
import pandas as pd
import psycopg2
import psycopg2.extras
from pathlib import Path
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

BB_FILE = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")
OUT = Path("REX_Asia_Monthly_Log.xlsx")

# ── Styling ──
HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILLS = {
    "IDENTITY":    PatternFill("solid", fgColor="2d3436"),
    "VENDOR":      PatternFill("solid", fgColor="0984e3"),
    "OURS":        PatternFill("solid", fgColor="27ae60"),
    "DERIVED":     PatternFill("solid", fgColor="e67e22"),
    "FLAGS":       PatternFill("solid", fgColor="e74c3c"),
}
FIRST_APPEAR_FILL = PatternFill("solid", fgColor="FFF59D")  # yellow
DELISTED_FILL = PatternFill("solid", fgColor="F5F5F5")  # grey-white
MISMATCH_FILL = PatternFill("solid", fgColor="FFCDD2")  # light red

def get_db():
    return psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                            cursor_factory=psycopg2.extras.RealDictCursor)

# ── Step 1: fetch Bloomberg month-end AUM + price for all REX tickers ──
def fetch_bloomberg_month(month_end: date) -> dict:
    """Return {ticker: {'global_aum_usd': x, 'price_usd': y}} for the given month_end."""
    aum = pd.read_excel(BB_FILE, sheet_name="data_aum", header=0)
    aum = aum.rename(columns={aum.columns[0]: "Date"})
    aum["Date"] = pd.to_datetime(aum["Date"], errors="coerce")
    rows = aum[(aum["Date"].dt.year == month_end.year) & (aum["Date"].dt.month == month_end.month)]
    if rows.empty:
        raise ValueError(f"No Bloomberg AUM data for {month_end}")
    feb_end_aum = rows.iloc[-1]
    feb_aum_date = feb_end_aum["Date"].date()

    price = pd.read_excel(BB_FILE, sheet_name="data_price", header=0)
    price = price.rename(columns={price.columns[0]: "Date"})
    price["Date"] = pd.to_datetime(price["Date"], errors="coerce")
    prows = price[(price["Date"].dt.year == month_end.year) & (price["Date"].dt.month == month_end.month)]
    feb_end_price = prows.iloc[-1] if not prows.empty else None

    # microsector overwrite for ETNs
    ms = pd.read_excel(BB_FILE, sheet_name="microsector", header=None)
    ms_tickers = ms.iloc[3, 1:].tolist()
    ms_data = ms.iloc[4:].copy()
    ms_data.columns = ["Date"] + ms_tickers
    ms_data["Date"] = pd.to_datetime(ms_data["Date"], errors="coerce")
    ms_rows = ms_data[(ms_data["Date"].dt.year == month_end.year) & (ms_data["Date"].dt.month == month_end.month)]
    feb_end_ms = ms_rows.iloc[-1] if not ms_rows.empty else None

    out = {}
    for col in feb_end_aum.index:
        if col == "Date":
            continue
        val = feb_end_aum[col]
        if pd.isna(val):
            continue
        # Column format: "TSLT US Equity" or "FEPI LN Equity"
        parts = str(col).split()
        if len(parts) < 3:
            continue
        ticker, listing, _ = parts[0], parts[1], parts[2]
        key = f"{ticker}_LN" if listing == "LN" else ticker
        out[key] = {
            "global_aum_usd": float(val) * 1_000_000,  # $M -> $
            "price_usd": float(feb_end_price[col]) if feb_end_price is not None and col in feb_end_price.index and pd.notna(feb_end_price[col]) else None,
            "bbg_source": "data_aum",
            "bbg_date": feb_aum_date,
        }
    # Apply microsector overwrite
    if feb_end_ms is not None:
        for t in ms_tickers:
            if t in out and pd.notna(feb_end_ms.get(t)):
                out[t]["global_aum_usd"] = float(feb_end_ms[t])  # raw $
                out[t]["bbg_source"] = "microsector"
    return out

# ── Step 2: fetch DB data for the month ──
def fetch_db_month(month_id: int) -> dict:
    """Returns {(ticker, exchange): aum_usd, ...}, and fund meta."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT e.ticker, e.name AS fund_name, pf.name AS family,
               ex.name AS exchange, c.name AS country,
               m.exchange_aum_usd, m.source_type, m.original_aum_usd
        FROM etp_exchange_monthly_aum m
        JOIN etp e USING (etp_id)
        JOIN product_family pf USING (family_id)
        JOIN exchange ex USING (exchange_id)
        JOIN country c USING (country_id)
        WHERE m.month_id = %s
        ORDER BY e.ticker, ex.name
    """, (month_id,))
    exchange_rows = cur.fetchall()

    cur.execute("""
        SELECT e.ticker, e.name AS fund_name, pf.name AS family,
               COALESCE(m.total_aum_usd, 0) AS db_global_aum, COALESCE(m.price_usd, 0) AS db_price
        FROM etp e
        JOIN product_family pf USING (family_id)
        LEFT JOIN etp_monthly_fund m ON m.etp_id = e.etp_id AND m.month_id = %s
        ORDER BY e.ticker
    """, (month_id,))
    fund_rows = cur.fetchall()

    conn.close()
    return exchange_rows, fund_rows

# ── Step 3: build the month sheet ──
def build_month_sheet(wb: Workbook, month_label: str, month_id: int, month_end: date, prior_month_id: int):
    ws = wb.create_sheet(month_label)

    # Fetch data
    exchange_rows, fund_rows = fetch_db_month(month_id)
    prior_exchange_rows, _ = fetch_db_month(prior_month_id)
    bbg = fetch_bloomberg_month(month_end)

    # Build exchange list (sorted Korea/Japan/HK/SG/MY/TW/TH/Other)
    country_order = {"Korea": 1, "Japan": 2, "Hong Kong": 3, "Singapore": 4, "Malaysia": 5, "Taiwan": 6, "Thailand": 7}
    exchanges_seen = {}
    for r in exchange_rows:
        key = (r["country"], r["exchange"])
        if key not in exchanges_seen:
            exchanges_seen[key] = r["exchange"]
    exchange_list = sorted(exchanges_seen.keys(), key=lambda k: (country_order.get(k[0], 99), k[1]))

    # Prior exchange AUM for first-appearance detection — key by (ticker, country, exchange)
    prior_aum = {}
    for r in prior_exchange_rows:
        prior_aum[(r["ticker"], r["country"], r["exchange"])] = float(r["exchange_aum_usd"])

    # Pivot Feb AUM per (ticker, country, exchange) — MooMoo/ViewTrade share names across countries
    feb_aum = {}
    feb_source = {}
    for r in exchange_rows:
        k = (r["ticker"], r["country"], r["exchange"])
        feb_aum[k] = float(r["exchange_aum_usd"])
        feb_source[k] = r["source_type"]

    # Fund metadata
    fund_meta = {r["ticker"]: r for r in fund_rows}

    # ── Header rows ──
    # Row 1: section banner
    # Row 2: column labels
    headers_row1 = []  # (section_name, span)
    headers_row2 = []

    # IDENTITY (3 cols)
    headers_row1.append(("IDENTITY", 3))
    headers_row2.extend(["Ticker", "Family", "Fund Name"])

    # VENDOR REPORTED: one column per exchange (USD)
    headers_row1.append(("VENDOR REPORTED (USD)", len(exchange_list)))
    for (country, exchange) in exchange_list:
        headers_row2.append(f"{country[:2].upper()} {exchange[:20]}")

    # OUR SIDE (Bloomberg)
    headers_row1.append(("OUR SIDE (BBG)", 3))
    headers_row2.extend(["BBG Global AUM", "BBG Price", "BBG Source"])

    # DERIVED
    headers_row1.append(("DERIVED", 5))
    headers_row2.extend(["Asia Total", "Prior Asia", "MoM $", "MoM %", "% in Asia"])

    # FLAGS
    headers_row1.append(("FLAGS", 2))
    headers_row2.extend(["First Appears", "Notes"])

    # Write row 1 (section banner with merge)
    col = 1
    for (name, span) in headers_row1:
        ws.cell(row=1, column=col, value=name).fill = SECTION_FILLS.get(name.split()[0], HEADER_FILL)
        ws.cell(row=1, column=col).font = HEADER_FONT
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        col += span

    # Row 2: column labels
    for i, h in enumerate(headers_row2, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "D3"

    # ── Data rows ──
    tickers = sorted(fund_meta.keys())
    row = 3
    for ticker in tickers:
        meta = fund_meta[ticker]
        bbg_data = bbg.get(ticker, {})

        col = 1
        ws.cell(row=row, column=col, value=ticker); col += 1
        ws.cell(row=row, column=col, value=meta["family"]); col += 1
        ws.cell(row=row, column=col, value=meta["fund_name"]); col += 1

        # VENDOR columns
        vendor_start = col
        asia_total = 0.0
        for (country, exchange) in exchange_list:
            k = (ticker, country, exchange)
            val = feb_aum.get(k)
            cell = ws.cell(row=row, column=col)
            if val is not None:
                cell.value = val
                cell.number_format = "$#,##0"
                asia_total += val
                # First appearance at (fund, country, exchange) — highlight
                if k not in prior_aum:
                    cell.fill = FIRST_APPEAR_FILL
                # Flag if repriced
                src = feb_source.get(k)
                if src == "repriced":
                    cell.font = Font(italic=True, color="757575")
            col += 1

        # OUR SIDE
        ws.cell(row=row, column=col, value=bbg_data.get("global_aum_usd")).number_format = "$#,##0"; col += 1
        ws.cell(row=row, column=col, value=bbg_data.get("price_usd")).number_format = "$#,##0.00"; col += 1
        ws.cell(row=row, column=col, value=bbg_data.get("bbg_source")); col += 1

        # DERIVED
        prior_total = sum(prior_aum.get((ticker, ctry, ex), 0) for (ctry, ex) in exchange_list)
        mom = asia_total - prior_total
        mom_pct = mom / prior_total if prior_total > 0 else None
        pct_asia = asia_total / bbg_data["global_aum_usd"] if bbg_data.get("global_aum_usd") else None

        ws.cell(row=row, column=col, value=asia_total if asia_total > 0 else None).number_format = "$#,##0"; col += 1
        ws.cell(row=row, column=col, value=prior_total if prior_total > 0 else None).number_format = "$#,##0"; col += 1
        ws.cell(row=row, column=col, value=mom if prior_total > 0 else None).number_format = "$#,##0;[Red]-$#,##0"; col += 1
        ws.cell(row=row, column=col, value=mom_pct).number_format = "0.0%;[Red]-0.0%"; col += 1
        ws.cell(row=row, column=col, value=pct_asia).number_format = "0.0%"; col += 1

        # FLAGS
        first_appears_here = (asia_total > 0 and prior_total == 0)
        ws.cell(row=row, column=col, value="YES" if first_appears_here else "").fill = FIRST_APPEAR_FILL if first_appears_here else PatternFill(); col += 1
        # Notes: flag if Bloomberg AUM missing
        notes = []
        if not bbg_data:
            notes.append("NOT in Bloomberg (delisted?)")
        if asia_total > 0 and not bbg_data.get("global_aum_usd"):
            notes.append("Has Asia AUM but no BBG global AUM")
        ws.cell(row=row, column=col, value="; ".join(notes) if notes else "")

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    for i in range(4, 4 + len(exchange_list)):
        ws.column_dimensions[get_column_letter(i)].width = 14
    for i in range(4 + len(exchange_list), ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    print(f"  Wrote {month_label}: {row - 3} fund rows, {len(exchange_list)} exchanges, {ws.max_column} columns")

# ── Step 4: reference sheets ──
def build_funds_sheet(wb: Workbook, month_id: int):
    """All REX tickers with metadata, sorted by family then ticker."""
    ws = wb.create_sheet("Funds")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT e.ticker, e.name AS fund_name, pf.name AS family,
               CASE WHEN m.etp_id IS NOT NULL THEN 'active' ELSE 'no current data' END AS status,
               COALESCE(m.total_aum_usd, 0) AS latest_global_aum
        FROM etp e
        JOIN product_family pf USING (family_id)
        LEFT JOIN etp_monthly_fund m ON m.etp_id = e.etp_id AND m.month_id = %s
        ORDER BY pf.name, e.ticker
    """, (month_id,))
    rows = cur.fetchall()
    conn.close()

    headers = ["Ticker", "Fund Name", "Family", "Status", "Latest Global AUM"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["ticker"])
        ws.cell(row=i, column=2, value=r["fund_name"])
        ws.cell(row=i, column=3, value=r["family"])
        ws.cell(row=i, column=4, value=r["status"])
        ws.cell(row=i, column=5, value=float(r["latest_global_aum"])).number_format = "$#,##0"
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    print(f"  Wrote Funds: {len(rows)} rows")

def build_vendors_sheet(wb: Workbook):
    ws = wb.create_sheet("Vendors")
    vendors = [
        # name, country, cadence, native_ccy, format, status, notes
        ("KSD Retail",           "Korea",    "monthly",   "USD",    "USD per fund",       "active",  "Aggregate + ACE ETF side table"),
        ("KIM ACE Tesla Fund",   "Korea",    "monthly",   "KRW",    "KRW + FX -> USD",    "active",  "Single fund, from Investor webpage"),
        ("SBI",                  "Japan",    "monthly",   "JPY",    "USD (also JPY)",     "active",  ""),
        ("Rakuten",              "Japan",    "monthly",   "JPY",    "USD",                "active",  ""),
        ("Monex",                "Japan",    "monthly",   "JPY",    "JPY + USD + FX",     "active",  "Clean, reproducible"),
        ("Matsui",               "Japan",    "monthly",   "JPY",    "USD",                "active",  ""),
        ("MooMoo Japan",         "Japan",    "monthly",   "USD",    "shares + USD",       "active",  "Upgraded from quarterly Mar26"),
        ("Futu/MooMoo HK",       "Hong Kong","quarterly", "USD",    "USD",                "FROZEN",  "Cannot send without commercial contract (Apr26)"),
        ("Oriental Harbour",     "Hong Kong","quarterly", "USD",    "13F filing",         "active",  "Q-end + 13F arrives mid-quarter-end+1"),
        ("SYFE",                 "Hong Kong","monthly",   "USD",    "verbal",             "active",  "No documentation; value is FEPI_LN only"),
        ("ViewTrade HK/SG/TW",   "Various",  "quarterly", "USD",    "shares + price + USD","active", "Physical hardcopy collection required"),
        ("MooMoo Singapore",     "Singapore","quarterly", "USD",    "USD",                "active",  "Grace waiting for Q1 data"),
        ("MooMoo Malaysia",      "Malaysia", "quarterly", "USD",    "USD",                "active",  "Grace waiting for Q1 data"),
        ("Asset Plus Thailand",  "Thailand", "monthly",   "USD",    "USD",                "ZERO",    "Sold DRNZ shares Apr26 -- now $0"),
    ]
    headers = ["Vendor", "Country", "Cadence", "Native CCY", "Format", "Status", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
    for i, v in enumerate(vendors, start=2):
        for j, val in enumerate(v, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if j == 6 and val in ("FROZEN", "ZERO"):
                cell.fill = MISMATCH_FILL
                cell.font = Font(bold=True, color="C62828")
    ws.freeze_panes = "A2"
    for col, w in enumerate([22, 12, 11, 12, 22, 10, 50], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    print(f"  Wrote Vendors: {len(vendors)} rows")

def build_overview_sheet(wb: Workbook):
    ws = wb.create_sheet("README", 0)
    content = [
        ("REX Asia Monthly Log", True),
        (f"Generated: {date.today().isoformat()}", False),
        ("", False),
        ("Purpose: per-fund per-month per-vendor log of Asia AUM reporting.", False),
        ("This workbook is the SANITY CHECK against the DB and report pipeline.", False),
        ("", False),
        ("Sheets:", True),
        ("  Funds   -- master list of all REX tickers (active + historical)", False),
        ("  Vendors -- broker list with cadence / format / status", False),
        ("  YYYY-MM -- one sheet per reporting month. Wide layout.", False),
        ("", False),
        ("Month sheet sections:", True),
        ("  IDENTITY  -- ticker, family, name", False),
        ("  VENDOR    -- each exchange as a column; USD values as reported/derived", False),
        ("  OUR SIDE  -- Bloomberg global AUM + price for our end", False),
        ("  DERIVED   -- Asia total, MoM, % in Asia", False),
        ("  FLAGS     -- first appearance (yellow), notes, anomalies", False),
        ("", False),
        ("Highlighting:", True),
        ("  YELLOW  -- first appearance of a fund at an exchange (expansion)", False),
        ("  ITALIC  -- value is REPRICED (quarterly reporter, data not fresh)", False),
        ("  RED     -- mismatch or frozen / zero status", False),
    ]
    for i, (text, bold) in enumerate(content, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 100

def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # README first
    build_overview_sheet(wb)

    # Reference sheets
    print("Building reference sheets...")
    build_funds_sheet(wb, month_id=13)
    build_vendors_sheet(wb)

    # Month sheets — backfill from earliest month forward
    print("Building month sheets...")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT month_id, month_end FROM calendar_month ORDER BY month_id")
    months = cur.fetchall()
    conn.close()

    for m in months:
        month_id = m["month_id"]
        month_end = m["month_end"]
        label = month_end.strftime("%Y-%m")
        prior_id = month_id - 1 if month_id > 1 else None
        if prior_id is None:
            continue  # skip month 1 (no prior)
        build_month_sheet(wb, label, month_id, month_end, prior_id)

    wb.save(OUT)
    print(f"\nSaved: {OUT.absolute()}  ({OUT.stat().st_size / 1024:.0f} KB)")

if __name__ == "__main__":
    main()
