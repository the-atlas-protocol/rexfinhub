"""Check NAV values vs Price across day windows around each month-end.
If NAV was day-shifted by mistake, the NAV/Price ratio will jump abruptly.
"""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
from datetime import date, timedelta

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")

OUT = open("_nav_date_check.txt", "w", encoding="utf-8")
def L(*a): print(*a); print(*a, file=OUT); OUT.flush()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

def index(sheet):
    ws = wb[sheet]
    h = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cols = {c: i for i, c in enumerate(h) if c and " Equity" in str(c)}
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d and hasattr(d, 'year'):
            rows[d.date() if hasattr(d, 'date') else d] = row
    return cols, rows

p_cols, p_rows = index("data_price")
n_cols, n_rows = index("data_nav")

def get(rows, cols, ticker, dt):
    col = f"{ticker} US Equity"
    if col not in cols: return None
    row = rows.get(dt)
    if not row: return None
    v = row[cols[col]]
    return float(v) if isinstance(v, (int, float)) and v > 0 else None

# Check 9 key tickers
TICKERS = ["TSLT", "MSTU", "BMNU", "NVDX", "FEPI", "AIPI", "GOOX", "ROBN", "BULZ", "GDXU", "FNGU"]

# Month-end transition windows for last 6 months
WINDOWS = [
    ("Oct→Nov 2025", [date(2025,10,28), date(2025,10,29), date(2025,10,30), date(2025,10,31), date(2025,11,3), date(2025,11,4)]),
    ("Nov→Dec 2025", [date(2025,11,25), date(2025,11,26), date(2025,11,27), date(2025,11,28), date(2025,12,1), date(2025,12,2)]),
    ("Dec25→Jan26", [date(2025,12,29), date(2025,12,30), date(2025,12,31), date(2026,1,1), date(2026,1,2), date(2026,1,5)]),
    ("Jan→Feb 2026", [date(2026,1,28), date(2026,1,29), date(2026,1,30), date(2026,2,2), date(2026,2,3)]),
    ("Feb→Mar 2026", [date(2026,2,25), date(2026,2,26), date(2026,2,27), date(2026,3,2), date(2026,3,3)]),
    ("Mar→Apr 2026", [date(2026,3,27), date(2026,3,30), date(2026,3,31), date(2026,4,1), date(2026,4,2)]),
]

for window_name, dates in WINDOWS:
    L(f"\n{'='*100}")
    L(f"WINDOW: {window_name}")
    L(f"{'='*100}")
    L(f"{'Ticker':<8}  " + "  ".join(f"{d.strftime('%a %m/%d'):<11}" for d in dates))
    L(f"{'  NAV':<8}")
    for t in TICKERS:
        navs = []
        for d in dates:
            v = get(n_rows, n_cols, t, d)
            navs.append(f"{v:>10.4f}" if v else f"{'—':>10}")
        L(f"  {t:<6}  " + "  ".join(f"{x:<11}" for x in navs))
    L(f"  Price")
    for t in TICKERS:
        prices = []
        for d in dates:
            v = get(p_rows, p_cols, t, d)
            prices.append(f"{v:>10.4f}" if v else f"{'—':>10}")
        L(f"  {t:<6}  " + "  ".join(f"{x:<11}" for x in prices))
    L(f"  NAV/Price ratio")
    for t in TICKERS:
        ratios = []
        for d in dates:
            n = get(n_rows, n_cols, t, d)
            p = get(p_rows, p_cols, t, d)
            if n and p:
                ratios.append(f"{n/p:>10.4f}")
            else:
                ratios.append(f"{'—':>10}")
        L(f"  {t:<6}  " + "  ".join(f"{x:<11}" for x in ratios))

# Now identify suspicious patterns:
# - Same NAV value on two consecutive trading days = potentially shifted/duplicated
# - NAV/Price ratio drifting smoothly is normal; sudden jump >2x baseline = suspicious
L(f"\n\n{'='*100}")
L("SUSPICIOUS PATTERNS — same NAV on consecutive trading days, or ratio jumps")
L(f"{'='*100}")

ALL_DATES = sorted(set(d for w in WINDOWS for d in w[1]))
for t in TICKERS:
    flags = []
    prev_d = None; prev_nav = None
    for d in ALL_DATES:
        nav = get(n_rows, n_cols, t, d)
        if nav is None: continue
        if prev_nav is not None and abs(nav - prev_nav) < 0.0001 and (d - prev_d).days <= 3:
            flags.append(f"DUPLICATE: {prev_d} and {d} both = {nav:.4f}")
        prev_d = d; prev_nav = nav
    if flags:
        L(f"\n  {t}:")
        for f in flags:
            L(f"    {f}")

# Specific scan: Mar 31 NAV value — does it match Mar 30 or Mar 31 economic value?
L(f"\n\n{'='*100}")
L(f"MAR 31 SPOT CHECK — Is Mar 31 NAV the actual Mar 31 close, or a copy of Mar 30?")
L(f"{'='*100}")
L(f"{'Ticker':<8} {'Mar27 NAV':>11} {'Mar30 NAV':>11} {'Mar31 NAV':>11} {'Apr1 NAV':>11} {'Mar31 Price':>13}")
for t in TICKERS:
    n_27 = get(n_rows, n_cols, t, date(2026,3,27))
    n_30 = get(n_rows, n_cols, t, date(2026,3,30))
    n_31 = get(n_rows, n_cols, t, date(2026,3,31))
    n_a1 = get(n_rows, n_cols, t, date(2026,4,1))
    p_31 = get(p_rows, p_cols, t, date(2026,3,31))
    fmt = lambda x: f"{x:>10.4f}" if x else f"{'—':>10}"
    L(f"  {t:<6} {fmt(n_27):>11} {fmt(n_30):>11} {fmt(n_31):>11} {fmt(n_a1):>11} {fmt(p_31):>13}")
    # Detect: if Mar31 NAV == Mar30 NAV exactly, suggests no Mar31 update
    if n_30 and n_31 and abs(n_30 - n_31) < 0.0001:
        L(f"     SUSPECT: Mar30 NAV == Mar31 NAV exactly")

wb.close(); OUT.close()
