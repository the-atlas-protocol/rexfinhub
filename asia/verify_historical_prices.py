"""
Split-aware historical price verification — compare yfinance closes to BBG data_nav
across all month-end dates for the REX universe.

yfinance auto-adjusts for splits (default). To compare apples-to-apples with BBG's
date-of-event prices, we un-adjust yfinance closes by walking forward through splits.

Usage:
    python verify_historical_prices.py
Output:
    _historical_verify.txt — flagged tickers with month-end gaps >2%
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
from datetime import date, timedelta
from collections import defaultdict

import openpyxl
import yfinance as yf
import psycopg2
import psycopg2.extras

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")
OUT = open("_historical_verify.txt", "w", encoding="utf-8")
def L(*a): print(*a); print(*a, file=OUT); OUT.flush()

# ── Index BBG sheets ───────────────────────────────────────────────
wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

def index(sheet):
    ws = wb[sheet]
    h = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cols = {c: i for i, c in enumerate(h) if c and " Equity" in str(c)}
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d and hasattr(d, 'year'):
            rows[d.date() if hasattr(d, 'date') else d] = row
    return cols, rows

p_cols, p_rows = index("data_price")
n_cols, n_rows = index("data_nav")

def get(rows, cols, ticker, dt):
    col = f"{ticker} US Equity"
    if col not in cols: return None
    row = rows.get(dt)
    v = row[cols[col]] if row else None
    return float(v) if isinstance(v, (int, float)) and v > 0 else None

# DB tickers
conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
cur.execute("SELECT ticker FROM etp")
tickers = [r["ticker"] for r in cur.fetchall()]

# Month-end dates we care about
MONTH_ENDS = [date(2025,2,28), date(2025,3,31), date(2025,4,30), date(2025,5,30),  # Sat→Fri
              date(2025,6,30), date(2025,7,31), date(2025,8,29),  # Sun→Fri
              date(2025,9,30), date(2025,10,31), date(2025,11,28), date(2025,12,31),
              date(2026,1,30), date(2026,2,27), date(2026,3,31)]

# ── Pull yfinance with splits info ─────────────────────────────────
L(f"Pulling yfinance history + splits for {len(tickers)} tickers ...")
yf_close = defaultdict(dict)
splits = {}
for t in tickers:
    try:
        ticker_obj = yf.Ticker(t)
        hist = ticker_obj.history(start="2025-01-01", end="2026-05-05", auto_adjust=False)
        if hist.empty: continue
        for ts, row in hist.iterrows():
            d = ts.date()
            close = float(row["Close"]) if row["Close"] > 0 else None
            if close:
                yf_close[t][d] = close
        # Get splits
        actions = ticker_obj.actions
        if actions is not None and not actions.empty:
            ticker_splits = []
            for ts, row in actions.iterrows():
                if row.get("Stock Splits", 0) and row["Stock Splits"] != 0:
                    ticker_splits.append((ts.date(), float(row["Stock Splits"])))
            if ticker_splits:
                splits[t] = ticker_splits
    except Exception as e:
        L(f"  yfinance error for {t}: {e}")

L(f"  yfinance pulled {len(yf_close)} tickers; {len(splits)} have splits")

def unadjust_close(t, raw_close, target_date):
    """yfinance auto-adjusts Close for FUTURE splits relative to today.
    To get price as-of target_date (raw), apply forward splits that happened AFTER target_date.
    For each split (date, factor): if split_date > target_date, multiply by factor (factor < 1 means reverse-split, raises raw price).
    """
    if t not in splits: return raw_close
    out = raw_close
    for split_date, factor in splits[t]:
        if split_date > target_date:
            # Future split (relative to target). yfinance close is post-split basis; un-adjust by multiplying inverse.
            # Reverse split factor = 0.04 means 1 share became 0.04 shares (price ×25 to maintain market cap).
            # yfinance shows post-split high price; raw historical low price = post-split / 25 = post-split × factor
            out = out * factor
    return out

# ── Compare per-month-end ──────────────────────────────────────────
L(f"\n{'='*100}")
L(f"MONTH-END NAV vs yfinance (split-adjusted to as-of date)")
L(f"{'='*100}")

issues = defaultdict(list)  # month_end -> list of (ticker, nav, yf_unadj, gap)
checked = defaultdict(int)
for me in MONTH_ENDS:
    for t in tickers:
        nav = get(n_rows, n_cols, t, me)
        # Walk back yfinance to last available trading day ≤ me
        yf_raw = None; yf_date = None
        cur_d = me
        for _ in range(7):
            if cur_d in yf_close.get(t, {}):
                yf_raw = yf_close[t][cur_d]; yf_date = cur_d; break
            cur_d -= timedelta(days=1)
        if not (nav and yf_raw): continue
        yf_unadj = unadjust_close(t, yf_raw, me)
        gap = (nav - yf_unadj) / yf_unadj
        checked[me] += 1
        if abs(gap) > 0.02:  # 2% threshold
            issues[me].append((t, nav, yf_unadj, gap, yf_date))

# Report
L(f"\n{'Month-end':<12} {'#checked':>10} {'#issues':>10} {'sample issues (top 5)':<60}")
for me in MONTH_ENDS:
    iss = issues[me]
    sample = ", ".join(f"{t}:{gap*100:+.1f}%" for t, _, _, gap, _ in sorted(iss, key=lambda x: -abs(x[3]))[:5])
    L(f"  {str(me):<11} {checked[me]:>10} {len(iss):>10}    {sample}")

# Per-ticker hot list (consistent issues across multiple month-ends)
L(f"\n{'='*100}")
L(f"TICKERS WITH ISSUES ON MULTIPLE MONTH-ENDS (NAV vs yf-unadj >2%)")
L(f"{'='*100}")
ticker_issue_count = defaultdict(int)
ticker_examples = defaultdict(list)
for me, iss in issues.items():
    for t, nav, yf_u, gap, _ in iss:
        ticker_issue_count[t] += 1
        ticker_examples[t].append((me, nav, yf_u, gap))

for t in sorted(ticker_issue_count.keys(), key=lambda x: -ticker_issue_count[x]):
    if ticker_issue_count[t] >= 2:
        L(f"\n  {t}: {ticker_issue_count[t]} month-ends with >2% gap")
        for me, nav, yf_u, gap in sorted(ticker_examples[t], key=lambda x: x[0])[:5]:
            L(f"    {str(me)}: NAV={nav:.4f} yf-unadj={yf_u:.4f} gap={gap*100:+.2f}%")

# Special: Mar 31 2026 specifically — Price vs yf-unadj
L(f"\n{'='*100}")
L(f"MAR 31 2026 — BBG data_price vs yfinance (split-adjusted) — confirms data_price corruption")
L(f"{'='*100}")
me = date(2026,3,31)
price_issues = []
for t in tickers:
    price = get(p_rows, p_cols, t, me)
    nav = get(n_rows, n_cols, t, me)
    yf_raw = yf_close.get(t, {}).get(me)
    if not yf_raw:
        # Walk back
        cur_d = me
        for _ in range(7):
            if cur_d in yf_close.get(t, {}):
                yf_raw = yf_close[t][cur_d]; break
            cur_d -= timedelta(days=1)
    if not (price and yf_raw): continue
    yf_unadj = unadjust_close(t, yf_raw, me)
    p_gap = (price - yf_unadj) / yf_unadj
    n_gap = (nav - yf_unadj) / yf_unadj if nav else None
    if abs(p_gap) > 0.005:  # 0.5% threshold for Mar 31
        price_issues.append((t, price, nav, yf_unadj, p_gap, n_gap))

L(f"\nFinal corrupt-Price ticker list for Mar 31: {len(price_issues)}")
L(f"{'Ticker':<8} {'BBG Price':>10} {'BBG NAV':>10} {'yf-unadj':>10} {'Pgap%':>8} {'Ngap%':>8}")
for t, p, n, y, pg, ng in sorted(price_issues, key=lambda x: -abs(x[4]))[:30]:
    ng_str = f"{ng*100:+.2f}" if ng is not None else "—"
    L(f"  {t:<8} {p:>10.4f} {n if n else '—':>10} {y:>10.4f} {pg*100:>+7.2f}% {ng_str:>7}%")

wb.close(); conn.close(); OUT.close()
