"""
Load Feb 2026 data into rex_asia database.
Steps:
  1. Load etp_monthly_fund (price + total AUM) from bloomberg_daily_file
  2. Parse all broker Excel files into staging rows
  3. Validate against Grace's summary
  4. Insert into etp_exchange_monthly_aum
  5. Carry forward quarterly vendors
  6. Verify views
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import psycopg2
from datetime import date, datetime
from decimal import Decimal

MONTH_ID = 13
MONTH_END = date(2026, 2, 28)
BBG_PATH = r"C:\Users\RyuEl-Asmar\REX Financial LLC\REX Financial LLC - MasterFiles\MASTER Data\bloomberg_daily_file.xlsm"
ATTACH_DIR = r"C:\Projects\rexfinhub\asia\attachments"

def get_conn():
    return psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia")

def get_etp_map(conn):
    """ticker -> etp_id"""
    cur = conn.cursor()
    cur.execute("SELECT ticker, etp_id FROM etp")
    return {row[0]: row[1] for row in cur.fetchall()}

def get_exchange_map(conn):
    """(exchange_name) -> exchange_id"""
    cur = conn.cursor()
    cur.execute("SELECT name, exchange_id FROM exchange")
    return {row[0]: row[1] for row in cur.fetchall()}

# ─── STEP 1: Load etp_monthly_fund ────────────────────────────────────────────

def load_etp_monthly_fund(conn):
    """Load price_usd from data_price and total_aum_usd from data_aum for Feb 28."""
    print("\n=== STEP 1: Loading etp_monthly_fund ===")
    wb = openpyxl.load_workbook(BBG_PATH, data_only=True, read_only=False)

    # --- Read prices from data_price ---
    ws_price = wb['data_price']
    price_header = [c.value for c in ws_price[1]]
    prices = {}  # ticker -> price

    for row in ws_price.iter_rows(min_row=2, max_row=4000):
        d = row[0].value
        if d and hasattr(d, 'month') and d.year == 2026 and d.month == 2 and d.day == 28:
            for i, col in enumerate(price_header[1:], 1):
                if col:
                    raw = str(col)
                    # Skip non-US listings (LN Equity, etc.)
                    if ' LN Equity' in raw:
                        continue
                    ticker = raw.replace(' US Equity', '')
                    val = row[i].value
                    if val and isinstance(val, (int, float)) and val > 0:
                        prices[ticker] = float(val)
            break

    # --- Read AUM from data_aum ---
    ws_aum = wb['data_aum']
    aum_header = list(next(ws_aum.iter_rows(min_row=1, max_row=1, values_only=True)))
    aums = {}  # ticker -> aum in millions

    for row in ws_aum.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d and hasattr(d, 'month') and d.year == 2026 and d.month == 2 and d.day == 28:
            for i, col in enumerate(aum_header[1:], 1):
                if col:
                    raw = str(col)
                    # Skip non-US listings (LN Equity, etc.)
                    if ' LN Equity' in raw:
                        continue
                    ticker = raw.replace(' US Equity', '')
                    val = row[i]
                    if val and isinstance(val, (int, float)) and val > 0:
                        aums[ticker] = float(val) * 1_000_000  # convert M to raw
            break

    # --- Fallback: read current AUM from w4 for any missing tickers ---
    ws_w4 = wb['w4']
    for row in ws_w4.iter_rows(min_row=2, max_row=9000, values_only=True):
        ticker_raw = str(row[0] or '').strip()
        ticker = ticker_raw.replace(' US', '').replace(' LN', '')
        if ticker not in aums and row[10] and isinstance(row[10], (int, float)) and row[10] > 0:
            aums[ticker] = float(row[10]) * 1_000_000  # w4 AUM is in millions

    wb.close()

    # --- Insert into DB ---
    etp_map = get_etp_map(conn)
    cur = conn.cursor()
    inserted = 0
    missing_price = []
    missing_aum = []

    for ticker, etp_id in etp_map.items():
        price = prices.get(ticker)
        aum = aums.get(ticker)

        if not price:
            missing_price.append(ticker)
            continue
        if not aum:
            missing_aum.append(ticker)
            # Use price * 0 as placeholder? No - skip for now
            continue

        cur.execute("""
            INSERT INTO etp_monthly_fund (etp_id, month_id, price_usd, total_aum_usd)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (etp_id, month_id) DO NOTHING
        """, (etp_id, MONTH_ID, Decimal(str(round(price, 6))), Decimal(str(round(aum, 4)))))
        inserted += 1

    conn.commit()
    print(f"  Inserted: {inserted} ETPs")
    if missing_price:
        print(f"  Missing price ({len(missing_price)}): {sorted(missing_price)}")
    if missing_aum:
        print(f"  Missing AUM ({len(missing_aum)}): {sorted(missing_aum)}")

# ─── STEP 2: Parse broker files ───────────────────────────────────────────────

def parse_korea_ksd(conn):
    """Korea KSD: hold_amt is already USD value. Derive shares = aum / price."""
    print("\n--- Korea KSD ---")
    wb = openpyxl.load_workbook(os.path.join(ATTACH_DIR, "2026 02 28 Korea REX report.xlsx"), data_only=True)
    ws = wb.active

    # Get prices to derive share counts
    cur = conn.cursor()
    cur.execute("SELECT e.ticker, emf.price_usd FROM etp_monthly_fund emf JOIN etp e ON e.etp_id = emf.etp_id WHERE emf.month_id = %s", (MONTH_ID,))
    prices = {row[0]: float(row[1]) for row in cur.fetchall()}

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        ticker = row[0]
        hold_amt = row[2]  # hold_amt = USD dollar value (NOT shares)
        if not ticker or not isinstance(ticker, str) or not ticker.isupper() or len(ticker) > 6:
            continue
        if not hold_amt or not isinstance(hold_amt, (int, float)):
            continue

        aum_usd = float(hold_amt)
        price = prices.get(ticker, 0)
        shares = aum_usd / price if price > 0 else 0

        rows.append({
            'ticker': ticker,
            'exchange': 'KSD (Korea Securities Depository) - Retail',
            'aum_usd': round(aum_usd, 4),
            'shares': round(shares, 6),
        })

    wb.close()
    total = sum(r['aum_usd'] for r in rows)
    print(f"  Parsed: {len(rows)} tickers, total AUM: ${total/1e6:.1f}M")
    return rows

def parse_korea_ace(conn):
    """Korea ACE TESLA Value Chain: single entry from Grace's summary."""
    print("\n--- Korea ACE TESLA ---")
    # From Grace's spreadsheet: $23.9M, and it's all TSLT position
    # The Korea report also has this: KRW 1,257.5B / FX 1,450 = $867M, 2.76% to TSLT = $23.94M
    wb = openpyxl.load_workbook(os.path.join(ATTACH_DIR, "2026 02 28 Korea REX report.xlsx"), data_only=True)
    ws = wb.active

    # Read the ACE ETF calculation from the side columns (row 2, cols J-N)
    ace_aum_usd = None
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        # Col J (idx 9) = KRW AUM, Col K (idx 10) = FX rate, Col L (idx 11) = USD AUM
        # Col M (idx 12) = allocation %, Col N (idx 13) = TSLT position
        tslt_position = row[13]
        if tslt_position and isinstance(tslt_position, (int, float)):
            ace_aum_usd = float(tslt_position)
    wb.close()

    if not ace_aum_usd:
        ace_aum_usd = 23_900_000  # fallback from Grace's summary

    # Get TSLT price for shares calculation
    cur = conn.cursor()
    cur.execute("SELECT emf.price_usd FROM etp_monthly_fund emf JOIN etp e ON e.etp_id = emf.etp_id WHERE e.ticker = 'TSLT' AND emf.month_id = %s", (MONTH_ID,))
    tslt_price = float(cur.fetchone()[0])
    shares = ace_aum_usd / tslt_price

    print(f"  ACE TSLT position: ${ace_aum_usd/1e6:.1f}M ({shares:.0f} shares @ ${tslt_price})")
    return [{
        'ticker': 'TSLT',
        'exchange': 'Korea Investment Management -ACE TESLA Value Chain ETF',
        'aum_usd': round(ace_aum_usd, 4),
        'shares': round(shares, 6),
    }]

def parse_japan_rakuten():
    """Rakuten: USD values in column 5 (Market Value USD)."""
    print("\n--- Japan Rakuten ---")
    wb = openpyxl.load_workbook(os.path.join(ATTACH_DIR, "2026 02 28 Japan Rakuten REX report.xlsx"), data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        ticker = row[0]
        shares = row[3]      # 数量 (quantity)
        usd_val = row[4]     # 時価評価額(US$)

        if not ticker or not isinstance(ticker, str) or not ticker.isupper():
            continue
        if not usd_val or not isinstance(usd_val, (int, float)):
            continue

        rows.append({
            'ticker': ticker.strip(),
            'exchange': 'Rakuten',
            'aum_usd': round(float(usd_val), 4),
            'shares': round(float(shares), 6) if shares else 0,
        })

    wb.close()
    total = sum(r['aum_usd'] for r in rows)
    print(f"  Parsed: {len(rows)} tickers, total AUM: ${total/1e6:.1f}M")
    return rows

def parse_japan_sbi():
    """SBI: USD values in column 4 ($ value)."""
    print("\n--- Japan SBI ---")
    wb = openpyxl.load_workbook(os.path.join(ATTACH_DIR, "2026 02 28 Japan SBI REX report.xlsx"), data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        ticker = row[0]
        shares = row[1]
        usd_val = row[3]  # $ value

        if not ticker or not isinstance(ticker, str) or not ticker.isupper():
            continue
        if not usd_val or not isinstance(usd_val, (int, float)):
            continue

        rows.append({
            'ticker': ticker.strip(),
            'exchange': 'SBI',
            'aum_usd': round(float(usd_val), 4),
            'shares': round(float(shares), 6) if shares else 0,
        })

    wb.close()
    total = sum(r['aum_usd'] for r in rows)
    print(f"  Parsed: {len(rows)} tickers, total AUM: ${total/1e6:.1f}M")
    return rows

def parse_japan_monex():
    """Monex: USD values in column 4 (Asset Balances USD)."""
    print("\n--- Japan Monex ---")
    wb = openpyxl.load_workbook(os.path.join(ATTACH_DIR, "02272026_Monex_AssetBalances.xlsx"), data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        ticker = row[0]
        usd_val = row[3]  # Asset Balances (USD)

        if not ticker or not isinstance(ticker, str) or not ticker.isupper():
            continue
        if not usd_val or not isinstance(usd_val, (int, float)):
            continue

        # Derive shares from JPY balance / USD balance * ... or just use AUM / price
        rows.append({
            'ticker': ticker.strip(),
            'exchange': 'Monex',
            'aum_usd': round(float(usd_val), 4),
            'shares': 0,  # will compute from price later
        })

    wb.close()
    total = sum(r['aum_usd'] for r in rows)
    print(f"  Parsed: {len(rows)} tickers, total AUM: ${total/1e6:.1f}M")
    return rows

def parse_japan_matsui():
    """Matsui: USD values in column 3 (Stock Balance $)."""
    print("\n--- Japan Matsui ---")
    wb = openpyxl.load_workbook(os.path.join(ATTACH_DIR, "Feb_Matsui_AssetBlalances.xlsx"), data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        ticker = row[0]
        shares = row[1]
        usd_val = row[2]  # Stock Balance ($)

        if not ticker or not isinstance(ticker, str) or not ticker.isupper():
            continue
        if not usd_val or not isinstance(usd_val, (int, float)):
            continue

        rows.append({
            'ticker': ticker.strip(),
            'exchange': 'Matsui',
            'aum_usd': round(float(usd_val), 4),
            'shares': round(float(shares), 6) if shares else 0,
        })

    wb.close()
    total = sum(r['aum_usd'] for r in rows)
    print(f"  Parsed: {len(rows)} tickers, total AUM: ${total/1e6:.1f}M")
    return rows

def parse_thailand(conn):
    """Thailand Asset Plus: DRNZ, 16000 shares."""
    print("\n--- Thailand Asset Plus ---")
    cur = conn.cursor()
    cur.execute("SELECT emf.price_usd FROM etp_monthly_fund emf JOIN etp e ON e.etp_id = emf.etp_id WHERE e.ticker = 'DRNZ' AND emf.month_id = %s", (MONTH_ID,))
    drnz_price = float(cur.fetchone()[0])

    shares = 16000
    aum = shares * drnz_price
    print(f"  DRNZ: 16,000 shares × ${drnz_price} = ${aum/1e6:.3f}M")

    return [{
        'ticker': 'DRNZ',
        'exchange': 'Asset Plus Asset Management',
        'aum_usd': round(aum, 4),
        'shares': 16000,
    }]

def parse_hk_syfe():
    """SYFE HK: verbal report, $1M on FEPILN. Mapped to FEPI."""
    print("\n--- HK SYFE ---")
    # Grace says "$FEPILN (Verbal report only)" = $1M
    # FEPILN is FEPI listed in London. Map to FEPI ticker.
    aum = 1_000_000
    print(f"  FEPI (via FEPILN): ${aum/1e6:.1f}M (verbal)")
    return [{
        'ticker': 'FEPI',
        'exchange': 'SYFE',
        'aum_usd': round(aum, 4),
        'shares': 0,  # not reported
    }]

# ─── STEP 3: Fill shares where missing ────────────────────────────────────────

def fill_shares(all_rows, conn):
    """For rows with aum_usd but no shares, compute shares = aum / price."""
    cur = conn.cursor()
    cur.execute("SELECT e.ticker, emf.price_usd FROM etp_monthly_fund emf JOIN etp e ON e.etp_id = emf.etp_id WHERE emf.month_id = %s", (MONTH_ID,))
    prices = {row[0]: float(row[1]) for row in cur.fetchall()}

    filled = 0
    for r in all_rows:
        if r['shares'] == 0 and r['aum_usd'] > 0:
            price = prices.get(r['ticker'])
            if price and price > 0:
                r['shares'] = round(r['aum_usd'] / price, 6)
                filled += 1
    print(f"\n  Filled {filled} missing share counts from price data")

# ─── STEP 4: Validate ─────────────────────────────────────────────────────────

def validate(all_rows):
    """Cross-check against Grace's summary totals."""
    print("\n=== STEP 3: Validation ===")

    # Grace's summary (Feb-26 column, in $M)
    grace = {
        'KSD (Korea Securities Depository) - Retail': 927.6,
        'Korea Investment Management -ACE TESLA Value Chain ETF': 23.9,
        'SBI': 29.5,
        'Rakuten': 61.9,
        'Monex': 7.3,
        'Matsui': 1.8,
        'Asset Plus Asset Management': 0.25,
        'SYFE': 1.0,
    }

    # Aggregate our parsed data
    by_exchange = {}
    for r in all_rows:
        ex = r['exchange']
        by_exchange[ex] = by_exchange.get(ex, 0) + r['aum_usd']

    print(f"  {'Exchange':<55s} {'Parsed':>10s} {'Grace':>10s} {'Diff':>10s}")
    print(f"  {'-'*55} {'-'*10} {'-'*10} {'-'*10}")

    all_ok = True
    for ex, grace_val in grace.items():
        parsed_val = by_exchange.get(ex, 0) / 1e6
        diff = parsed_val - grace_val
        flag = "OK" if abs(diff) < 5 else "CHECK"  # $5M tolerance
        if abs(diff) >= 5:
            all_ok = False
        print(f"  {ex:<55s} {parsed_val:>9.1f}M {grace_val:>9.1f}M {diff:>+9.1f}M  {flag}")

    total_parsed = sum(by_exchange.values()) / 1e6
    print(f"\n  Total parsed (fresh data only): ${total_parsed:.1f}M")
    return all_ok

# ─── STEP 5: Insert into fact table ───────────────────────────────────────────

def insert_exchange_aum(all_rows, conn):
    """Insert parsed rows into etp_exchange_monthly_aum."""
    print("\n=== STEP 4: Inserting into etp_exchange_monthly_aum ===")

    etp_map = get_etp_map(conn)
    exchange_map = get_exchange_map(conn)
    cur = conn.cursor()

    inserted = 0
    skipped = []
    for r in all_rows:
        etp_id = etp_map.get(r['ticker'])
        exchange_id = exchange_map.get(r['exchange'])

        if not etp_id:
            skipped.append(f"no etp_id for {r['ticker']}")
            continue
        if not exchange_id:
            skipped.append(f"no exchange_id for {r['exchange']}")
            continue

        cur.execute("""
            INSERT INTO etp_exchange_monthly_aum (etp_id, exchange_id, month_id, exchange_aum_usd, shares_outstanding)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (etp_id, exchange_id, month_id) DO NOTHING
        """, (etp_id, exchange_id, MONTH_ID,
              Decimal(str(r['aum_usd'])), Decimal(str(r['shares']))))
        inserted += 1

    conn.commit()
    print(f"  Inserted: {inserted} rows")
    if skipped:
        for s in set(skipped):
            print(f"  SKIPPED: {s}")

# ─── STEP 6: Carry forward quarterly vendors ─────────────────────────────────

def carry_forward_quarterly(conn):
    """Copy month 12 data for quarterly exchanges into month 13."""
    print("\n=== STEP 5: Carry forward quarterly vendors ===")
    cur = conn.cursor()

    # Quarterly exchanges: MooMoo (Japan=5, Singapore=11, Malaysia=13, HK=9), Oriental Harbour (10)
    quarterly_ids = [5, 9, 10, 11, 13]

    cur.execute("""
        INSERT INTO etp_exchange_monthly_aum (etp_id, exchange_id, month_id, exchange_aum_usd, shares_outstanding)
        SELECT etp_id, exchange_id, %s, exchange_aum_usd, shares_outstanding
        FROM etp_exchange_monthly_aum
        WHERE month_id = 12 AND exchange_id = ANY(%s)
        ON CONFLICT (etp_id, exchange_id, month_id) DO NOTHING
    """, (MONTH_ID, quarterly_ids))

    conn.commit()
    print(f"  Carried forward: {cur.rowcount} rows from month 12")

# ─── STEP 7: Verify ──────────────────────────────────────────────────────────

def verify(conn):
    """Check the views produce sane numbers."""
    print("\n=== STEP 6: Verification ===")
    cur = conn.cursor()

    cur.execute("""
        SELECT e.name, c.name as country,
               count(*) as etps,
               round(sum(exa.exchange_aum_usd)/1000000, 1) as aum_mil
        FROM etp_exchange_monthly_aum exa
        JOIN exchange e ON e.exchange_id = exa.exchange_id
        JOIN country c ON c.country_id = e.country_id
        WHERE exa.month_id = %s
        GROUP BY e.name, c.name
        ORDER BY aum_mil DESC
    """, (MONTH_ID,))

    print(f"\n  {'Exchange':<55s} {'Country':>10s} {'ETPs':>5s} {'AUM($M)':>10s}")
    print(f"  {'-'*55} {'-'*10} {'-'*5} {'-'*10}")
    total = 0
    for row in cur.fetchall():
        print(f"  {row[0]:<55s} {row[1]:>10s} {row[2]:>5d} {float(row[3]):>10.1f}")
        total += float(row[3])
    print(f"\n  TOTAL: ${total:.1f}M")

    # Compare with Grace's total
    print(f"  Grace's total: $1,362.45M")
    print(f"  Difference: ${total - 1362.45:+.1f}M")

    # Check family rollup view
    cur.execute("""
        SELECT month_end, total_asian_aum, total_rex_aum, pct_asia_aum
        FROM asia_family_rollup
        WHERE month_end = '2026-02-28'
    """)
    row = cur.fetchone()
    if row:
        print(f"\n  asia_family_rollup for Feb 2026:")
        print(f"    Total Asian AUM: ${float(row[1])/1e6:.1f}M")
        print(f"    Total REX AUM:   ${float(row[2])/1e6:.1f}M")
        print(f"    % in Asia:       {float(row[3])*100:.1f}%")

# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    conn = get_conn()

    # Step 1: Fund-level price + AUM
    load_etp_monthly_fund(conn)

    # Step 2: Parse all broker files
    print("\n=== STEP 2: Parsing broker files ===")
    all_rows = []
    all_rows.extend(parse_korea_ksd(conn))
    all_rows.extend(parse_korea_ace(conn))
    all_rows.extend(parse_japan_rakuten())
    all_rows.extend(parse_japan_sbi())
    all_rows.extend(parse_japan_monex())
    all_rows.extend(parse_japan_matsui())
    all_rows.extend(parse_thailand(conn))
    all_rows.extend(parse_hk_syfe())

    # Fill missing shares
    fill_shares(all_rows, conn)

    # Step 3: Validate
    validate(all_rows)

    # Step 4: Insert
    insert_exchange_aum(all_rows, conn)

    # Step 5: Carry forward quarterly
    carry_forward_quarterly(conn)

    # Step 6: Verify
    verify(conn)

    conn.close()
    print("\nDone.")
