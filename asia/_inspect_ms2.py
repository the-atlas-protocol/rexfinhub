"""Inspect microsector sheet structure of fresh Bloomberg file."""
import openpyxl
from pathlib import Path
import traceback

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")

OUT = open("_ms_result.txt", "w", encoding="utf-8")
def P(*a, **k): print(*a, **k, file=OUT); OUT.flush()

try:
    wb = openpyxl.load_workbook(BB, data_only=True, read_only=False)
    ws = wb["microsector"]
    P(f"microsector sheet: {ws.max_row}r x {ws.max_column}c")
    P()

    # Show row 4 (tickers) fully, col by col
    P("=== Row 4 (tickers/headers) ===")
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=c).value
        P(f"  col {c:>3}: {v!r}")

    P()
    P("=== Row 3 (descriptions) first 5 cols ===")
    for c in range(1, 6):
        v = ws.cell(row=3, column=c).value
        P(f"  col {c}: {v!r}")

    # Find last Mar 2026 and Feb 2026 dates
    P()
    P("=== Scanning for month-end rows ===")
    last_mar_row = last_feb_row = last_apr_row = last_date_row = None
    for r in range(5, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if d and hasattr(d, "year") and d.year == 2026:
            last_date_row = r
            if d.month == 2: last_feb_row = r
            if d.month == 3: last_mar_row = r
            if d.month == 4: last_apr_row = r
    P(f"Feb last row: {last_feb_row}, date: {ws.cell(row=last_feb_row, column=1).value if last_feb_row else 'none'}")
    P(f"Mar last row: {last_mar_row}, date: {ws.cell(row=last_mar_row, column=1).value if last_mar_row else 'none'}")
    P(f"Apr last row: {last_apr_row}, date: {ws.cell(row=last_apr_row, column=1).value if last_apr_row else 'none'}")
    P(f"Most recent date row: {last_date_row}, date: {ws.cell(row=last_date_row, column=1).value if last_date_row else 'none'}")

    P()
    P("=== Key tickers Feb27 / Mar31 / last Apr ===")
    KEY = ["GDXU", "FNGU", "BULZ", "SHNY", "OILU", "NRGU", "WTIU", "DULL"]
    for ticker in KEY:
        col = None
        for c in range(2, ws.max_column + 1):
            if ws.cell(row=4, column=c).value == ticker:
                col = c; break
        if not col:
            P(f"  {ticker}: not found")
            continue
        feb_v = ws.cell(row=last_feb_row, column=col).value if last_feb_row else None
        mar_v = ws.cell(row=last_mar_row, column=col).value if last_mar_row else None
        apr_v = ws.cell(row=last_apr_row, column=col).value if last_apr_row else None
        def f(v): return f"${v/1e6:,.2f}M" if isinstance(v, (int, float)) else str(v)
        P(f"  {ticker:6s}  Feb27: {f(feb_v):>14}   Mar31: {f(mar_v):>14}   Apr-last: {f(apr_v):>14}")

    wb.close()
    P("\nDONE")
except Exception as e:
    P(f"ERROR: {e}")
    P(traceback.format_exc())
finally:
    OUT.close()
