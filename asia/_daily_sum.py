"""Compute our Bloomberg total for specific dates, compare to Seamus's daily emails."""
import openpyxl
from pathlib import Path
import psycopg2
import psycopg2.extras
from datetime import date

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")

OUT = open("_daily_sum.txt", "w", encoding="utf-8")
def P(*a, **k): print(*a, **k, file=OUT); OUT.flush()

# DB tickers
conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
cur.execute("SELECT ticker FROM etp")
rex_tickers = set(r["ticker"] for r in cur.fetchall())
conn.close()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

# data_aum header
ws = wb["data_aum"]
aum_header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

# Build ticker -> col index map
ticker_to_col = {}
for i, col in enumerate(aum_header[1:], 1):
    if not col or " Equity" not in str(col): continue
    parts = str(col).split()
    base, listing = parts[0], parts[1]
    key = f"{base}_LN" if listing == "LN" else base
    ticker_to_col[key] = i

# Microsector setup
ws_ms = wb["microsector"]
ms_tickers_row = list(next(ws_ms.iter_rows(min_row=4, max_row=4, values_only=True)))
ms_tickers_to_col = {}
for i, t in enumerate(ms_tickers_row[1:], 1):
    if t and isinstance(t, str):
        ms_tickers_to_col[t.strip()] = i

# Dates to check
TARGET_DATES = [
    (date(2026, 2, 27), "Feb 27", 6_711_700_000),
    (date(2026, 3, 26), "Mar 26", 5_821_500_000),
    (date(2026, 3, 27), "Mar 27", None),
    (date(2026, 3, 30), "Mar 30", 5_604_400_000),
    (date(2026, 3, 31), "Mar 31", 6_061_200_000),
    (date(2026, 4, 1),  "Apr 1",  None),
]

# Pre-index data_aum and microsector rows by date
aum_by_date = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year'):
        aum_by_date[d.date() if hasattr(d, 'date') else d] = row

ms_by_date = {}
for row in ws_ms.iter_rows(min_row=5, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year'):
        ms_by_date[d.date() if hasattr(d, 'date') else d] = row

def compute_rex_total(target: date):
    aum_row = aum_by_date.get(target)
    if not aum_row: return None, 0
    ms_row = ms_by_date.get(target)

    total = 0.0
    n = 0
    for t in rex_tickers:
        col = ticker_to_col.get(t)
        if col is None: continue
        val = aum_row[col] if col < len(aum_row) else None
        if not (isinstance(val, (int, float)) and val > 0): continue
        # Apply microsector overwrite
        if t in ms_tickers_to_col and ms_row:
            mi = ms_tickers_to_col[t]
            ms_val = ms_row[mi] if mi < len(ms_row) else None
            if isinstance(ms_val, (int, float)) and ms_val > 0:
                total += float(ms_val)
                n += 1
                continue
        total += float(val) * 1_000_000  # $M -> $
        n += 1
    return total, n

P(f"{'Date':<10}  {'Seamus':>18}  {'My BBG':>18}  {'Gap':>15}  {'# tick':>8}")
P("-" * 82)
for dt, label, seamus in TARGET_DATES:
    total, n = compute_rex_total(dt)
    if total is None:
        P(f"{label:<10}  (no data for {dt})")
        continue
    gap = total - seamus if seamus else None
    gap_str = f"${gap/1e6:+10.1f}M" if gap is not None else "—"
    seamus_str = f"${seamus/1e9:.4f}B" if seamus else "—"
    P(f"{label:<10}  {seamus_str:>18}  ${total/1e9:.4f}B  {gap_str:>15}  {n:>8}")

# Specific diagnostic for Mar 31: which tickers contribute most
P()
P("=== Mar 31 top 20 tickers (for compare with Seamus's table) ===")
aum_row = aum_by_date.get(date(2026, 3, 31))
ms_row = ms_by_date.get(date(2026, 3, 31))
per_ticker = {}
for t in rex_tickers:
    col = ticker_to_col.get(t)
    if col is None: continue
    val = aum_row[col] if col < len(aum_row) else None
    if not isinstance(val, (int, float)) or val <= 0: continue
    if t in ms_tickers_to_col and ms_row:
        mi = ms_tickers_to_col[t]
        ms_val = ms_row[mi] if mi < len(ms_row) else None
        if isinstance(ms_val, (int, float)) and ms_val > 0:
            per_ticker[t] = float(ms_val)
            continue
    per_ticker[t] = float(val) * 1_000_000

top = sorted(per_ticker.items(), key=lambda x: -x[1])
for t, v in top[:25]:
    src = "microsector" if t in ms_tickers_to_col else "data_aum"
    P(f"  {t:<10}  ${v/1e6:>10,.1f}M   [{src}]")

wb.close()
OUT.close()
