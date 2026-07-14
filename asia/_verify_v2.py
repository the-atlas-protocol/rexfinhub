"""Reconcile v2 Feb26 against the shipped PDF numbers."""
from openpyxl import load_workbook

wb = load_workbook("REX_Asia_Monthly_Log_v2.xlsx")
ws = wb["2026-02"]

# Map row 2 header text -> column
cols = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=2, column=c).value
    if v:
        cols[str(v).replace("\n", " ").strip()] = c

print("Asia total reconciliation:")
asia_col = None
for label, idx in cols.items():
    if "Asia Total" in label:
        asia_col = idx; break

total_musd = 0.0
by_family = {}
for r in range(3, ws.max_row + 1):
    fam = ws.cell(row=r, column=2).value
    v = ws.cell(row=r, column=asia_col).value if asia_col else None
    if isinstance(v, (int, float)):
        total_musd += v
        by_family[fam] = by_family.get(fam, 0) + v

target = 1319.8532549129998
print(f"  v2 Feb26 total: ${total_musd:,.4f}M")
print(f"  DB target:      ${target:,.4f}M")
print(f"  diff:           ${total_musd - target:+.4f}M  ({(total_musd - target)*1e6:+,.2f} raw $)")
print()
print("By family:")
for f, v in sorted(by_family.items(), key=lambda x: -x[1]):
    print(f"  {f:16s}  ${v:,.2f}M")

print()
# Top funds vs PDF
print("Top funds v2 vs PDF:")
pdf_top = {"FNGU": 178, "TSLT": 164, "MSTU": 156, "BMNU": 146, "GDXU": 137,
           "BULZ": 114, "NVDX": 70.5, "AIPI": 40.7, "FEPI": 37.3, "SHNY": 37.1}
for r in range(3, ws.max_row + 1):
    t = ws.cell(row=r, column=1).value
    if t in pdf_top:
        v = ws.cell(row=r, column=asia_col).value
        if isinstance(v, (int, float)):
            print(f"  {t:6s}  v2 ${v:,.2f}M   pdf ${pdf_top[t]:.1f}M   diff ${v - pdf_top[t]:+.2f}M")

# Exchange totals
print()
print("Vendor USD totals (Feb26):")
# Find vendor USD columns (even col after each shares col in VENDOR section)
vendor_usd_cols = []  # (country, exchange, col)
# Walk row 1 banners then row 2 sub-headers
# We know structure: first 3 cols IDENTITY, then vendor groups, then OUR SIDE (4), DERIVED (5), FLAGS (3)
max_c = ws.max_column
last_vendor_col = max_c - 4 - 5 - 3  # IDENTITY=3 but the 3 trails at end
# Actually simpler: iterate row 2 looking for "USD $M" in label
for c in range(1, ws.max_column + 1):
    h2 = str(ws.cell(row=2, column=c).value or "")
    h1 = str(ws.cell(row=1, column=c).value or "")
    if "USD $M" in h2:
        # Country: scan left for non-empty merged row1
        country = None
        for cc in range(c, 0, -1):
            v = ws.cell(row=1, column=cc).value
            if v:
                country = v; break
        exch = h2.split("\n")[0].replace("USD $M", "").strip()
        vendor_usd_cols.append((country, exch, c))

for (country, exch, c) in vendor_usd_cols:
    total = 0.0
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            total += v
    print(f"  {country:10s} {exch:40s}  ${total:,.2f}M")
