"""
audit_grace_vs_expected.py

Parses Grace's Asset report summary file for a given month,
cross-references against config/vendor_status.yaml, and flags:
  - active vendors with blank Grace value (Grace delayed?)
  - zeroed vendors with non-zero Grace value (needs config update)
  - frozen/waiting vendors with fresh Grace value (unblock?)
  - vendors in Grace's file that vendor_status.yaml doesn't know about (new vendor)

Usage:
    python audit_grace_vs_expected.py --month 2026-04
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path

import yaml
import openpyxl

ap = argparse.ArgumentParser()
ap.add_argument("--month", required=True, help="YYYY-MM")
args = ap.parse_args()
YEAR, MONTH = map(int, args.month.split("-"))

CONFIG = yaml.safe_load(open("config/vendor_status.yaml"))
vendors = CONFIG["vendors"]

DATA_DIR = Path(f"grace_data/{args.month}")
summary_files = list(DATA_DIR.glob("Asset report summary*.xlsx"))
if not summary_files:
    print(f"ERROR: No 'Asset report summary' file in {DATA_DIR}")
    sys.exit(1)
summary = summary_files[0]
print(f"Reading: {summary}")

wb = openpyxl.load_workbook(summary, data_only=True)
ws = wb.active

# Find the month column (header row)
month_col = None
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=2, column=c).value  # row 2 typically has dates
    if v and hasattr(v, "year") and v.year == YEAR and v.month == MONTH:
        month_col = c; break
if not month_col:
    # Try scanning all rows
    for r in range(1, 5):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and hasattr(v, "year") and v.year == YEAR and v.month == MONTH:
                month_col = c; break
        if month_col: break
if not month_col:
    print(f"ERROR: Could not find {args.month} column in summary file")
    sys.exit(1)
print(f"Found {args.month} in column {month_col}")

# Parse each row: col A=country, col B=source name, month_col=value
grace_values = {}  # (country, source) -> value (None if blank)
for r in range(3, ws.max_row + 1):
    country = ws.cell(row=r, column=1).value
    source = ws.cell(row=r, column=2).value
    val = ws.cell(row=r, column=month_col).value
    if not country or not source:
        continue
    country = str(country).strip().replace("�", "").strip()  # clean unicode artifacts
    source = str(source).strip()
    if isinstance(val, (int, float)) and val > 0:
        grace_values[(country, source)] = float(val)
    else:
        grace_values[(country, source)] = None  # blank

print(f"Parsed {len(grace_values)} vendor rows from summary")
print()

# ── Cross-check ────────────────────────────────────────────────────────
issues = []
matched_yaml = set()

SRC_NAME_MAP = {
    # Grace's source-column name -> our YAML vendor exchange name
    "KSD (Korea Securities Depository) - Retail": "KSD (Korea Securities Depository) - Retail",
    "Korea Investment Management -ACE TESLA Value Chain ETF": "Korea Investment Management -ACE TESLA Value Chain ETF",
    "SBI": "SBI", "Rakuten": "Rakuten", "Monex": "Monex",
    "Mastui": "Matsui", "Matsui": "Matsui",
    "MooMoo": None,  # ambiguous — need country context
    "Futu/MooMoo": "Futu/MooMoo",
    "Oriental Harbour *": "Oriental Harbour *", "Oriental Harbour": "Oriental Harbour *",
    "SYFE": "SYFE",
    "ViewTrade": "ViewTrade",
    "Asset Plus Asset Management": "Asset Plus Asset Management",
}

def resolve_vendor_key(grace_country, grace_src):
    """Map Grace's (country, source) to our YAML key 'Country/Exchange'."""
    # Grace bundles "Taiwan, SG, HK" for ViewTrade into one row — handled below as known bundle
    if "Taiwan" in grace_country and "SG" in grace_country and "HK" in grace_country and "ViewTrade" in grace_src:
        return "Hong Kong/ViewTrade"  # map to primary; all 3 share same status
    ex = SRC_NAME_MAP.get(grace_src, grace_src)
    if ex is None:  # MooMoo — use country directly
        ex = grace_src
    return f"{grace_country}/{ex}"

print(f"{'STATUS':<25}  {'VENDOR':<55}  {'GRACE':<12}  NOTE")
print("-" * 130)
for (country, source), val in grace_values.items():
    ykey = resolve_vendor_key(country, source)
    vcfg = vendors.get(ykey)

    if vcfg is None:
        # Grace's row is something YAML doesn't know
        issues.append(f"UNKNOWN VENDOR in Grace: {country}/{source} = {val}")
        print(f"{'UNKNOWN_VENDOR':<25}  {ykey[:55]:<55}  {str(val):<12}  not in config")
        continue
    matched_yaml.add(ykey)
    status = vcfg["status"]

    if status in ("active", "active_aggregate_only"):
        if val is None:
            issues.append(f"MISSING: {status} vendor {ykey} has blank Grace value")
            print(f"{'MISSING':<25}  {ykey[:55]:<55}  blank        Grace delayed?")
        else:
            print(f"{'OK (' + status + ')':<25}  {ykey[:55]:<55}  ${val:>8.2f}M  expected")
    elif status == "zeroed":
        if val is None or val == 0:
            print(f"{'OK (zeroed)':<25}  {ykey[:55]:<55}  blank        correct")
        else:
            issues.append(f"UNEXPECTED: zeroed vendor {ykey} has fresh Grace value ${val}M")
            print(f"{'UNEXPECTED_VALUE':<25}  {ykey[:55]:<55}  ${val:>8.2f}M  YAML says zeroed — update?")
    elif status in ("frozen_permanent", "waiting_13f", "waiting_hardcopy", "waiting"):
        if val is None:
            print(f"{'OK (' + status + ')':<25}  {ykey[:55]:<55}  blank        waiting for data")
        else:
            issues.append(f"UNBLOCK? {status} vendor {ykey} has fresh Grace value ${val}M")
            print(f"{'NEW_DATA':<25}  {ykey[:55]:<55}  ${val:>8.2f}M  was {status} — unblock?")
    else:
        print(f"{'UNKNOWN_STATUS':<25}  {ykey[:55]:<55}  {str(val)}")

# Check vendors in YAML not found in Grace's file
print()
for ykey, vcfg in vendors.items():
    if ykey in matched_yaml: continue
    if vcfg["status"] == "active":
        issues.append(f"ABSENT from Grace: active vendor {ykey}")
        print(f"{'NOT_IN_GRACE':<25}  {ykey[:55]:<55}  —           active but missing")

print()
print("=" * 60)
if issues:
    print(f"FOUND {len(issues)} ISSUES:")
    for i in issues:
        print(f"  - {i}")
    print()
    print("Action: review each issue and either update config/vendor_status.yaml or investigate data.")
    sys.exit(1)
else:
    print(f"CLEAN — {args.month} matches config expectations.")
    sys.exit(0)

wb.close()
