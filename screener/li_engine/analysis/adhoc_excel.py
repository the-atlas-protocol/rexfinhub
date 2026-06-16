"""Ad-hoc L&I analysis -> Excel, driven entirely off the T-REX Recommendations
report's own loaders so the numbers match what we send. No parallel logic.

Columns (per Ryu, 2026-06-11): Ticker | Score | Pctile | REX Position |
# Competitors | Top Comp AUM | Verdict.  Score = 0-100 (report style), NOT the
raw z-composite. Score History sheet shows the 0-100 percentile over time.

Usage: python adhoc_excel.py OUT.xlsx GROUP=TK,TK ...
"""
import sys, sqlite3
import pandas as pd

sys.path.insert(0, "/home/jarvis/rexfinhub")
from screener.li_engine.analysis.trex_combined_v9 import (
    load_rex_position, load_underlier_competition, load_underlier_live, _canon,
)
from screener.li_engine.persistence import DB

# ---- args
out = sys.argv[1]
groups, order = {}, []
for arg in sys.argv[2:]:
    label, csv = arg.split("=", 1)
    for tk in csv.split(","):
        t = _canon(tk)
        if t and t not in groups:
            groups[t] = label; order.append(t)

# ---- report loaders (same data the sent report uses)
rex_pos = load_rex_position()              # u -> ("Live"/"Filed"/"—", color)
comp    = load_underlier_competition()     # u -> {"n": # non-REX filers, "eff": date}
live    = load_underlier_live()            # u -> {"top_aum", "has_long", "has_inv"}

# ---- production score (li_engine_daily.final_score, v1.0.1) + 0-100 percentile + history
conn = sqlite3.connect(str(DB))
latest = conn.execute("SELECT MAX(run_date) FROM li_engine_daily").fetchone()[0]
uni = pd.read_sql_query("SELECT ticker, final_score FROM li_engine_daily WHERE run_date=?", conn, params=[latest])
uni["t"] = uni["ticker"].map(_canon)
uni = uni.dropna(subset=["final_score"]).drop_duplicates("t").sort_values("final_score", ascending=False).reset_index(drop=True)
uni["rank"] = uni["final_score"].rank(ascending=False, method="min").astype(int)
uni["pct"] = (uni["final_score"].rank(pct=True) * 100)
n_univ = len(uni)
scoremap = uni.set_index("t")[["final_score", "rank", "pct"]].to_dict("index")
allhist = pd.read_sql_query("SELECT run_date, ticker, final_score FROM li_engine_daily", conn)
allhist["t"] = allhist["ticker"].map(_canon)
# percentile per run_date (0-100) so history reads on the same 0-100 scale as the report
allhist["pct"] = allhist.groupby("run_date")["final_score"].rank(pct=True) * 100
conn.close()

def verdict(rexlabel, n_filers, topaum, has_long, has_score):
    """Separate LIVE competition (top_aum / has_long) from the FILING race
    (# filers). $0M top with filers = nobody live yet = still whitespace."""
    if not has_score:
        return "No score — delisted / sub-floor; skip"
    if rexlabel == "Live":
        return "REX already live — skip"
    live = has_long and topaum > 0
    if live:
        if rexlabel == "Filed":
            return f"REX filed; competitor live (top ${topaum:.0f}M, {n_filers} filer(s)) — launch to compete"
        if 1 <= n_filers <= 2 and topaum > 100:
            return f"Proven demand (top ${topaum:.0f}M), beatable head-count — strong enter"
        return f"Crowded — competitor live top ${topaum:.0f}M, {n_filers} filers — lower priority"
    race = f"{n_filers} filer(s) racing" if n_filers else "no filers"
    if rexlabel == "Filed":
        return f"REX filed; no live product yet, {race} — launch fast to be first"
    return f"Whitespace, no live product ({race}) — file" + (" fast" if n_filers else " (first mover)")

rows = []
for t in order:
    sc = scoremap.get(t)
    rexlabel = rex_pos.get(t, ("—",))[0]
    ncomp = int(comp.get(t, {}).get("n", 0))
    topaum = float(live.get(t, {}).get("top_aum", 0.0))
    has_long = bool(live.get(t, {}).get("has_long", False))
    rec = {
        "Group": groups[t],
        "Ticker": t,
        "Score": round(sc["pct"], 1) if sc else None,
        "Pctile": f"{sc['pct']:.0f}th (rank {sc['rank']}/{n_univ})" if sc else "—",
        "REX Position": {"Live": "Live (we trade it)", "Filed": "Filed (in registration)"}.get(rexlabel, "Not in"),
        "# Competitor Filers": ncomp,
        "Top Comp AUM (live)": f"${topaum:.0f}M" if topaum > 0 else ("none live" if ncomp else "—"),
        "Comp Live Long?": "yes" if has_long else "no",
        "Verdict": verdict(rexlabel, ncomp, topaum, has_long, sc is not None),
        "Final Score (raw)": round(sc["final_score"], 2) if sc else None,
    }
    rows.append(rec)
df = pd.DataFrame(rows)

# ---- score history (0-100 percentile over time)
hp = allhist[allhist["t"].isin(order)].pivot_table(index="t", columns="run_date", values="pct", aggfunc="max")
hist_cols = sorted(hp.columns)
hdf = hp.reindex(order)[hist_cols].round(0)
hdf.insert(0, "Group", [groups[t] for t in hdf.index])
if len(hist_cols) >= 2:
    hdf["Δ pct"] = (hp[hist_cols[-1]] - hp[hist_cols[0]]).reindex(order).round(0)
hdf = hdf.reset_index().rename(columns={"t": "Ticker"})

meta = pd.DataFrame({"Field": [
    "Generated","Latest scoring run","Scorer","Universe size","Score column","Pctile column",
    "REX Position source","# Competitors source","Top Comp AUM source","Verdict","History scale"],
    "Value": [
    pd.Timestamp.today().strftime("%Y-%m-%d %H:%M"), latest, "li_engine_daily v1.0.1 (same as sent T-REX report)",
    n_univ, "0-100 percentile (report style — NOT the raw z-composite, which reads ~ -1..2)",
    "percentile + integer rank across the scored universe",
    "trex_combined_v9.load_rex_position() — Live / Filed / Not in (from competitor_counts)",
    "trex_combined_v9.load_underlier_competition() — distinct non-REX L&I filers on the underlier",
    "trex_combined_v9.load_underlier_live() — largest live L&I product AUM on the underlier",
    "Mirrors report Section 5 logic: REX-not-in + 1-2 competitors + top AUM>$100M = strong enter; >2 = crowded",
    "0-100 percentile per run_date (so values are comparable across days)"]})

with pd.ExcelWriter(out, engine="openpyxl") as xw:
    df.to_excel(xw, sheet_name="Analysis", index=False)
    hdf.to_excel(xw, sheet_name="Score History", index=False)
    meta.to_excel(xw, sheet_name="Methodology", index=False)

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
wb = load_workbook(out)
hf, hfont = PatternFill("solid", fgColor="231F20"), Font(bold=True, color="FFFFFF")
for sh, frame in [("Analysis", df), ("Score History", hdf), ("Methodology", meta)]:
    ws = wb[sh]
    ws.freeze_panes = "C2" if sh != "Methodology" else "A2"
    for c, name in enumerate(frame.columns, 1):
        cell = ws.cell(1, c); cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(vertical="center")
        bmax = max([len(str(v)) for v in frame[name].tolist()] + [len(str(name))])
        ws.column_dimensions[get_column_letter(c)].width = max(10, min(42, bmax + 2))
if "Score" in df.columns:
    sc = get_column_letter(list(df.columns).index("Score") + 1)
    wb["Analysis"].conditional_formatting.add(f"{sc}2:{sc}{len(df)+1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=50, mid_color="FFEB84",
                       end_type="num", end_value=100, end_color="63BE7B"))
wb.save(out)
print(f"WROTE {out} ({len(df)} tickers, {len(hist_cols)} history cols, universe {n_univ})")
