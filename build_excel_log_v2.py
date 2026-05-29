"""
Build REX_Asia_Monthly_Log_v2.xlsx — shares-first ledger, wide per-month layout.

Design per 2026-04-22 session:
- Shares are the fundamental unit. USD = shares * price (derived).
- For vendors that send shares directly: logged as reported. USD derived from BBG price.
- For vendors that send USD only: shares inferred as vendor_usd / bbg_price (italic/grey to flag).
- For frozen reporters (Futu HK, etc.): shares carried forward, USD floats with price.
- Precision preserved internally. Rounding via number_format only.
- Funds sorted by Asia Total DESC within family; non-Asia funds appended in a muted block.
- Two-row country header: merged country banner over per-vendor column pairs (shares | USD).
- Row zebra stripes. Family colors on family column.
- Output: REX_Asia_Monthly_Log_v2.xlsx (v1 stays untouched).
"""
import json
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import psycopg2
import psycopg2.extras

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")
OUT = Path("REX_Asia_Monthly_Log.xlsx")

# ── Style palette ──
FILL_HEADER     = PatternFill("solid", fgColor="1a1a2e")
FONT_HEADER     = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
FILL_COUNTRY    = PatternFill("solid", fgColor="0984e3")   # blue banner
FILL_OURS       = PatternFill("solid", fgColor="27ae60")   # green
FILL_DERIVED    = PatternFill("solid", fgColor="e67e22")   # orange
FILL_IDENT      = PatternFill("solid", fgColor="2d3436")   # dark
FILL_ZEBRA      = PatternFill("solid", fgColor="F4F6F8")
FILL_FIRST      = PatternFill("solid", fgColor="FFF59D")   # yellow — first appearance
FILL_FROZEN     = PatternFill("solid", fgColor="ECEFF1")   # grey — frozen reporter
FILL_SEP        = PatternFill("solid", fgColor="B0BEC5")

FAMILY_FILL = {
    "T-REX":        PatternFill("solid", fgColor="D6EAF8"),
    "MicroSectors": PatternFill("solid", fgColor="D5F5E3"),
    "Income":       PatternFill("solid", fgColor="E8DAEF"),
    "REX Osprey":   PatternFill("solid", fgColor="FCF3CF"),
    "Other":        PatternFill("solid", fgColor="F2F3F4"),
}

GREY_ITALIC = Font(italic=True, color="757575", size=10, name="Calibri")
BLACK_REG   = Font(color="000000", size=10, name="Calibri")
BLACK_BOLD  = Font(bold=True, color="000000", size=10, name="Calibri")
BORDER_THIN = Border(right=Side(style="thin", color="CFD8DC"))

COUNTRY_ORDER = {"Korea": 1, "Japan": 2, "Hong Kong": 3, "Singapore": 4, "Malaysia": 5, "Taiwan": 6, "Thailand": 7}

# ── Data loading ──
def db():
    return psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                            cursor_factory=psycopg2.extras.RealDictCursor)

def load_classification():
    return json.loads(Path("tickers_classification.json").read_text())

def load_fx():
    return json.loads(Path("fx_rates.json").read_text())

def fetch_bloomberg_monthly():
    """Returns {month_label: {ticker: {'aum_musd': x, 'price': y, 'source': 'data_aum|microsector'}}}"""
    aum = pd.read_excel(BB, sheet_name="data_aum", header=0)
    aum = aum.rename(columns={aum.columns[0]: "Date"})
    aum["Date"] = pd.to_datetime(aum["Date"], errors="coerce")

    price = pd.read_excel(BB, sheet_name="data_price", header=0)
    price = price.rename(columns={price.columns[0]: "Date"})
    price["Date"] = pd.to_datetime(price["Date"], errors="coerce")

    ms_raw = pd.read_excel(BB, sheet_name="microsector", header=None)
    ms_tickers = ms_raw.iloc[3, 1:].tolist()
    ms = ms_raw.iloc[4:].copy()
    ms.columns = ["Date"] + ms_tickers
    ms["Date"] = pd.to_datetime(ms["Date"], errors="coerce")
    ms = ms.dropna(subset=["Date"])

    def last_of_month(df, y, m):
        mask = (df["Date"].dt.year == y) & (df["Date"].dt.month == m)
        rows = df[mask]
        return rows.iloc[-1] if not rows.empty else None

    result = {}
    for y, m in [(2025, mm) for mm in range(2, 13)] + [(2026, mm) for mm in range(1, 5)]:
        label = f"{y:04d}-{m:02d}"
        aum_row = last_of_month(aum, y, m)
        price_row = last_of_month(price, y, m)
        ms_row = last_of_month(ms, y, m)
        if aum_row is None:
            continue
        per_ticker = {}
        for col in aum_row.index:
            if col == "Date" or not isinstance(col, str) or " Equity" not in col:
                continue
            parts = col.split()
            if len(parts) < 3: continue
            bbg_base, suffix = parts[0], parts[1]
            # Map back to DB ticker conv: LN->{base}_LN, else base
            db_ticker = f"{bbg_base}_LN" if suffix == "LN" else bbg_base
            v_aum = aum_row.get(col)
            v_pr = price_row.get(col) if price_row is not None else None
            if pd.isna(v_aum): continue
            per_ticker[db_ticker] = {
                "aum_musd": float(v_aum),       # $M
                "price": float(v_pr) if (v_pr is not None and pd.notna(v_pr) and isinstance(v_pr, (int, float))) else None,
                "source": "data_aum",
                "bbg_date": aum_row["Date"].date().isoformat(),
            }
        # Microsector overwrite
        if ms_row is not None:
            for t in ms_tickers:
                v = ms_row.get(t)
                if pd.notna(v):
                    if t not in per_ticker:
                        per_ticker[t] = {"aum_musd": 0, "price": None, "source": None, "bbg_date": ms_row["Date"].date().isoformat()}
                    per_ticker[t]["aum_musd"] = float(v) / 1e6  # raw $ -> $M for consistency
                    per_ticker[t]["source"] = "microsector"
                    per_ticker[t]["bbg_date"] = ms_row["Date"].date().isoformat()
                    # Price: microsector sheet doesn't have price; rely on data_price
        result[label] = per_ticker
    return result

def fetch_db_month(month_id: int, prior_month_id: int | None):
    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT e.ticker, pf.name AS family, ex.name AS exchange, c.name AS country,
               m.exchange_aum_usd, m.source_type
        FROM etp_exchange_monthly_aum m
        JOIN etp e USING (etp_id) JOIN product_family pf USING (family_id)
        JOIN exchange ex USING (exchange_id) JOIN country c USING (country_id)
        WHERE m.month_id = %s
    """, (month_id,))
    cur_rows = cur.fetchall()

    prior_rows = []
    if prior_month_id is not None:
        cur.execute("""
            SELECT e.ticker, ex.name AS exchange, c.name AS country, m.exchange_aum_usd
            FROM etp_exchange_monthly_aum m
            JOIN etp e USING (etp_id) JOIN exchange ex USING (exchange_id) JOIN country c USING (country_id)
            WHERE m.month_id = %s
        """, (prior_month_id,))
        prior_rows = cur.fetchall()

    cur.execute("""
        SELECT e.ticker, e.name AS fund_name, pf.name AS family
        FROM etp e JOIN product_family pf USING (family_id)
        ORDER BY pf.name, e.ticker
    """)
    fund_rows = cur.fetchall()

    cur.execute("SELECT month_id, month_end FROM calendar_month ORDER BY month_id")
    months = cur.fetchall()
    conn.close()
    return cur_rows, prior_rows, fund_rows, months

# ── Layout helpers ──
def write_header_rows(ws, vendor_columns_by_country):
    """Two-row headers with country banner.
    Row 1: section bands (IDENTITY | <country banners> | OUR SIDE | DERIVED | FLAGS)
    Row 2: per-vendor sub-headers (each vendor = 2 cols: shares | USD $M)
    Start data at row 3.
    Returns: list of (col_index, field_type, meta) for data-write phase."""
    col = 1
    layout = []

    # IDENTITY: 3 cols
    ws.cell(row=1, column=col, value="IDENTITY").fill = FILL_IDENT
    ws.cell(row=1, column=col).font = FONT_HEADER
    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
    for i, h in enumerate(["Ticker", "Family", "Fund Name"]):
        c = ws.cell(row=2, column=col + i, value=h); c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")
    for i, field in enumerate(["ticker", "family", "fund_name"]):
        layout.append((col + i, field, None))
    col += 3

    # VENDOR sections — group by country
    for (country, exchanges) in vendor_columns_by_country:
        span = len(exchanges) * 2  # each exchange has shares + USD cols
        banner = ws.cell(row=1, column=col, value=country)
        banner.fill = FILL_COUNTRY; banner.font = FONT_HEADER
        banner.alignment = Alignment(horizontal="center")
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        for exch in exchanges:
            c1 = ws.cell(row=2, column=col, value=f"{exch}\nshares"); c1.fill = FILL_HEADER; c1.font = FONT_HEADER
            c1.alignment = Alignment(horizontal="center", wrap_text=True)
            c2 = ws.cell(row=2, column=col + 1, value=f"{exch}\nUSD $M"); c2.fill = FILL_HEADER; c2.font = FONT_HEADER
            c2.alignment = Alignment(horizontal="center", wrap_text=True)
            layout.append((col, "vendor_shares", (country, exch)))
            layout.append((col + 1, "vendor_usd_musd", (country, exch)))
            col += 2

    # OUR SIDE
    ws.cell(row=1, column=col, value="OUR SIDE (BBG)").fill = FILL_OURS; ws.cell(row=1, column=col).font = FONT_HEADER
    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
    for i, h in enumerate(["Global AUM $M", "Price $", "Global Shares (M)", "Source"]):
        c = ws.cell(row=2, column=col + i, value=h); c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, field in enumerate(["bbg_aum_musd", "bbg_price", "bbg_shares_mm", "bbg_source"]):
        layout.append((col + i, field, None))
    col += 4

    # DERIVED
    ws.cell(row=1, column=col, value="DERIVED").fill = FILL_DERIVED; ws.cell(row=1, column=col).font = FONT_HEADER
    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 4)
    for i, h in enumerate(["Asia Total $M", "Prior Asia $M", "MoM $M", "MoM %", "% in Asia"]):
        c = ws.cell(row=2, column=col + i, value=h); c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, field in enumerate(["asia_total_musd", "prior_asia_musd", "mom_musd", "mom_pct", "pct_in_asia"]):
        layout.append((col + i, field, None))
    col += 5

    # FLAGS
    ws.cell(row=1, column=col, value="FLAGS").fill = FILL_HEADER; ws.cell(row=1, column=col).font = FONT_HEADER
    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
    for i, h in enumerate(["Lifecycle", "First Appears", "Notes"]):
        c = ws.cell(row=2, column=col + i, value=h); c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")
    for i, field in enumerate(["lifecycle", "first_appears", "notes"]):
        layout.append((col + i, field, None))

    return layout, col + 2

def build_month_sheet(wb, label, month_id, month_end, prior_month_id, bbg_months, classification):
    ws = wb.create_sheet(label)

    cur_rows, prior_rows, fund_rows, _ = fetch_db_month(month_id, prior_month_id)

    # Exchanges seen in this month, grouped by country
    cols_by_country = {}
    for r in cur_rows:
        cols_by_country.setdefault(r["country"], set()).add(r["exchange"])
    ordered = sorted(cols_by_country.items(), key=lambda kv: COUNTRY_ORDER.get(kv[0], 99))
    vendor_columns_by_country = [(c, sorted(exs)) for c, exs in ordered]

    layout, max_col = write_header_rows(ws, vendor_columns_by_country)

    # Pivot month AUM by (ticker, country, exchange)
    cur_aum = {(r["ticker"], r["country"], r["exchange"]): float(r["exchange_aum_usd"]) for r in cur_rows}
    cur_source = {(r["ticker"], r["country"], r["exchange"]): r["source_type"] for r in cur_rows}
    prior_aum = {(r["ticker"], r["country"], r["exchange"]): float(r["exchange_aum_usd"]) for r in prior_rows}

    # BBG for this month
    bbg_this = bbg_months.get(label, {})

    # Order funds: by family with Asia-active at top (by Asia Total DESC), then non-Asia inactive
    fund_asia = {}
    for r in fund_rows:
        t = r["ticker"]
        total = sum(v for k, v in cur_aum.items() if k[0] == t)
        fund_asia[t] = total

    def fund_sort_key(fr):
        t = fr["ticker"]
        has_asia = fund_asia.get(t, 0) > 0
        return (0 if has_asia else 1, fr["family"], -fund_asia.get(t, 0), t)

    fund_rows_sorted = sorted(fund_rows, key=fund_sort_key)

    # Write data rows
    row = 3
    prev_section_has_asia = True
    for fr in fund_rows_sorted:
        t = fr["ticker"]
        cls = classification.get(t, {})
        has_asia = fund_asia.get(t, 0) > 0

        # Section separator when we transition from Asia-active to not
        if prev_section_has_asia and not has_asia:
            sep = ws.cell(row=row, column=1, value="— non-Asia REX funds —")
            sep.fill = FILL_SEP; sep.font = Font(italic=True, color="37474F", size=9)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
            sep.alignment = Alignment(horizontal="center")
            row += 1
            prev_section_has_asia = False

        # Zebra stripe (every other data row)
        zebra = (row % 2 == 1)
        family_fill = FAMILY_FILL.get(fr["family"], FAMILY_FILL["Other"])

        # BBG data
        bbg = bbg_this.get(t, {})
        bbg_aum_musd = bbg.get("aum_musd")
        bbg_price = bbg.get("price")
        bbg_shares_mm = (bbg_aum_musd / bbg_price) if (bbg_aum_musd is not None and bbg_price) else None  # $M / $ = M shares
        bbg_source = bbg.get("source")

        # Asia total for this fund
        asia_total_usd = sum(v for k, v in cur_aum.items() if k[0] == t)
        prior_total_usd = sum(v for k, v in prior_aum.items() if k[0] == t)
        mom_usd = asia_total_usd - prior_total_usd if prior_total_usd > 0 else None
        mom_pct = (mom_usd / prior_total_usd) if (mom_usd is not None and prior_total_usd > 0) else None
        pct_in_asia = (asia_total_usd / (bbg_aum_musd * 1e6)) if (bbg_aum_musd and asia_total_usd > 0) else None

        # First appears ANYWHERE new this month?
        first_appears_anywhere = False
        for k, v in cur_aum.items():
            if k[0] == t and v > 0 and k not in prior_aum:
                first_appears_anywhere = True
                break

        # Notes
        notes = []
        if cls.get("lifecycle") == "pending_launch":
            notes.append("pending launch")
        elif cls.get("lifecycle") == "delisted":
            notes.append(f"delisted (last {cls.get('last_seen')})")
        elif cls.get("days_stale", 0) > 7:
            notes.append(f"bbg stale {cls['days_stale']}d")
        if cls.get("bbg_suffix") == "LN":
            notes.append("UCITS (LN)")

        # Write the row
        values = {
            "ticker": t,
            "family": fr["family"],
            "fund_name": fr["fund_name"],
            "bbg_aum_musd": bbg_aum_musd,
            "bbg_price": bbg_price,
            "bbg_shares_mm": bbg_shares_mm,
            "bbg_source": bbg_source,
            "asia_total_musd": (asia_total_usd / 1e6) if asia_total_usd > 0 else None,
            "prior_asia_musd": (prior_total_usd / 1e6) if prior_total_usd > 0 else None,
            "mom_musd": (mom_usd / 1e6) if mom_usd is not None else None,
            "mom_pct": mom_pct,
            "pct_in_asia": pct_in_asia,
            "lifecycle": cls.get("lifecycle", "unknown"),
            "first_appears": "●" if first_appears_anywhere else "",
            "notes": "; ".join(notes),
        }

        for (col_idx, field, meta) in layout:
            cell = ws.cell(row=row, column=col_idx)
            if field == "vendor_shares" or field == "vendor_usd_musd":
                country, exch = meta
                k = (t, country, exch)
                vendor_usd = cur_aum.get(k)  # raw $
                if vendor_usd is None:
                    continue
                if field == "vendor_shares":
                    # Infer shares from vendor_usd / bbg_price (we don't yet ingest reported shares)
                    if bbg_price and bbg_price > 0:
                        shares = vendor_usd / bbg_price
                        cell.value = shares
                        cell.number_format = "#,##0"
                        cell.font = GREY_ITALIC  # inferred
                elif field == "vendor_usd_musd":
                    cell.value = vendor_usd / 1e6  # $M
                    cell.number_format = "#,##0.00"
                    # Repriced source -> italic
                    if cur_source.get(k) == "repriced":
                        cell.font = GREY_ITALIC
                    else:
                        cell.font = BLACK_REG
                    # First appearance highlight on USD cell
                    if vendor_usd > 0 and k not in prior_aum:
                        cell.fill = FILL_FIRST
                    elif zebra:
                        cell.fill = FILL_ZEBRA
                cell.border = BORDER_THIN
            else:
                val = values.get(field)
                if val is not None and val != "":
                    cell.value = val
                # Format specifics
                if field == "family":
                    cell.fill = family_fill
                    cell.font = BLACK_BOLD
                elif field == "bbg_aum_musd" or field == "asia_total_musd" or field == "prior_asia_musd" or field == "mom_musd":
                    cell.number_format = "#,##0.00;[Red]-#,##0.00"
                elif field == "bbg_price":
                    cell.number_format = "#,##0.00"
                elif field == "bbg_shares_mm":
                    cell.number_format = "#,##0.000"
                elif field == "mom_pct":
                    cell.number_format = "0.0%;[Red]-0.0%"
                elif field == "pct_in_asia":
                    cell.number_format = "0.0%"
                elif field == "first_appears" and values.get("first_appears"):
                    cell.fill = FILL_FIRST
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True, color="F57F17")
                elif field == "lifecycle":
                    if val == "pending_launch":
                        cell.fill = PatternFill("solid", fgColor="FFF9C4")
                    elif val == "delisted":
                        cell.fill = PatternFill("solid", fgColor="FFCDD2")
                # Apply zebra to non-family, non-highlighted cells
                if zebra and not cell.fill.start_color.rgb and field not in ("family",):
                    cell.fill = FILL_ZEBRA

        row += 1

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 9
    ws.column_dimensions[get_column_letter(2)].width = 14
    ws.column_dimensions[get_column_letter(3)].width = 28
    # Vendor cols narrow
    for c in range(4, 4 + sum(len(exs) for _, exs in vendor_columns_by_country) * 2):
        ws.column_dimensions[get_column_letter(c)].width = 11
    # Our side + derived cols
    for c in range(4 + sum(len(exs) for _, exs in vendor_columns_by_country) * 2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "D3"
    ws.sheet_view.showGridLines = False

    print(f"  Wrote {label}: {row - 3} fund rows, {sum(len(exs) for _, exs in vendor_columns_by_country)} exchanges, {max_col} columns")

# ── Reference sheets ──
def build_readme(wb):
    ws = wb.create_sheet("README", 0)
    lines = [
        ("REX Asia Monthly Log", True, 14),
        (f"Built: {date.today().isoformat()}", False, 10),
        ("", False, 10),
        ("Purpose", True, 12),
        ("  Per-fund per-month per-vendor ledger of Asia AUM reporting.", False, 10),
        ("  Independent of the DB and the report pipeline — a human-auditable record.", False, 10),
        ("", False, 10),
        ("Schema — SHARES are fundamental, USD is derived", True, 12),
        ("  For each vendor per month we log shares (reported where we have them, inferred as USD/BBG_price otherwise).", False, 10),
        ("  USD in this sheet is derived: shares × current-month BBG price.", False, 10),
        ("  For frozen reporters (e.g., Futu HK, Oriental Harbour), shares carry forward; USD floats with price.", False, 10),
        ("", False, 10),
        ("Sheets", True, 12),
        ("  Summary — at-a-glance dashboard for the latest report month (matches the PDFs)", False, 10),
        ("  README  — this page", False, 10),
        ("  Funds   — 96 REX tickers with lifecycle (active / delisted / pending_launch)", False, 10),
        ("  Vendors — 14 brokers with cadence, native currency, status", False, 10),
        ("  FX      — month-end FX rates from yfinance, 7 currencies x 14 months", False, 10),
        ("  YYYY-MM — one sheet per month, wide layout", False, 10),
        ("", False, 10),
        ("Monthly-sheet conventions", True, 12),
        ("  Rows sorted: Asia-active funds first (within family, by Asia Total DESC), then a separator, then non-Asia funds", False, 10),
        ("  IDENTITY (dark)  — ticker, family (colored by suite), fund name", False, 10),
        ("  Country banners (blue) — merged header over each vendor within that country. Each vendor = 2 cols: shares, USD $M", False, 10),
        ("  OUR SIDE (green) — BBG global AUM $M, price, implied global shares (M), source (data_aum or microsector)", False, 10),
        ("  DERIVED (orange) — Asia total $M, prior Asia $M, MoM $M, MoM %, % in Asia", False, 10),
        ("  FLAGS           — lifecycle, first-appears dot, notes (stale, UCITS, delisted)", False, 10),
        ("", False, 10),
        ("Formatting cues", True, 12),
        ("  Grey italic shares   = inferred from vendor USD ÷ BBG price (not reported as shares)", False, 10),
        ("  Grey italic USD      = vendor was repriced (quarterly reporter, not fresh)", False, 10),
        ("  Yellow cell          = first appearance of fund at this vendor vs prior month", False, 10),
        ("  Yellow lifecycle     = pending launch", False, 10),
        ("  Red lifecycle        = delisted", False, 10),
        ("  Zebra stripes        = every other data row for scan", False, 10),
        ("", False, 10),
        ("Precision", True, 12),
        ("  Internal values stored as floats (full precision). Display rounding via number format only.", False, 10),
    ]
    for i, (text, bold, size) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size, name="Calibri")
    ws.column_dimensions["A"].width = 120
    ws.sheet_view.showGridLines = False

def build_funds_sheet(wb, classification):
    ws = wb.create_sheet("Funds")
    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT e.ticker, e.name AS fund_name, pf.name AS family
        FROM etp e JOIN product_family pf USING (family_id)
        ORDER BY pf.name, e.ticker
    """)
    rows = cur.fetchall(); conn.close()

    headers = ["Ticker", "Fund Name", "Family", "BBG Ticker", "Suffix", "Lifecycle", "First Seen (BBG)", "Last Seen (BBG)", "Latest AUM $M", "Days Stale"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(rows, start=2):
        t = r["ticker"]; cls = classification.get(t, {})
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=r["fund_name"])
        fc = ws.cell(row=i, column=3, value=r["family"])
        fc.fill = FAMILY_FILL.get(r["family"], FAMILY_FILL["Other"])
        fc.font = BLACK_BOLD
        ws.cell(row=i, column=4, value=cls.get("bbg_ticker") or "")
        ws.cell(row=i, column=5, value=cls.get("bbg_suffix") or "")
        lifecell = ws.cell(row=i, column=6, value=cls.get("lifecycle") or "")
        if cls.get("lifecycle") == "pending_launch":
            lifecell.fill = PatternFill("solid", fgColor="FFF9C4")
        elif cls.get("lifecycle") == "delisted":
            lifecell.fill = PatternFill("solid", fgColor="FFCDD2")
        ws.cell(row=i, column=7, value=cls.get("first_seen") or "")
        ws.cell(row=i, column=8, value=cls.get("last_seen") or "")
        aum_cell = ws.cell(row=i, column=9, value=cls.get("latest_aum_musd"))
        aum_cell.number_format = "#,##0.00"
        ws.cell(row=i, column=10, value=cls.get("days_stale"))
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFGHIJ", [10, 52, 14, 20, 8, 14, 16, 16, 14, 12]):
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False
    print(f"  Funds: {len(rows)} rows")

def build_vendors_sheet(wb):
    ws = wb.create_sheet("Vendors")
    vendors = [
        ("KSD Retail",           "Korea",    "monthly",   "USD",    "USD per fund",       "active",  "Aggregate + ACE ETF side table"),
        ("KIM ACE Tesla Fund",   "Korea",    "monthly",   "KRW",    "KRW + FX -> USD",    "active",  "Single fund, from Investor webpage"),
        ("SBI",                  "Japan",    "monthly",   "JPY",    "USD (Japanese hdrs)","active",  ""),
        ("Rakuten",              "Japan",    "monthly",   "JPY",    "USD (Japanese hdrs)","active",  ""),
        ("Monex",                "Japan",    "monthly",   "JPY",    "JPY + USD + FX",     "active",  "Clean, cross-checkable"),
        ("Matsui",               "Japan",    "monthly",   "JPY",    "USD (Japanese hdrs)","active",  ""),
        ("MooMoo Japan",         "Japan",    "monthly",   "USD",    "shares + USD",       "active",  "Upgraded from quarterly in Mar26"),
        ("Futu/MooMoo HK",       "Hong Kong","quarterly", "USD",    "USD",                "FROZEN",  "Cannot send without commercial contract (21 Apr 2026)"),
        ("Oriental Harbour",     "Hong Kong","quarterly", "USD",    "13F filing",         "active",  "Q-end + 13F arrives mid-quarter+1"),
        ("SYFE",                 "Hong Kong","monthly",   "USD",    "verbal",             "active",  "No documentation; value is FEPI_LN only"),
        ("ViewTrade HK/SG/TW",   "Various",  "quarterly", "USD",    "shares + price + USD","active", "Physical hardcopy collection required"),
        ("MooMoo Singapore",     "Singapore","quarterly", "USD",    "USD",                "active",  "Waiting for Q1 2026 data"),
        ("MooMoo Malaysia",      "Malaysia", "quarterly", "USD",    "USD",                "active",  "Waiting for Q1 2026 data"),
        ("Asset Plus Thailand",  "Thailand", "monthly",   "USD",    "USD",                "ZERO",    "Sold DRNZ Apr26 — now $0"),
    ]
    headers = ["Vendor", "Country", "Cadence", "Native CCY", "Format", "Status", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")
    for i, v in enumerate(vendors, start=2):
        for j, val in enumerate(v, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if j == 6 and val in ("FROZEN", "ZERO"):
                cell.fill = PatternFill("solid", fgColor="FFCDD2")
                cell.font = Font(bold=True, color="C62828")
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFG", [24, 12, 11, 12, 24, 10, 60]):
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False
    print(f"  Vendors: {len(vendors)} rows")

def build_fx_sheet(wb, fx):
    ws = wb.create_sheet("FX")
    pairs = ["KRW", "JPY", "HKD", "SGD", "THB", "MYR", "TWD"]
    headers = ["Month"] + pairs + ["As of"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")
    row = 2
    for month in sorted(fx.keys()):
        ws.cell(row=row, column=1, value=month)
        as_of = None
        for j, p in enumerate(pairs, start=2):
            rec = fx[month].get(p)
            if rec:
                c = ws.cell(row=row, column=j, value=rec["rate"])
                c.number_format = "#,##0.0000"
                as_of = rec["as_of"]
        ws.cell(row=row, column=len(pairs) + 2, value=as_of)
        row += 1
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 10
    for i, p in enumerate(pairs, start=2):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions[get_column_letter(len(pairs) + 2)].width = 12
    ws.sheet_view.showGridLines = False
    print(f"  FX: {len(fx)} months")


def build_summary(wb):
    """At-a-glance dashboard for the latest report month. Mirrors the PDF headline
    numbers exactly by reading enriched_report_data.json (the report's own output)."""
    import json as _json
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    path = Path("enriched_report_data.json")
    if not path.exists():
        ws.cell(1, 1, "Summary unavailable - run enrich_report_data.py first").font = Font(italic=True)
        return
    d = _json.loads(path.read_text(encoding="utf-8"))
    H = d["headlines"]; funds = d["funds"]
    month = (d.get("narrative", {}) or {}).get("month_long") or d["meta"]["report_month"]
    EPI = {"AIPI", "FEPI", "CEPI", "ATCL"}
    def disp(f):
        fam, t = f["family_name"], f["ticker"]
        if t == "ULTI": return "Growth & Income"
        if t == "ATCL": return "Equity Premium Income"
        if fam == "Income": return "Equity Premium Income" if t in EPI else "Growth & Income"
        return fam
    suites = {}
    for f in funds:
        suites[disp(f)] = suites.get(disp(f), 0.0) + f["asia_aum"]
    countries = sorted(d["countries"], key=lambda c: -c["aum"])
    top = sorted(funds, key=lambda f: -f["asia_aum"])[:10]
    asia = H["total_asia_aum"]; prior = H["total_asia_aum_prior"]

    TITLE = Font(bold=True, size=16, name="Calibri")
    SUB = Font(italic=True, size=10, color="666666", name="Calibri")
    H2 = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    KLAB = Font(size=9, color="666666", name="Calibri")
    KVAL = Font(bold=True, size=14, name="Calibri")
    HDR = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    BAND = PatternFill("solid", fgColor="1a1a2e")
    M = '#,##0.0,,"M"'; P = '0.0%'
    r = 1
    ws.cell(r, 1, "REX Asia Report - Summary").font = TITLE; r += 1
    ws.cell(r, 1, month).font = SUB; r += 2
    # KPI block
    kpis = [("Total Asia AUM", asia, M), ("% of REX Global AUM", H["pct_in_asia"], P),
            ("MoM Change ($)", asia - prior, M), ("MoM Change (%)", H.get("mom_pct") or 0, P),
            ("Market Move", H["total_market_move"], M), ("Est. Net Flows", H["total_flows"], M)]
    for i, (lab, val, fmt) in enumerate(kpis):
        col = 1 + i * 2
        ws.cell(r, col, lab).font = KLAB
        c = ws.cell(r + 1, col, val); c.font = KVAL; c.number_format = fmt
    r += 3
    ws.cell(r, 1, f"REX Global AUM (Bloomberg): {H['total_global_aum']/1e9:.2f}B   |   Funds with Asia AUM: {len(funds)}").font = SUB
    r += 2
    def table(title, headers, rows, startcol=1):
        nonlocal r
        cell = ws.cell(r, startcol, title); cell.font = H2; cell.fill = BAND
        for j in range(len(headers)):
            ws.cell(r, startcol + j).fill = BAND
        r += 1
        for j, h in enumerate(headers):
            c = ws.cell(r, startcol + j, h); c.font = HDR; c.fill = PatternFill("solid", fgColor="34495e")
        r += 1
        for row in rows:
            for j, (val, fmt) in enumerate(row):
                c = ws.cell(r, startcol + j, val)
                if fmt: c.number_format = fmt
                if j == 0: c.font = Font(bold=True, size=10, name="Calibri")
            r += 1
        r += 1
    table("By Product Suite", ["Suite", "Asia AUM", "% of Asia"],
          [[(k, None), (v, M), (v / asia, P)] for k, v in sorted(suites.items(), key=lambda x: -x[1])])
    table("By Country", ["Country", "Asia AUM", "% of Asia"],
          [[(c["country"], None), (c["aum"], M), (c["aum"] / asia, P)] for c in countries])
    table("Top 10 Funds by Asia AUM", ["Ticker", "Fund", "Asia AUM", "% of Asia"],
          [[(f["ticker"], None), (f.get("fund_name") or "", None), (f["asia_aum"], M), (f["asia_aum"] / asia, P)] for f in top])
    widths = [26, 16, 14, 16, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    classification = load_classification()
    fx = load_fx()
    print("Fetching Bloomberg monthly data...")
    bbg_months = fetch_bloomberg_monthly()
    print(f"  {len(bbg_months)} months of BBG data")

    wb = Workbook(); wb.remove(wb.active)
    build_readme(wb)
    print("Reference sheets...")
    build_funds_sheet(wb, classification)
    build_vendors_sheet(wb)
    build_fx_sheet(wb, fx)
    build_summary(wb)

    conn = db(); cur = conn.cursor()
    cur.execute("SELECT month_id, month_end FROM calendar_month ORDER BY month_id")
    months = cur.fetchall(); conn.close()

    print("Monthly sheets...")
    for m in months:
        month_id = m["month_id"]; month_end = m["month_end"]
        label = month_end.strftime("%Y-%m")
        prior = month_id - 1 if month_id > 1 else None
        if prior is None: continue
        build_month_sheet(wb, label, month_id, month_end, prior, bbg_months, classification)

    wb.save(OUT)
    print(f"\nSaved: {OUT.absolute()}  ({OUT.stat().st_size / 1024:.0f} KB)")

if __name__ == "__main__":
    main()
