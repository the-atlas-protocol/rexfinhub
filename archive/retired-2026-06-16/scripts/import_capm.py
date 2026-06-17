raise SystemExit("RETIRED 2026-06-16 - quarantined to archive/retired-2026-06-16/, do not run. 3-gate proven 0 live refs; pending final sweep.")
#!/usr/bin/env python3
"""Import Capital Markets Product List from Excel into rex_products.

Reads per-suite operational sheets (T-REX, REX, REX-OSPREY, BMO) and the
ALL PRODUCTS LIST classification sheet, merges by ticker, and upserts the
CapM fields into rex_products. (capm_products was retired in Track 4a —
rex_products is the single product table.)

Usage:
    python scripts/import_capm.py
    python scripts/import_capm.py --file "path/to/file.xlsx"
    python scripts/import_capm.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime, time
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
os.chdir(PROJECT_ROOT)

import pandas as pd
from sqlalchemy import func, text

from webapp.database import init_db, SessionLocal
from webapp.models import CapMTrustAP, RexProduct

DEFAULT_FILE = Path.home() / "Downloads" / "Capital Markets Product List .xlsx"

# Suite sheets and their column mappings (column index -> field name).
# T-REX, REX, REX-OSPREY share the same layout.
STANDARD_COLS = {
    0: "fund_name",
    1: "ticker",
    2: "bb_ticker",
    3: "inception_date",
    4: "trust",
    5: "exchange",
    6: "cu_size",
    7: "fixed_fee",
    8: "variable_fee",
    9: "cut_off",
    10: "custodian",
    11: "lmm",
    12: "prospectus_link",
}

# BMO has an extra first column (BMO Suites)
BMO_COLS = {
    0: "bmo_suite",
    1: "fund_name",
    2: "ticker",
    3: "bb_ticker",
    4: "inception_date",
    5: "issuer",
    6: "exchange",
    7: "cu_size",
    8: "fixed_fee",
    9: "variable_fee",
    10: "cut_off",
    11: "custodian",
    12: "lmm",
    13: "prospectus_link",
}

ALL_PRODUCTS_COLS = {
    0: "ticker",
    1: "fund_name",
    2: "inception_date",
    3: "our_category",
    4: "product_type",
    5: "category",
    6: "sub_category",
    7: "direction",
    8: "leverage",
    9: "underlying_ticker",
    10: "underlying_name",
    11: "expense_ratio",
    12: "competitor_products",
}


def _clean(v) -> str | None:
    """Clean a cell value to string or None."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return str(v)
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, (int, float)):
        # Keep numbers that aren't NaN
        if pd.isna(v):
            return None
        return str(v)
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan", "nat"):
        return None
    return s


def _to_date(v) -> date | None:
    """Parse various date formats into a date object."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan", "nat"):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_float(v) -> float | None:
    """Parse expense ratio or other float values."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if pd.isna(v):
            return None
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan", "#value!"):
        return None
    try:
        return float(s.replace("%", "").replace(",", ""))
    except ValueError:
        return None


def read_suite_sheet(wb_path: str, sheet_name: str, col_map: dict, suite_label: str) -> dict[str, dict]:
    """Read a per-suite sheet and return {ticker: field_dict}."""
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]

    products = {}
    # Row 1 is header; data starts at row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for col_idx, field_name in col_map.items():
            if col_idx < len(row):
                record[field_name] = row[col_idx]
            else:
                record[field_name] = None

        ticker = _clean(record.get("ticker"))
        fund_name = _clean(record.get("fund_name"))
        if not ticker or not fund_name:
            continue

        # Build cleaned record
        cleaned = {
            "fund_name": fund_name,
            "ticker": ticker,
            "bb_ticker": _clean(record.get("bb_ticker")),
            "inception_date": _to_date(record.get("inception_date")),
            "trust": _clean(record.get("trust")),
            "issuer": _clean(record.get("issuer")),
            "exchange": _clean(record.get("exchange")),
            "cu_size": _clean(record.get("cu_size")),
            "fixed_fee": _clean(record.get("fixed_fee")),
            "variable_fee": _clean(record.get("variable_fee")),
            "cut_off": _clean(record.get("cut_off")) if not isinstance(record.get("cut_off"), time) else record["cut_off"].strftime("%H:%M"),
            "custodian": _clean(record.get("custodian")),
            "lmm": _clean(record.get("lmm")),
            "prospectus_link": _clean(record.get("prospectus_link")),
            "suite_source": suite_label,
            "bmo_suite": _clean(record.get("bmo_suite")),
        }

        # BMO sheets carry forward the suite name from the first column
        # Handle merged cells: if bmo_suite is None but we're in BMO, carry forward
        if suite_label == "BMO" and not cleaned["bmo_suite"]:
            # Will be filled in post-processing
            pass

        products[ticker] = cleaned

    wb.close()
    return products


def read_all_products(wb_path: str) -> dict[str, dict]:
    """Read the ALL PRODUCTS LIST sheet and return {ticker: classification_dict}."""
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb["ALL PRODUCTS LIST"]

    products = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for col_idx, field_name in ALL_PRODUCTS_COLS.items():
            if col_idx < len(row):
                record[field_name] = row[col_idx]
            else:
                record[field_name] = None

        ticker = _clean(record.get("ticker"))
        if not ticker:
            continue

        products[ticker] = {
            "our_category": _clean(record.get("our_category")),
            "product_type": _clean(record.get("product_type")),
            "category": _clean(record.get("category")),
            "sub_category": _clean(record.get("sub_category")),
            "direction": _clean(record.get("direction")),
            "leverage": _clean(record.get("leverage")),
            "underlying_ticker": _clean(record.get("underlying_ticker")),
            "underlying_name": _clean(record.get("underlying_name")),
            "expense_ratio": _to_float(record.get("expense_ratio")),
            "competitor_products": _clean(record.get("competitor_products")),
            # Also capture fund_name and inception_date as fallback
            "fund_name": _clean(record.get("fund_name")),
            "inception_date": _to_date(record.get("inception_date")),
        }

    wb.close()
    return products


def fill_bmo_suites(products: dict[str, dict], wb_path: str) -> None:
    """Post-process BMO products to fill in bmo_suite from merged cells.

    The BMO sheet has the suite name (FANG+, Gold Miners, etc.) in column A,
    but only on the first row of each group (merged cells). We need to carry
    the value forward.
    """
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb["BMO"]

    current_suite = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        suite_val = _clean(row[0]) if len(row) > 0 else None
        ticker_val = _clean(row[2]) if len(row) > 2 else None

        if suite_val:
            current_suite = suite_val

        if ticker_val and ticker_val in products:
            products[ticker_val]["bmo_suite"] = current_suite

    wb.close()


# --- Track 4a: import target is rex_products (capm_products retired) -------
# Identity columns — fill only when rex_products' value is empty.
_COND_FILLS = {
    "fund_name":           "name",
    "suite_source":        "product_suite",
    "prospectus_link":     "latest_prospectus_link",
    "competitor_products": "competitors",
}
# CapM-managed columns — the Excel is the source of truth; overwrite, except
# any field the admin has pinned via rex_products.manually_edited_fields.
_CAPM_OVERWRITE = {
    "bb_ticker":         "bb_ticker",
    "inception_date":    "inception_date",
    "trust":             "trust",
    "issuer":            "issuer",
    "exchange":          "exchange",
    "cu_size":           "cu_size",
    "fixed_fee":         "fixed_fee",
    "variable_fee":      "variable_fee",
    "cut_off":           "cut_off",
    "custodian":         "custodian",
    "lmm":               "lmm",
    "our_category":      "our_category",
    "product_type":      "product_type",
    "category":          "category",
    "sub_category":      "sub_category",
    "direction":         "direction",
    "leverage":          "leverage",
    "underlying_ticker": "underlying_ticker",
    "underlying_name":   "underlying_name",
    "expense_ratio":     "expense_ratio",
    "bmo_suite":         "bmo_suite",
}


def _empty(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def import_capm(file_path: str, dry_run: bool = False) -> dict:
    """Main import function. Returns summary stats."""
    print(f"Reading: {file_path}")

    # 1. Read per-suite operational sheets
    all_products: dict[str, dict] = {}

    suite_counts = {}
    for sheet_name, col_map, suite_label in [
        ("T-REX", STANDARD_COLS, "T-REX"),
        ("REX", STANDARD_COLS, "REX"),
        ("REX-OSPREY", STANDARD_COLS, "REX-OSPREY"),
        ("BMO", BMO_COLS, "BMO"),
    ]:
        suite_data = read_suite_sheet(file_path, sheet_name, col_map, suite_label)
        suite_counts[suite_label] = len(suite_data)
        print(f"  {suite_label}: {len(suite_data)} products")
        all_products.update(suite_data)

    # Post-process BMO suite names (merged cell carry-forward)
    fill_bmo_suites(all_products, file_path)

    # 2. Read ALL PRODUCTS LIST for classification data
    classification = read_all_products(file_path)
    print(f"  ALL PRODUCTS LIST: {len(classification)} products")

    # 3. Merge classification into operational data
    # Also add products that are ONLY in ALL PRODUCTS LIST (no suite sheet)
    for ticker, cls_data in classification.items():
        if ticker in all_products:
            # Merge classification fields into existing record
            for key in ("our_category", "product_type", "category", "sub_category",
                        "direction", "leverage", "underlying_ticker", "underlying_name",
                        "expense_ratio", "competitor_products"):
                if cls_data.get(key) is not None:
                    all_products[ticker][key] = cls_data[key]
        else:
            # Product only in ALL PRODUCTS LIST, no suite sheet data
            all_products[ticker] = {
                "fund_name": cls_data.get("fund_name", ticker),
                "ticker": ticker,
                "inception_date": cls_data.get("inception_date"),
                "suite_source": None,
                **{k: cls_data.get(k) for k in (
                    "our_category", "product_type", "category", "sub_category",
                    "direction", "leverage", "underlying_ticker", "underlying_name",
                    "expense_ratio", "competitor_products",
                )},
            }

    print(f"\n  Total unique products: {len(all_products)}")

    if dry_run:
        print("\n  [DRY RUN] No database changes made.")
        for ticker, data in sorted(all_products.items()):
            suite = data.get("suite_source") or "--"
            name = (data.get("fund_name") or "?")[:60]
            print(f"    {ticker:8s} | {suite:12s} | {name}")
        return {"total": len(all_products), "inserted": 0, "updated": 0}

    # 4. Upsert CapM fields into rex_products. capm_products was retired in
    #    Track 4a — rex_products is the single product table. Match by ticker;
    #    identity columns fill-only-if-empty, CapM-managed columns overwrite
    #    (except fields the admin pinned via manually_edited_fields).
    init_db()
    db = SessionLocal()
    updated = 0
    unmatched: list[str] = []

    try:
        for ticker, data in all_products.items():
            rex = (
                db.query(RexProduct)
                .filter(func.upper(func.trim(RexProduct.ticker)) == ticker.upper().strip())
                .first()
            )
            if rex is None:
                unmatched.append(ticker)
                continue

            try:
                edited = set(json.loads(rex.manually_edited_fields or "[]"))
            except (json.JSONDecodeError, TypeError):
                edited = set()

            changed = False
            # Identity columns — fill only when rex_products' value is empty.
            for src, dst in _COND_FILLS.items():
                v = data.get(src)
                if not _empty(v) and _empty(getattr(rex, dst, None)):
                    setattr(rex, dst, v)
                    changed = True
            # CapM-managed columns — Excel is source of truth; skip admin-pinned.
            for src, dst in _CAPM_OVERWRITE.items():
                if dst in edited:
                    continue
                v = data.get(src)
                if _empty(v):
                    continue
                if dst == "cu_size":
                    try:
                        v = int(str(v).replace(",", "").strip())
                    except (ValueError, TypeError):
                        continue
                if getattr(rex, dst, None) != v:
                    setattr(rex, dst, v)
                    changed = True

            if changed:
                rex.updated_at = datetime.utcnow()
                updated += 1

        db.commit()
        print(f"\n  rex_products rows updated: {updated}")
        if unmatched:
            print(f"  Unmatched tickers (no rex_products row — skipped): {len(unmatched)}")
            for t in sorted(unmatched):
                print(f"    {t}")
    except Exception as e:
        db.rollback()
        print(f"\n  ERROR: {e}")
        raise
    finally:
        db.close()

    return {"total": len(all_products), "updated": updated, "unmatched": len(unmatched)}


def read_trust_aps(wb_path: str) -> list[dict]:
    """Read the 'Trust & APs' sheet and return a list of {trust, ap, order} rows.

    The sheet lays trusts out horizontally: trust names sit on row 3 (1-indexed
    row 4) in columns D/F/H/J (0-indexed 3, 5, 7, 9), and each trust's APs are
    listed vertically in the same column beneath it. Blank columns (E/G/I) are
    spacer columns used for visual separation.
    """
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb["Trust & APs"]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the trust-header row: the first row that has >= 2 non-empty cells
    # past column C and whose contents look like trust names (contain "Trust" or "Fund").
    header_row_idx = None
    for i, row in enumerate(rows[:10]):
        vals = [(_clean(c), idx) for idx, c in enumerate(row) if _clean(c)]
        trust_like = [v for v, _ in vals if v and ("trust" in v.lower() or "fund" in v.lower())]
        if len(trust_like) >= 2:
            header_row_idx = i
            break

    if header_row_idx is None:
        print("  Trust & APs: could not locate header row; skipping")
        return []

    header_row = rows[header_row_idx]
    trust_columns: list[tuple[int, str]] = []
    for col_idx, cell in enumerate(header_row):
        name = _clean(cell)
        if name:
            trust_columns.append((col_idx, name))

    records: list[dict] = []
    for col_idx, trust_name in trust_columns:
        order = 0
        for row in rows[header_row_idx + 1:]:
            if col_idx >= len(row):
                continue
            ap = _clean(row[col_idx])
            if not ap:
                continue
            order += 1
            records.append({
                "trust_name": trust_name,
                "ap_name": ap,
                "sort_order": order,
            })

    return records


def import_trust_aps(file_path: str, dry_run: bool = False) -> int:
    """Import the Trust & APs sheet into capm_trust_aps. Returns rows upserted."""
    records = read_trust_aps(file_path)
    print(f"  Trust & APs: {len(records)} (trust, AP) rows parsed")

    if not records:
        return 0

    if dry_run:
        current = None
        for r in records:
            if r["trust_name"] != current:
                current = r["trust_name"]
                print(f"    {current}")
            print(f"      {r['sort_order']:2d}. {r['ap_name']}")
        return len(records)

    init_db()
    db = SessionLocal()
    try:
        # Replace strategy: clear existing rows, then insert fresh. The sheet
        # is the source of truth and is small enough (< 50 rows) that a wipe-
        # and-reload is cleaner than per-row upserts.
        db.query(CapMTrustAP).delete()
        for r in records:
            db.add(CapMTrustAP(
                trust_name=r["trust_name"],
                ap_name=r["ap_name"],
                sort_order=r["sort_order"],
            ))
        db.commit()
        print(f"  Trust & APs: wrote {len(records)} rows")
    except Exception as e:
        db.rollback()
        print(f"  Trust & APs ERROR: {e}")
        raise
    finally:
        db.close()

    return len(records)


def main():
    parser = argparse.ArgumentParser(description="Import Capital Markets Product List from Excel")
    parser.add_argument("--file", type=str, default=str(DEFAULT_FILE),
                        help="Path to the Excel file")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview without writing to database")
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"File not found: {file_path}")
        sys.exit(1)

    import_capm(str(file_path), dry_run=args.dry_run)
    import_trust_aps(str(file_path), dry_run=args.dry_run)


if __name__ == "__main__":
    main()
