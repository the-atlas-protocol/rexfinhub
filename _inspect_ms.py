"""Inspect microsector sheet Feb vs Mar values."""
import openpyxl
from pathlib import Path
import datetime as dt

OUT = open("_ms_result.txt", "w", encoding="utf-8")
def P(*args, **kwargs):
    print(*args, **kwargs, file=OUT)
    OUT.flush()

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")
P(f"File: {BB}")
P(f"Last modified: {dt.datetime.fromtimestamp(BB.stat().st_mtime)}")
P(f"Size: {BB.stat().st_size / 1024 / 1024:.1f} MB")
P()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)
ws = wb["microsector"]
tickers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
P(f"Num tickers in sheet: {sum(1 for t in tickers[1:] if t)}")
P(f"Tickers: {[t for t in tickers[1:] if t]}")
P()

# Find last rows per month
rows_by_mo = {}
for r in range(5, ws.max_row + 1):
    d = ws.cell(row=r, column=1).value
    if d and hasattr(d, "year"):
        key = (d.year, d.month)
        rows_by_mo[key] = r

feb_row = rows_by_mo.get((2026, 2))
mar_row = rows_by_mo.get((2026, 3))
apr_row = rows_by_mo.get((2026, 4))

P(f"Feb last row: {feb_row}  date: {ws.cell(row=feb_row, column=1).value if feb_row else None}")
P(f"Mar last row: {mar_row}  date: {ws.cell(row=mar_row, column=1).value if mar_row else None}")
P(f"Apr last row: {apr_row}  date: {ws.cell(row=apr_row, column=1).value if apr_row else None}")
P()

P(f"{'Ticker':<8} {'Feb 27':>14} {'Mar 31':>14} {'Apr last':>14} {'Mar chg':>10} {'Apr chg':>10}")
P("-" * 80)
for c in range(2, ws.max_column + 1):
    t = tickers[c - 1]
    if not t: continue
    feb_v = ws.cell(row=feb_row, column=c).value if feb_row else None
    mar_v = ws.cell(row=mar_row, column=c).value if mar_row else None
    apr_v = ws.cell(row=apr_row, column=c).value if apr_row else None

    def fmt(v):
        if not isinstance(v, (int, float)): return "—"
        return f"${v/1e6:,.1f}M"

    def pct(old, new):
        if not (isinstance(old, (int, float)) and isinstance(new, (int, float)) and old > 0): return "—"
        return f"{(new/old-1)*100:+.1f}%"

    print(f"{t:<8} {fmt(feb_v):>14} {fmt(mar_v):>14} {fmt(apr_v):>14} {pct(feb_v, mar_v):>10} {pct(mar_v, apr_v):>10}")

wb.close()
