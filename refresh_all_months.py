"""
Refresh etp_monthly_fund for ALL months with corrected logic.

RULES:
  - Walk back from calendar_month.month_end to the last available row in data_aum,
    data_price, and microsector INDEPENDENTLY (microsector has no weekend rows).
  - For each ticker × month:
      * If inception_dt (W1 col 5) > month_end  -> NO ROW (not launched yet)
      * If delist_date  (W1 col 23) <= month_end -> NO ROW (already delisted)
      * Else: write AUM = microsector value (raw $) if ticker is ETN,
              else data_aum value in $M * 1e6. Price from data_price.
  - DOES NOT TOUCH etp_exchange_monthly_aum (Asia positions).

SAFETY:
  - DRY RUN by default. Produces _refresh_all_months_diff.txt with every change.
  - Sanity check: Asia AUM sum BEFORE and AFTER must be identical.
  - Single transaction on apply; rollback on any error.

USAGE:
  python refresh_all_months.py              # dry-run
  python refresh_all_months.py --apply      # apply writes
"""
import os
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
import pandas as pd
import psycopg2
import psycopg2.extras

BB_PATH = r"C:\Users\RyuEl-Asmar\REX Financial LLC\REX Financial LLC - MasterFiles\MASTER Data\bloomberg_daily_file.xlsm"
DRY_RUN = "--apply" not in sys.argv

LOG_PATH = "_refresh_all_months_diff.txt"
LOG = open(LOG_PATH, "w", encoding="utf-8")
def logp(*a, **k):
    print(*a, **k); print(*a, **k, file=LOG); LOG.flush()

logp(f"{'='*80}")
logp(f"refresh_all_months — MODE: {'DRY RUN (no DB writes)' if DRY_RUN else 'APPLY'}")
logp(f"Started {date.today()}")
logp(f"{'='*80}\n")

# ── Connection + calendar ──────────────────────────────────────────────
conn = psycopg2.connect(host="localhost", port=int(os.environ.get("REX_ASIA_PORT","5433")), user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
cur.execute("SELECT month_id, month_end FROM calendar_month ORDER BY month_id")
months = cur.fetchall()
cur.execute("SELECT ticker, etp_id FROM etp")
etp_map = {r["ticker"]: r["etp_id"] for r in cur.fetchall()}

# ── Load W1 lifecycle ───────────────────────────────────────────────────
logp("Loading W1 lifecycle dates (inception + delist)...")
w1 = pd.read_excel(BB_PATH, sheet_name="w1", header=0)
w1["base"] = w1["Ticker"].astype(str).str.split().str[0]
w1["listing"] = w1["Ticker"].astype(str).str.split().str[1]
w1["db_key"] = w1.apply(lambda r: f"{r['base']}_LN" if r["listing"] == "LN" else r["base"], axis=1)

lifecycle = {}
for _, r in w1.iterrows():
    if r["db_key"] in etp_map:
        inception = pd.Timestamp(r["Inception Dt"]).date() if pd.notna(r["Inception Dt"]) else None
        delist = pd.Timestamp(r["Delist Date"]).date() if pd.notna(r["Delist Date"]) else None
        lifecycle[r["db_key"]] = {"inception": inception, "delist": delist,
                                   "mkt_status": r.get("Market Status")}

# For any DB ticker not in W1: assume active forever
for t in etp_map:
    if t not in lifecycle:
        lifecycle[t] = {"inception": None, "delist": None, "mkt_status": None}

logp(f"  W1 lifecycle loaded: {len(lifecycle)} tickers")
in_w1 = sum(1 for v in lifecycle.values() if v["mkt_status"])
logp(f"  Matched to W1: {in_w1}, assumed-active (not in W1): {len(lifecycle)-in_w1}")

# ── Bloomberg workbook ─────────────────────────────────────────────────
logp(f"\nLoading Bloomberg workbook: {BB_PATH}")
wb = openpyxl.load_workbook(BB_PATH, data_only=True, read_only=True)

ws_aum = wb["data_aum"]
aum_header = list(next(ws_aum.iter_rows(min_row=1, max_row=1, values_only=True)))
ticker_to_col = {}
for i, col in enumerate(aum_header[1:], 1):
    if not col: continue
    _p = str(col).split()
    if len(_p) < 2 or _p[1] not in ("US","LN"): continue
    parts = str(col).split()
    key = f"{parts[0]}_LN" if parts[1] == "LN" else parts[0]
    ticker_to_col[key] = i
aum_rows_by_date = {}
for row in ws_aum.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, "year"):
        aum_rows_by_date[d.date() if hasattr(d, "date") else d] = row

ws_price = wb["data_price"]
price_header = list(next(ws_price.iter_rows(min_row=1, max_row=1, values_only=True)))
price_ticker_cols = {}
for i, col in enumerate(price_header[1:], 1):
    if not col: continue
    _p = str(col).split()
    if len(_p) < 2 or _p[1] not in ("US","LN"): continue
    parts = str(col).split()
    key = f"{parts[0]}_LN" if parts[1] == "LN" else parts[0]
    price_ticker_cols[key] = i
price_rows_by_date = {}
for row in ws_price.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, "year"):
        price_rows_by_date[d.date() if hasattr(d, "date") else d] = row

# data_nav — primary source for price_usd (data_price has known corruption issues e.g. Mar 31 2026)
ws_nav = wb["data_nav"]
nav_header = list(next(ws_nav.iter_rows(min_row=1, max_row=1, values_only=True)))
nav_ticker_cols = {}
for i, col in enumerate(nav_header[1:], 1):
    if not col: continue
    _p = str(col).split()
    if len(_p) < 2 or _p[1] not in ("US","LN"): continue
    parts = str(col).split()
    key = f"{parts[0]}_LN" if parts[1] == "LN" else parts[0]
    nav_ticker_cols[key] = i
nav_rows_by_date = {}
for row in ws_nav.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, "year"):
        nav_rows_by_date[d.date() if hasattr(d, "date") else d] = row
logp(f"  data_nav: {len(nav_ticker_cols)} tickers, {len(nav_rows_by_date)} dates")

ws_ms = wb["microsector_aum"]
ms_tickers_row = list(next(ws_ms.iter_rows(min_row=4, max_row=4, values_only=True)))
ms_ticker_cols = {}
for i, t in enumerate(ms_tickers_row[1:], 1):
    if t and isinstance(t, str):
        ms_ticker_cols[t.strip()] = i
ms_rows_by_date = {}
for row in ws_ms.iter_rows(min_row=5, values_only=True):
    d = row[0]
    if d and hasattr(d, "year"):
        ms_rows_by_date[d.date() if hasattr(d, "date") else d] = row

def walk_back(by_date: dict, target: date, max_days: int = 10) -> date | None:
    cur_d = target
    for _ in range(max_days):
        if cur_d in by_date: return cur_d
        cur_d -= timedelta(days=1)
    return None

# ── Snapshot current Asia sums per month (must not change) ─────────────
cur.execute("""
    SELECT month_id, COALESCE(SUM(exchange_aum_usd), 0) AS asia_total
    FROM etp_exchange_monthly_aum GROUP BY month_id ORDER BY month_id
""")
asia_before = {r["month_id"]: float(r["asia_total"]) for r in cur.fetchall()}

# ── Compute desired state per (ticker, month) ──────────────────────────
desired = {}   # {(month_id, etp_id): {"price": x, "aum": y, "source": s}}
per_month_meta = []

for m in months:
    mid = m["month_id"]; cal_end = m["month_end"]
    aum_date = walk_back(aum_rows_by_date, cal_end)
    price_date = walk_back(price_rows_by_date, cal_end)
    nav_date = walk_back(nav_rows_by_date, cal_end)
    ms_date = walk_back(ms_rows_by_date, cal_end)
    per_month_meta.append({"mid": mid, "cal_end": cal_end, "dow": cal_end.strftime("%a"),
                           "aum_date": aum_date, "price_date": price_date,
                           "nav_date": nav_date, "ms_date": ms_date})

    aum_row = aum_rows_by_date.get(aum_date) if aum_date else None
    price_row = price_rows_by_date.get(price_date) if price_date else None
    nav_row = nav_rows_by_date.get(nav_date) if nav_date else None
    ms_row = ms_rows_by_date.get(ms_date) if ms_date else None

    for ticker, etp_id in etp_map.items():
        lc = lifecycle[ticker]
        inception = lc["inception"]; delist = lc["delist"]
        # Lifecycle filter
        if inception is not None and inception > cal_end:
            continue  # not launched yet → no row
        if delist is not None and delist <= cal_end:
            continue  # already delisted → no row

        # Price: prefer NAV (data_price had Mar 31 2026 corruption); fallback to data_price
        new_price = None
        price_src = None
        if nav_row is not None and ticker in nav_ticker_cols:
            v = nav_row[nav_ticker_cols[ticker]]
            if isinstance(v, (int, float)) and v > 0:
                new_price = float(v); price_src = f"nav({nav_date})"
        if new_price is None and price_row is not None and ticker in price_ticker_cols:
            v = price_row[price_ticker_cols[ticker]]
            if isinstance(v, (int, float)) and v > 0:
                new_price = float(v); price_src = f"price({price_date})"

        # AUM (microsector overwrite priority for ETNs)
        new_aum = None; source = None
        if ticker in ms_ticker_cols and ms_row is not None:
            v = ms_row[ms_ticker_cols[ticker]]
            if isinstance(v, (int, float)) and v > 0:
                new_aum = float(v); source = f"microsector({ms_date})"
        if new_aum is None and aum_row is not None and ticker in ticker_to_col:
            v = aum_row[ticker_to_col[ticker]]
            if isinstance(v, (int, float)) and v > 0:
                new_aum = float(v) * 1_000_000; source = f"data_aum({aum_date})"

        # Only keep if we have at least AUM (price may be missing for some tickers e.g. ETN)
        if new_aum is not None:
            desired[(mid, etp_id)] = {"price": new_price, "aum": new_aum, "source": source,
                                       "ticker": ticker}

# ── Compare desired vs existing ────────────────────────────────────────
cur.execute("SELECT etp_id, month_id, price_usd, total_aum_usd FROM etp_monthly_fund")
existing = {(r["month_id"], r["etp_id"]): (
    float(r["price_usd"]) if r["price_usd"] else None,
    float(r["total_aum_usd"]) if r["total_aum_usd"] else None
) for r in cur.fetchall()}

id_to_ticker = {v: k for k, v in etp_map.items()}

writes = []    # (mid, etp_id, new_price, new_aum)
deletes = []   # (mid, etp_id) — exists but shouldn't (delisted or pre-launch)
unchanged = 0

for (mid, etp_id), ex in existing.items():
    if (mid, etp_id) not in desired:
        deletes.append((mid, etp_id))

for (mid, etp_id), d in desired.items():
    old_price, old_aum = existing.get((mid, etp_id), (None, None))
    chg = False
    if d["price"] is not None and (old_price is None or abs(d["price"] - old_price) > 0.0001):
        chg = True
    if d["aum"] is not None and (old_aum is None or abs(d["aum"] - old_aum) > 100):
        chg = True
    if chg:
        writes.append({"mid": mid, "etp_id": etp_id, "ticker": d["ticker"],
                       "old_price": old_price, "new_price": d["price"],
                       "old_aum": old_aum, "new_aum": d["aum"], "source": d["source"]})
    else:
        unchanged += 1

# ── Report ─────────────────────────────────────────────────────────────
logp(f"\nMONTH META:")
logp(f"{'mid':<4} {'cal_end':<11} {'dow':<4} {'aum_src':<11} {'ms_src':<11} {'px_src':<11}")
logp("-" * 60)
for m in per_month_meta:
    logp(f"{m['mid']:<4} {str(m['cal_end']):<11} {m['dow']:<4} "
         f"{str(m['aum_date']):<11} {str(m['ms_date']):<11} {str(m['price_date']):<11}")

logp(f"\nCHANGE COUNTS:")
logp(f"  Writes (insert or update):  {len(writes)}")
logp(f"  Deletes (delisted/pre-launch rows to remove): {len(deletes)}")
logp(f"  Unchanged:                  {unchanged}")

# Per-month before/after totals
logp(f"\n{'mid':<4} {'cal_end':<11} {'dow':<4} {'# funds (old→new)':<20} {'old $M':>12} {'new $M':>12} {'delta $M':>12}")
logp("-" * 90)
for m in per_month_meta:
    mid = m["mid"]
    cur.execute("SELECT COUNT(*) n, COALESCE(SUM(total_aum_usd),0) t FROM etp_monthly_fund WHERE month_id = %s", (mid,))
    cr = cur.fetchone(); old_n = cr["n"]; old_t = float(cr["t"])
    # Simulate new: start from existing then apply writes + deletes
    sim = {}
    for (emid, eid), (p, a) in existing.items():
        if emid == mid: sim[eid] = a or 0
    for d_mid, d_eid in deletes:
        if d_mid == mid: sim.pop(d_eid, None)
    for w in writes:
        if w["mid"] == mid: sim[w["etp_id"]] = w["new_aum"] or 0
    new_n = len(sim); new_t = sum(sim.values())
    logp(f"{mid:<4} {str(m['cal_end']):<11} {m['dow']:<4} "
         f"{old_n:>4} → {new_n:<4}          "
         f"${old_t/1e6:>10.2f} ${new_t/1e6:>10.2f} {(new_t-old_t)/1e6:>+11.2f}")

# Deletes detail per month
logp(f"\nDELETES (per month):")
del_by_mid = {}
for mid, eid in deletes:
    del_by_mid.setdefault(mid, []).append(id_to_ticker.get(eid, f"etp{eid}"))
for mid in sorted(del_by_mid.keys()):
    tickers = sorted(del_by_mid[mid])
    cal_end = next(m["cal_end"] for m in per_month_meta if m["mid"] == mid)
    logp(f"  Month {mid} ({cal_end}): {len(tickers)} deletes -> {tickers}")

# Top writes per month
logp(f"\nTOP 5 WRITES PER MONTH BY AUM DELTA:")
writes_by_mid = {}
for w in writes:
    writes_by_mid.setdefault(w["mid"], []).append(w)
for mid in sorted(writes_by_mid.keys()):
    ws_list = writes_by_mid[mid]
    ws_list.sort(key=lambda w: abs((w["new_aum"] or 0) - (w["old_aum"] or 0)), reverse=True)
    cal_end = next(m["cal_end"] for m in per_month_meta if m["mid"] == mid)
    logp(f"\n  Month {mid} ({cal_end}): {len(ws_list)} writes")
    for w in ws_list[:5]:
        old = (w["old_aum"] or 0) / 1e6; new = (w["new_aum"] or 0) / 1e6
        logp(f"    {w['ticker']:<10}  ${old:>10.2f}M → ${new:>10.2f}M  ({new-old:+.2f}M)  [{w['source']}]")

# ── SAFETY CHECK: simulate Asia sum invariance ─────────────────────────
logp(f"\n{'='*80}")
logp(f"SAFETY CHECK — Asia AUM (etp_exchange_monthly_aum) must NOT change:")
for mid, before in asia_before.items():
    logp(f"  Month {mid}: Asia=${before/1e6:.2f}M (will remain unchanged — we only touch etp_monthly_fund)")

# ── Apply ──────────────────────────────────────────────────────────────
if DRY_RUN:
    logp(f"\n{'='*80}")
    logp(f"DRY RUN — no DB writes performed.")
    logp(f"Log: {LOG_PATH}")
    logp(f"To apply: python refresh_all_months.py --apply")
    logp(f"Rollback backup: rex_asia_pre_month_refresh_20260423_1405.backup")
else:
    logp(f"\n{'='*80}")
    logp(f"APPLYING — single transaction, {len(writes)} writes + {len(deletes)} deletes")
    try:
        cur2 = conn.cursor()
        # Deletes first
        for mid, eid in deletes:
            cur2.execute("DELETE FROM etp_monthly_fund WHERE month_id = %s AND etp_id = %s", (mid, eid))
        # Upserts — handle NOT NULL on price_usd for inserts where we don't have a new price
        # Strategy: INSERT requires a price (fallback 0 if truly unknown); UPDATE uses COALESCE to preserve existing
        skipped_no_price = []
        for w in writes:
            # Check if row exists
            cur2.execute("SELECT price_usd FROM etp_monthly_fund WHERE etp_id=%s AND month_id=%s",
                         (w["etp_id"], w["mid"]))
            existing = cur2.fetchone()
            new_price = w["new_price"]
            if existing is None:
                # INSERT path — need a price. If none from Bloomberg, skip this write.
                if new_price is None:
                    skipped_no_price.append((w["ticker"], w["mid"]))
                    continue
                price = Decimal(str(round(new_price, 6)))
                aum = Decimal(str(round(w["new_aum"], 4))) if w["new_aum"] is not None else None
                cur2.execute("""
                    INSERT INTO etp_monthly_fund (etp_id, month_id, price_usd, total_aum_usd)
                    VALUES (%s, %s, %s, %s)
                """, (w["etp_id"], w["mid"], price, aum))
            else:
                # UPDATE path — preserve existing price if no new price
                price = Decimal(str(round(new_price, 6))) if new_price is not None else None
                aum = Decimal(str(round(w["new_aum"], 4))) if w["new_aum"] is not None else None
                cur2.execute("""
                    UPDATE etp_monthly_fund
                    SET price_usd = COALESCE(%s, price_usd),
                        total_aum_usd = COALESCE(%s, total_aum_usd)
                    WHERE etp_id = %s AND month_id = %s
                """, (price, aum, w["etp_id"], w["mid"]))
        if skipped_no_price:
            logp(f"  Skipped {len(skipped_no_price)} inserts (no price data): {skipped_no_price[:10]}...")

        # Post-apply Asia sanity check WITHIN THE TRANSACTION
        cur2.execute("""SELECT month_id, COALESCE(SUM(exchange_aum_usd),0) t
                        FROM etp_exchange_monthly_aum GROUP BY month_id""")
        asia_after = {r["month_id"]: float(r["t"]) for r in cur2.fetchall()}
        mismatches = []
        for mid in set(asia_before) | set(asia_after):
            b = asia_before.get(mid, 0); a = asia_after.get(mid, 0)
            if abs(a - b) > 0.01:
                mismatches.append((mid, b, a))
        if mismatches:
            raise RuntimeError(f"ASIA DATA MUTATED — rolling back. Mismatches: {mismatches}")
        logp(f"  Asia sanity check passed — unchanged across all months.")
        conn.commit()
        logp(f"  COMMITTED {len(writes)} writes and {len(deletes)} deletes.")
    except Exception as e:
        conn.rollback()
        logp(f"  ERROR — transaction rolled back: {e}")
        raise

    logp(f"\nROLLBACK (if needed):")
    logp(f"  pg_restore -h localhost -p 5433 -U postgres -c -d rex_asia \\")
    logp(f"    C:/Projects/rex-asia/rex_asia_pre_month_refresh_20260423_1405.backup")
    logp(f"{'='*80}")

wb.close()
conn.close()
LOG.close()
