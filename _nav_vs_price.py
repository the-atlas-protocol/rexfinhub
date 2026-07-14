"""Compare data_nav vs data_price for our REX universe at Mar 31 2026.
Quantify the impact on shares-invariant repricing for ETF positions
at frozen/quarterly vendors (Futu HK, OH, ViewTrade, MooMoo SG/MY).
"""
import openpyxl
from pathlib import Path
import psycopg2
import psycopg2.extras
from datetime import date

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")

OUT = open("_nav_impact.txt", "w", encoding="utf-8")
def L(*a, **k): print(*a, **k); print(*a, **k, file=OUT); OUT.flush()

# ── Load Bloomberg sheets ──────────────────────────────────────────
wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

# data_price — market close
ws_p = wb["data_price"]
p_header = list(next(ws_p.iter_rows(min_row=1, max_row=1, values_only=True)))
p_cols = {}
for i, c in enumerate(p_header[1:], 1):
    if c and " Equity" in str(c):
        parts = str(c).split()
        key = f"{parts[0]}_LN" if parts[1] == "LN" else parts[0]
        p_cols[key] = i

# data_nav — NAV per share
ws_n = wb["data_nav"]
n_header = list(next(ws_n.iter_rows(min_row=1, max_row=1, values_only=True)))
n_cols = {}
for i, c in enumerate(n_header[1:], 1):
    if c and " Equity" in str(c):
        parts = str(c).split()
        key = f"{parts[0]}_LN" if parts[1] == "LN" else parts[0]
        n_cols[key] = i
L(f"data_price has {len(p_cols)} tickers")
L(f"data_nav   has {len(n_cols)} tickers")

# Find Mar 31 2026 row in each
def get_row(ws, target):
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d and hasattr(d, 'year') and d.date() == target if hasattr(d, 'date') else d == target:
            return row
    return None

target = date(2026, 3, 31)
p_row = None
for row in ws_p.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year') and d.year == 2026 and d.month == 3 and d.day == 31:
        p_row = row; break
n_row = None
for row in ws_n.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year') and d.year == 2026 and d.month == 3 and d.day == 31:
        n_row = row; break
L(f"Found Mar 31 row in data_price: {p_row is not None}, data_nav: {n_row is not None}")
L("")

# Microsector tickers (ETN)
ws_ms = wb["microsector"]
ms_tickers_row = list(next(ws_ms.iter_rows(min_row=4, max_row=4, values_only=True)))
ms_tickers = set(str(t).strip() for t in ms_tickers_row[1:] if t and isinstance(t, str))
L(f"MicroSector ETN tickers: {len(ms_tickers)}")

# DB tickers
conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
cur.execute("SELECT ticker FROM etp")
rex_tickers = set(r["ticker"] for r in cur.fetchall())
L(f"REX universe: {len(rex_tickers)}")

# ── Per-ticker price vs NAV comparison ─────────────────────────────
L("")
L("=== TICKER-LEVEL: data_price vs data_nav (Mar 31 2026) ===")
L(f"{'Ticker':<8} {'is_ETN':<7} {'price':>10} {'NAV':>10} {'diff':>10} {'%diff':>8}")
L("-" * 70)
gaps = []
for t in sorted(rex_tickers):
    is_etn = t in ms_tickers
    p_col = p_cols.get(t)
    n_col = n_cols.get(t)
    p_val = p_row[p_col] if p_col and p_row and isinstance(p_row[p_col], (int, float)) else None
    n_val = n_row[n_col] if n_col and n_row and isinstance(n_row[n_col], (int, float)) else None
    if p_val and n_val:
        diff = n_val - p_val
        pct = diff / p_val * 100 if p_val else 0
        gaps.append({"ticker": t, "is_etn": is_etn, "price": p_val, "nav": n_val, "diff": diff, "pct": pct})
        if abs(pct) > 0.05 or t in ms_tickers:  # show all ETNs and any >5bp gap
            L(f"{t:<8} {'ETN' if is_etn else 'ETF':<7} {p_val:>10.4f} {n_val:>10.4f} {diff:>+10.4f} {pct:>+7.2f}%")

# Summary stats
etfs = [g for g in gaps if not g["is_etn"]]
etns = [g for g in gaps if g["is_etn"]]
L("")
L(f"=== SUMMARY ===")
L(f"ETFs with both price+NAV: {len(etfs)}")
L(f"  median |%diff|: {sorted([abs(g['pct']) for g in etfs])[len(etfs)//2]:.3f}%")
L(f"  max |%diff|: {max([abs(g['pct']) for g in etfs] or [0]):.3f}%")
L(f"  ETFs with |diff| >0.5%: {sum(1 for g in etfs if abs(g['pct']) > 0.5)}")

L(f"ETNs with both price+NAV: {len(etns)}")
if etns:
    L(f"  median |%diff|: {sorted([abs(g['pct']) for g in etns])[len(etns)//2]:.3f}%")
    L(f"  max |%diff|: {max([abs(g['pct']) for g in etns]):.3f}%")
    L(f"  ETNs with |diff| >1%: {sum(1 for g in etns if abs(g['pct']) > 1)}")

# ── DOLLAR IMPACT on shares-invariant reprice ──────────────────────
# For each fund at a frozen/quarterly vendor, recompute Mar AUM:
#   current method: prior_aum / prior_price × cur_price
#   proposed: ETFs use NAV instead of price (both prior and current); ETNs keep price
L("")
L("=== DOLLAR IMPACT ON FROZEN/QUARTERLY VENDOR POSITIONS ===")
# Get prior price/NAV for each fund
prior_p_row = None
prior_n_row = None
for row in ws_p.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year') and d.year == 2026 and d.month == 2 and d.day == 27:  # last Feb trading day
        prior_p_row = row; break
for row in ws_n.iter_rows(min_row=2, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year') and d.year == 2026 and d.month == 2 and d.day == 27:
        prior_n_row = row; break
L(f"Prior (Feb 27): price_row={prior_p_row is not None}, nav_row={prior_n_row is not None}")

# Pull frozen/quarterly positions from DB Feb (month_id 13)
FROZEN_EXCHANGE_IDS = [9, 10, 11, 13, 15, 16, 17]  # Futu HK, OH, MooMoo SG, MooMoo MY, ViewTrade HK/SG/TW
cur.execute("""
    SELECT e.ticker, ex.name ex_name, c.name country,
           m.exchange_aum_usd feb_aum, mf.price_usd feb_price_db
    FROM etp_exchange_monthly_aum m JOIN etp e USING (etp_id)
    JOIN exchange ex USING (exchange_id) JOIN country c USING (country_id)
    JOIN etp_monthly_fund mf ON mf.etp_id = m.etp_id AND mf.month_id = m.month_id
    WHERE m.exchange_id = ANY(%s) AND m.month_id = 13
    AND m.exchange_aum_usd > 0 AND mf.price_usd > 0
""", (FROZEN_EXCHANGE_IDS,))
positions = cur.fetchall()
L(f"Stale/frozen Feb positions: {len(positions)}")

total_old = 0; total_new = 0
etf_old = 0; etf_new = 0
etn_old = 0; etn_new = 0
big_movers = []
for r in positions:
    t = r["ticker"]
    feb_aum = float(r["feb_aum"])
    is_etn = t in ms_tickers
    p_col = p_cols.get(t); n_col = n_cols.get(t)
    cur_price = p_row[p_col] if p_col and p_row and isinstance(p_row[p_col], (int, float)) else None
    cur_nav = n_row[n_col] if n_col and n_row and isinstance(n_row[n_col], (int, float)) else None
    prior_price = prior_p_row[p_col] if p_col and prior_p_row and isinstance(prior_p_row[p_col], (int, float)) else None
    prior_nav = prior_n_row[n_col] if n_col and prior_n_row and isinstance(prior_n_row[n_col], (int, float)) else None
    if not (cur_price and prior_price):
        continue

    # OLD: shares = feb_aum / prior_price; new_aum = shares × cur_price
    old_aum = feb_aum * (cur_price / prior_price)

    # NEW: ETN uses price, ETF uses NAV (if available, else price fallback)
    if is_etn:
        new_aum = feb_aum * (cur_price / prior_price)  # same as old
        etn_old += old_aum; etn_new += new_aum
    else:
        if cur_nav and prior_nav:
            new_aum = feb_aum * (cur_nav / prior_nav)
        else:
            new_aum = old_aum  # no NAV — fallback to price
        etf_old += old_aum; etf_new += new_aum

    total_old += old_aum
    total_new += new_aum
    diff = new_aum - old_aum
    if abs(diff) > 50_000:
        big_movers.append((t, r["country"], r["ex_name"], old_aum, new_aum, diff))

L("")
L(f"OLD method (all use market close price):")
L(f"  Total stale-repriced: ${total_old/1e6:.2f}M")
L(f"NEW method (ETNs use price, ETFs use NAV):")
L(f"  Total stale-repriced: ${total_new/1e6:.2f}M")
L(f"  Δ vs old: ${(total_new - total_old)/1e6:+.2f}M")
L("")
L(f"Breakdown:")
L(f"  ETFs OLD: ${etf_old/1e6:.2f}M  NEW (NAV): ${etf_new/1e6:.2f}M  Δ ${(etf_new - etf_old)/1e6:+.2f}M")
L(f"  ETNs OLD: ${etn_old/1e6:.2f}M  NEW (same): ${etn_new/1e6:.2f}M  Δ ${(etn_new - etn_old)/1e6:+.2f}M")
L("")
L(f"Big movers (|diff| > $50K):")
for t, country, ex, old, new, diff in sorted(big_movers, key=lambda x: -abs(x[5]))[:20]:
    L(f"  {t:<8} {country:<11} {ex[:25]:<25}  ${old/1e6:>7.2f}M → ${new/1e6:>7.2f}M  ({diff/1e6:+.3f}M)")

# Coverage check
L("")
L("=== COVERAGE: which REX tickers are in data_nav vs data_price? ===")
in_p_only = [t for t in rex_tickers if t in p_cols and t not in n_cols]
in_n_only = [t for t in rex_tickers if t in n_cols and t not in p_cols]
in_both = [t for t in rex_tickers if t in n_cols and t in p_cols]
in_neither = [t for t in rex_tickers if t not in n_cols and t not in p_cols]
L(f"  In both: {len(in_both)}")
L(f"  In data_price only: {len(in_p_only)}  -> {sorted(in_p_only)[:10]}{'...' if len(in_p_only) > 10 else ''}")
L(f"  In data_nav only: {len(in_n_only)}  -> {sorted(in_n_only)[:10]}")
L(f"  In neither: {len(in_neither)}  -> {sorted(in_neither)}")

# ETNs not in data_nav (expected — Ryu's point about price indexes)
etn_not_in_nav = [t for t in rex_tickers if t in ms_tickers and t not in n_cols]
L(f"  ETNs not in data_nav (expected): {len(etn_not_in_nav)}  -> {etn_not_in_nav}")

wb.close(); conn.close(); OUT.close()
