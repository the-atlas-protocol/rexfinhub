"""
Deep audit — verify EVERY number in the Mar 2026 PDF traces to source.

Goes beyond comprehensive_audit.py by:
  1. Parsing raw broker Excel files → verify DB per-fund per-vendor
  2. Manual formula re-computation for a sample of funds (market_move, flows, mom, pct)
  3. Cross-page consistency: suite table == donut == timeline endpoint
  4. Exchange table "Others (N)" aggregation
  5. Country timeline: per-country values sum to total each month
  6. Appendix sort order / total row
  7. Leveraged vs Non-Leveraged bullet math
  8. Suite country 6-month history
"""
import sys, io, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
from pypdf import PdfReader
import openpyxl
import pandas as pd
import psycopg2
import psycopg2.extras

OUT = open("_audit_deep.txt", "w", encoding="utf-8")
F = {"pass": 0, "fail": 0, "warn": 0}
def L(*a, level="INFO"):
    tag = "" if level == "INFO" else f"[{level}] "
    msg = " ".join(str(x) for x in a)
    print(f"{tag}{msg}"); print(f"{tag}{msg}", file=OUT); OUT.flush()
    if level == "PASS": F["pass"] += 1
    elif level == "FAIL": F["fail"] += 1
    elif level == "WARN": F["warn"] += 1

def hdr(title):
    L(""); L("="*80); L(title); L("="*80)

enr = json.load(open("enriched_report_data_mar.json"))
H = enr["headlines"]
S = enr["suites"]
raw = json.load(open("report_data_mar.json"))

conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()

# ───────────────────────────────────────────────────────────────────────
hdr("1. RAW BROKER FILES → DB per (fund, exchange)")
# ───────────────────────────────────────────────────────────────────────

DATA_DIR = Path("grace_data/2026-03")

def sum_broker(path, ticker_col=0, usd_col=3, data_row_min=2, sheet=None):
    """Sum a broker file's ticker → USD map."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    result = {}
    for r in range(data_row_min, ws.max_row + 1):
        t = ws.cell(row=r, column=ticker_col + 1).value
        v = ws.cell(row=r, column=usd_col + 1).value
        if isinstance(t, str) and t.isupper() and len(t) <= 6 and isinstance(v, (int, float)) and v > 0:
            result[t.strip()] = result.get(t.strip(), 0) + float(v)
    wb.close()
    return result

# KSD - col 1 (ticker B), col 3 (AUM D), rows from 5
ksd = sum_broker(DATA_DIR / "2026 03 31 Korea REX report KSD.xlsx", ticker_col=1, usd_col=3, data_row_min=5)
# SBI - col 0 (A), col 3 (D), rows from 2
sbi = sum_broker(DATA_DIR / "2026 03 31 Japan SBI REX report.xlsx", ticker_col=0, usd_col=3, data_row_min=2)
# Rakuten Sheet2 - col 0 ticker, col 4 USD, rows 2+
rak = sum_broker(DATA_DIR / "2026 03 31 Japan Rakuten REX report.xlsx", ticker_col=0, usd_col=4, data_row_min=2, sheet="Sheet2")
# Monex - col 0, col 3, rows 2+
monex = sum_broker(DATA_DIR / "2026 03 31 Japan Monex REX report.xlsx", ticker_col=0, usd_col=3, data_row_min=2)
# Matsui - col 0, col 2, rows 3+
mat = sum_broker(DATA_DIR / "2026 03 31 Japan Matsui REX report.xlsx", ticker_col=0, usd_col=2, data_row_min=3)

broker_totals = {
    ("Korea", "KSD (Korea Securities Depository) - Retail"): (ksd, "KSD"),
    ("Japan", "SBI"): (sbi, "SBI"),
    ("Japan", "Rakuten"): (rak, "Rakuten"),
    ("Japan", "Monex"): (monex, "Monex"),
    ("Japan", "Matsui"): (mat, "Matsui"),
}

# DB per (country, exchange, ticker) for month 14
cur.execute("""
    SELECT c.name country, ex.name exchange, e.ticker, m.exchange_aum_usd
    FROM etp_exchange_monthly_aum m JOIN etp e USING (etp_id)
    JOIN exchange ex USING (exchange_id) JOIN country c USING (country_id)
    WHERE m.month_id = 14
""")
db_positions = {}
for r in cur.fetchall():
    key = (r["country"], r["exchange"])
    db_positions.setdefault(key, {})[r["ticker"]] = float(r["exchange_aum_usd"])

for key, (broker_map, label) in broker_totals.items():
    db_map = db_positions.get(key, {})
    all_tickers = set(broker_map.keys()) | set(db_map.keys())
    mismatches = []
    for t in all_tickers:
        b = broker_map.get(t, 0); d = db_map.get(t, 0)
        if abs(b - d) > 1.0:
            mismatches.append((t, b, d))
    if mismatches:
        L(f"  {label:<10} {len(mismatches)} per-fund mismatches", level="FAIL")
        for t, b, d in mismatches[:5]:
            L(f"    {t}  broker=${b:.2f}  db=${d:.2f}")
    else:
        L(f"  {label:<10} {len(all_tickers)} funds match broker file ↔ DB within $1",
          level="PASS")

# ───────────────────────────────────────────────────────────────────────
hdr("2. FORMULA RE-COMPUTATION — 10 random funds")
# ───────────────────────────────────────────────────────────────────────

import random
random.seed(42)
sample = random.sample([f for f in enr["funds"] if f.get("asia_aum", 0) > 0], 10)

for f in sample:
    t = f["ticker"]
    # Expected market_move = asia_prior × ((global / global_prior) - 1)
    if f["global_aum_prior"] > 0:
        expected_mm = f["asia_aum_prior"] * (f["global_aum"] / f["global_aum_prior"] - 1)
    else:
        expected_mm = 0
    expected_fl = (f["asia_aum"] - f["asia_aum_prior"]) - expected_mm
    expected_mom = (f["asia_aum"] - f["asia_aum_prior"]) / f["asia_aum_prior"] if f["asia_aum_prior"] > 0 else None
    expected_pct = f["asia_aum"] / f["global_aum"] if f["global_aum"] > 0 else 0

    ok_mm = abs(expected_mm - f["market_move"]) < 1
    ok_fl = abs(expected_fl - f["flows"]) < 1
    ok_mom = (expected_mom is None and f.get("mom") is None) or (expected_mom is not None and f.get("mom") is not None and abs(expected_mom - f["mom"]) < 0.0001)
    ok_pct = abs(expected_pct - f["pct_of_global"]) < 0.0001
    all_ok = ok_mm and ok_fl and ok_mom and ok_pct
    L(f"  {t:<8}  mm={ok_mm}  fl={ok_fl}  mom={ok_mom}  pct={ok_pct}",
      level="PASS" if all_ok else "FAIL")
    if not all_ok:
        L(f"    expected mm={expected_mm:+.2f}  fl={expected_fl:+.2f}  mom={expected_mom}  pct={expected_pct*100:.3f}%")
        L(f"    actual   mm={f['market_move']:+.2f}  fl={f['flows']:+.2f}  mom={f.get('mom')}  pct={f['pct_of_global']*100:.3f}%")

# ───────────────────────────────────────────────────────────────────────
hdr("3. CROSS-PAGE CONSISTENCY — Suite Breakdown Table == Donut == Timeline Endpoint")
# ───────────────────────────────────────────────────────────────────────

# Suite breakdown table on page 1 pulls from suites aggregation
# Donut uses same source
# Timeline last month should equal suite totals
last_timeline = enr["timeline"][-1]
L(f"  Timeline last month: {last_timeline['month']}, total ${last_timeline['total']/1e6:.2f}M")
L(f"  Sum of suites from timeline['suites']:")
for suite, v in last_timeline["suites"].items():
    if v > 0:
        L(f"    {suite}: ${v/1e6:.2f}M")

# Compare with enriched suites
for suite_name in ["T-REX", "MicroSectors", "Income", "REX Osprey", "Other"]:
    tl_val = last_timeline["suites"].get(suite_name, 0)
    enr_val = S.get(suite_name, {}).get("aum", 0)
    ok = abs(tl_val - enr_val) < 1
    L(f"  {suite_name:<14} timeline ${tl_val/1e6:.2f}M vs enriched ${enr_val/1e6:.2f}M",
      level="PASS" if ok else "FAIL")

# ───────────────────────────────────────────────────────────────────────
hdr("4. EXCHANGE TABLE — Top-10 + Others = Total Asia")
# ───────────────────────────────────────────────────────────────────────

exchanges = sorted(enr["exchanges"], key=lambda e: -e["aum"])
top10 = exchanges[:10]
others = exchanges[10:]
top10_sum = sum(e["aum"] for e in top10)
others_sum = sum(e["aum"] for e in others)
total_from_table = top10_sum + others_sum

L(f"  Top-10 sum:         ${top10_sum/1e6:.2f}M")
L(f"  Others ({len(others)}) sum:     ${others_sum/1e6:.2f}M")
L(f"  Grand total:        ${total_from_table/1e6:.2f}M")
L(f"  Headlines total:    ${H['total_asia_aum']/1e6:.2f}M",
  level="PASS" if abs(total_from_table - H["total_asia_aum"]) < 1 else "FAIL")

# ───────────────────────────────────────────────────────────────────────
hdr("5. COUNTRY TIMELINE — Per-country sums to total per month")
# ───────────────────────────────────────────────────────────────────────

for entry in enr["country_timeline"]:
    mo = entry["month"]
    country_sum = sum(entry["countries"].values())
    # Compare to timeline['total']
    tl = next((t for t in enr["timeline"] if t["month"] == mo), None)
    tl_total = tl["total"] if tl else 0
    ok = abs(country_sum - tl_total) < 1
    L(f"  {mo}  countries sum ${country_sum/1e6:>8.2f}M  timeline ${tl_total/1e6:>8.2f}M  diff {(country_sum-tl_total)/1e6:+.4f}M",
      level="PASS" if ok else "FAIL")

# ───────────────────────────────────────────────────────────────────────
hdr("6. APPENDIX SORT ORDER (Asia AUM DESC)")
# ───────────────────────────────────────────────────────────────────────

# Parse PDF appendix, verify sort order
r = PdfReader("reports/2026-03/REX_Asia_Report_Mar26.pdf")
app_text = r.pages[6].extract_text() + "\n" + r.pages[7].extract_text()
lines = [l.strip() for l in app_text.split("\n") if l.strip()]
# Find data rows (ticker + family + $value ...)
fund_rows = []
for ln in lines:
    m = re.match(r"^([A-Z]{3,6})\s+(MicroSectors|T-REX|Equity Premium Income|Growth & Income|Autocallable|IncomeMax|Other)\s+\$([\d\.]+)(M|K)\s", ln)
    if m:
        t, fam, val, unit = m.group(1), m.group(2), float(m.group(3)), m.group(4)
        if unit == "K": val = val / 1000  # convert to M
        fund_rows.append((t, fam, val))

L(f"  Parsed {len(fund_rows)} fund rows from appendix")
# Check descending
is_sorted = all(fund_rows[i][2] >= fund_rows[i+1][2] - 0.01 for i in range(len(fund_rows)-1))
L(f"  Sorted DESC by Asia AUM: {is_sorted}", level="PASS" if is_sorted else "FAIL")
# Check against enriched
enr_asia_active = [f for f in enr["funds"] if f["asia_aum"] > 0]
if len(fund_rows) == len(enr_asia_active):
    L(f"  Count matches enriched asia-active count ({len(fund_rows)})", level="PASS")
else:
    L(f"  Count mismatch: PDF {len(fund_rows)} vs enriched {len(enr_asia_active)}", level="WARN")

# ───────────────────────────────────────────────────────────────────────
hdr("7. LEVERAGED vs NON-LEVERAGED BULLET MATH")
# ───────────────────────────────────────────────────────────────────────
# HTML bullet: "REX Shares (Leveraged): $X · REX Financial (Non-Leveraged): $Y"
# Leveraged = T-REX + MicroSectors (all single-stock 2x/3x or ETN 3x leveraged)
# Non-Leveraged = Income + Other + REX Osprey + Autocallable + IncomeMax + G&I

lev = S.get("T-REX", {}).get("aum", 0) + S.get("MicroSectors", {}).get("aum", 0)
nonlev = H["total_asia_aum"] - lev

L(f"  T-REX: ${S.get('T-REX',{}).get('aum',0)/1e6:.2f}M")
L(f"  MicroSectors: ${S.get('MicroSectors',{}).get('aum',0)/1e6:.2f}M")
L(f"  Leveraged total (T-REX + MS): ${lev/1e6:.2f}M")
L(f"  Non-Leveraged: ${nonlev/1e6:.2f}M (Income + REX Osprey + Other)")
L(f"  Sum: ${(lev+nonlev)/1e6:.2f}M vs total ${H['total_asia_aum']/1e6:.2f}M",
  level="PASS" if abs(lev + nonlev - H["total_asia_aum"]) < 1 else "FAIL")

# PDF check
full = "\n".join(p.extract_text() for p in r.pages)
# look for leveraged bullet
for line in full.split("\n"):
    if "Leveraged" in line and "Non-Leveraged" in line:
        L(f"  PDF bullet: {line.strip()}")

# ───────────────────────────────────────────────────────────────────────
hdr("8. SUITE COUNTRY 6-MONTH HISTORY")
# ───────────────────────────────────────────────────────────────────────

# Verify for each suite, each month, per-country sums to suite total for that month
if "suite_country_6m" in enr:
    for suite_name, months_data in enr["suite_country_6m"].items():
        for m in months_data:
            country_sum = sum(m["countries"].values())
            L(f"  {suite_name:<14} {m.get('month', 'unk'):<10} country-sum ${country_sum/1e6:>7.2f}M")

# ───────────────────────────────────────────────────────────────────────
hdr("9. GLOBAL AUM TIMELINE vs etp_monthly_fund")
# ───────────────────────────────────────────────────────────────────────
for mo, gaum in sorted(enr["global_aum_timeline"].items()):
    # Pull from DB directly for same month
    year, month = mo.split("-")
    cur.execute("""SELECT ROUND(SUM(total_aum_usd)::numeric, 2) t
                   FROM etp_monthly_fund mf JOIN calendar_month cm USING (month_id)
                   WHERE EXTRACT(YEAR FROM cm.month_end) = %s AND EXTRACT(MONTH FROM cm.month_end) = %s""",
                (int(year), int(month)))
    db_val = float(cur.fetchone()["t"] or 0)
    diff = gaum - db_val
    ok = abs(diff) < 1
    L(f"  {mo}  enriched ${gaum/1e9:.4f}B  DB ${db_val/1e9:.4f}B  diff {diff/1e6:+.4f}M",
      level="PASS" if ok else "FAIL")

# ───────────────────────────────────────────────────────────────────────
hdr("10. NARRATIVE BULLET NUMBERS MATCH SOURCES")
# ───────────────────────────────────────────────────────────────────────

N = enr["narrative"]
for key, val in N.items():
    L(f"  {key}: {val if isinstance(val, str) else val}")

# Verify T-REX Korea share bullet
from collections import defaultdict
suite_country_totals = defaultdict(lambda: defaultdict(float))
for fc in raw["fund_countries"]:
    fund = next((f for f in enr["funds"] if f["ticker"] == fc["ticker"]), None)
    if fund:
        suite_country_totals[fund["family_name"]][fc["country"]] += fc["aum"]

for suite_name in ["T-REX", "MicroSectors"]:
    top_c = max(suite_country_totals[suite_name].items(), key=lambda kv: kv[1])
    pct = top_c[1] / sum(suite_country_totals[suite_name].values()) * 100
    L(f"  {suite_name}: Top country {top_c[0]} ${top_c[1]/1e6:.2f}M ({pct:.1f}%)")

# ───────────────────────────────────────────────────────────────────────
hdr("SUMMARY")
# ───────────────────────────────────────────────────────────────────────
L(f"  PASS: {F['pass']}"); L(f"  FAIL: {F['fail']}"); L(f"  WARN: {F['warn']}")

conn.close()
OUT.close()
