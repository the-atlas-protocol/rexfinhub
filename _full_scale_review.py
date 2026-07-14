"""
Full-scale review for the Mar 31 2026 data quality.
Compares for EVERY REX ticker:
  - Bloomberg data_price (used for repricing in current pipeline)
  - Bloomberg data_nav (newly fixed by Ryu)
  - microsector sheet AUM (used for ETN global AUM)
  - yfinance close (external truth)

Goal: identify every Mar 31 data discrepancy that would corrupt the report.
"""
import openpyxl, sys, io, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
from datetime import date, timedelta

import yfinance as yf
import psycopg2
import psycopg2.extras

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")
OUT = open("_full_review.txt", "w", encoding="utf-8")
def L(*a): print(*a); print(*a, file=OUT); OUT.flush()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

def index(sheet):
    ws = wb[sheet]
    h = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cols = {c: i for i, c in enumerate(h) if c and " Equity" in str(c)}
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d and hasattr(d, 'year'):
            rows[d.date() if hasattr(d, 'date') else d] = row
    return cols, rows

p_cols, p_rows = index("data_price")
n_cols, n_rows = index("data_nav")

# microsector
ws_ms = wb["microsector"]
ms_tickers_row = list(next(ws_ms.iter_rows(min_row=4, max_row=4, values_only=True)))
ms_cols = {}
for i, t in enumerate(ms_tickers_row[1:], 1):
    if t and isinstance(t, str):
        ms_cols[t.strip()] = i
ms_rows = {}
for row in ws_ms.iter_rows(min_row=5, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year'):
        ms_rows[d.date() if hasattr(d, 'date') else d] = row

ms_tickers = set(ms_cols.keys())
L(f"Loaded {len(p_cols)} data_price tickers, {len(n_cols)} data_nav, {len(ms_cols)} microsector ETN tickers")

# DB tickers
conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
cur.execute("SELECT ticker, etp_id FROM etp ORDER BY ticker")
rex = [(r["ticker"], r["etp_id"]) for r in cur.fetchall()]
L(f"REX universe: {len(rex)} tickers")

def get_price(t, d):
    col = f"{t} US Equity"
    if col not in p_cols: return None
    row = p_rows.get(d)
    v = row[p_cols[col]] if row else None
    return float(v) if isinstance(v, (int, float)) and v > 0 else None

def get_nav(t, d):
    col = f"{t} US Equity"
    if col not in n_cols: return None
    row = n_rows.get(d)
    v = row[n_cols[col]] if row else None
    return float(v) if isinstance(v, (int, float)) and v > 0 else None

def get_ms_aum(t, d):
    if t not in ms_cols: return None
    row = ms_rows.get(d)
    v = row[ms_cols[t]] if row else None
    return float(v) if isinstance(v, (int, float)) and v > 0 else None

# ── Pull yfinance for ALL active REX tickers (Mar 30 - Apr 2) ──────
L(f"\nPulling yfinance closes for Mar 30, 31, Apr 1 ...")
yf_data = {}
TARGET_DATES = [date(2026,3,30), date(2026,3,31), date(2026,4,1)]

# yfinance batch download
tickers_str = " ".join(t for t,_ in rex)
try:
    df = yf.download(tickers_str, start="2026-03-25", end="2026-04-04",
                     auto_adjust=False, progress=False, group_by="ticker", threads=True)
    L(f"  yfinance returned shape: {df.shape}")
    # Process per-ticker
    for t, _ in rex:
        try:
            if t in df.columns.get_level_values(0):
                close_series = df[t]["Close"].dropna()
                yf_data[t] = {ts.date(): float(v) for ts, v in close_series.items()}
        except Exception:
            pass
    L(f"  yfinance pulled {len(yf_data)} tickers with data")
except Exception as e:
    L(f"  yfinance batch error: {e}")
    L(f"  Falling back to per-ticker (slow)...")
    for t, _ in rex:
        try:
            h = yf.Ticker(t).history(start="2026-03-25", end="2026-04-04", auto_adjust=False)
            yf_data[t] = {ts.date(): float(c) for ts, c in h["Close"].items()}
        except Exception:
            pass

# ── COMPARISON: Mar 31 — every ticker ──────────────────────────────
L(f"\n{'='*100}")
L(f"MAR 31 2026 — ALL TICKERS: yfinance close vs BBG data_price vs BBG data_nav vs microsector AUM")
L(f"{'='*100}")

issues = []
all_results = []
for t, etp_id in rex:
    is_etn = t in ms_tickers
    yf_close = yf_data.get(t, {}).get(date(2026,3,31))
    bbg_price = get_price(t, date(2026,3,31))
    bbg_nav = get_nav(t, date(2026,3,31))
    ms_aum = get_ms_aum(t, date(2026,3,31))

    rec = {"ticker": t, "is_etn": is_etn, "yf": yf_close, "price": bbg_price,
           "nav": bbg_nav, "ms_aum": ms_aum}
    all_results.append(rec)

    # Flag if available data disagrees
    if yf_close and bbg_price:
        gap_p = abs(yf_close - bbg_price) / yf_close
        if gap_p > 0.005:  # >0.5%
            issues.append((t, "PRICE", yf_close, bbg_price, gap_p, is_etn))
    if yf_close and bbg_nav:
        gap_n = abs(yf_close - bbg_nav) / yf_close
        if gap_n > 0.005:
            # ETN NAV may legitimately differ from market close
            if not is_etn:
                issues.append((t, "NAV", yf_close, bbg_nav, gap_n, is_etn))
            else:
                # tag ETN NAV gaps separately (informational)
                pass

# Header
L(f"{'Ticker':<8} {'Type':<4} {'yf 3/31':>9} {'price':>9} {'P-vs-yf':>9} {'NAV':>10} {'NAV-vs-yf':>10}")
for r in sorted(all_results, key=lambda x: x["ticker"]):
    t = r["ticker"]; is_etn = r["is_etn"]
    yfv = r["yf"]; pr = r["price"]; nv = r["nav"]
    fmt = lambda v: f"{v:>9.4f}" if v else f"{'—':>9}"
    fmt10 = lambda v: f"{v:>10.4f}" if v else f"{'—':>10}"
    p_gap = (pr - yfv) / yfv * 100 if (yfv and pr) else None
    n_gap = (nv - yfv) / yfv * 100 if (yfv and nv) else None
    L(f"{t:<8} {'ETN' if is_etn else 'ETF':<4} {fmt(yfv)} {fmt(pr)} "
      f"{(f'{p_gap:>+8.2f}%' if p_gap is not None else f'{'—':>9}')} "
      f"{fmt10(nv)} {(f'{n_gap:>+9.2f}%' if n_gap is not None else f'{'—':>10}')}")

# ── ISSUES SUMMARY ─────────────────────────────────────────────────
L(f"\n{'='*100}")
L(f"ISSUES SUMMARY")
L(f"{'='*100}")
L(f"\nTickers where BBG data_price differs from yfinance close by >0.5% on Mar 31:")
price_issues = [i for i in issues if i[1] == "PRICE"]
L(f"  Total: {len(price_issues)}")
etf_p = [i for i in price_issues if not i[5]]
etn_p = [i for i in price_issues if i[5]]
L(f"  ETFs: {len(etf_p)}, ETNs: {len(etn_p)}")
if price_issues:
    L(f"\n  {'Ticker':<8} {'Type':<4} {'yf':>9} {'BBG Price':>10} {'gap%':>8}")
    for t, kind, yfv, bbg, gap, is_etn in sorted(price_issues, key=lambda x: -x[4]):
        L(f"  {t:<8} {'ETN' if is_etn else 'ETF':<4} {yfv:>9.4f} {bbg:>10.4f} {gap*100:>7.2f}%")

L(f"\nETF NAV vs yfinance gaps on Mar 31 (ETN NAV ignored — those legitimately differ):")
etf_nav_issues = [i for i in issues if i[1] == "NAV" and not i[5]]
L(f"  Total ETF NAV issues: {len(etf_nav_issues)}")
for t, kind, yfv, bbg, gap, is_etn in sorted(etf_nav_issues, key=lambda x: -x[4])[:10]:
    L(f"  {t:<8} ETF  yf={yfv:.4f} NAV={bbg:.4f} gap={gap*100:.2f}%")

# ── microsector AUM sanity check ───────────────────────────────────
L(f"\n{'='*100}")
L(f"MICROSECTOR AUM (used for ETN global AUM) — Mar 31 vs Mar 30 vs Mar 27")
L(f"{'='*100}")
L(f"{'Ticker':<8} {'Mar 27':>14} {'Mar 30':>14} {'Mar 31':>14} {'30→31 %':>10} {'27→31 %':>10}")
for t in sorted(ms_tickers):
    a27 = get_ms_aum(t, date(2026,3,27))
    a30 = get_ms_aum(t, date(2026,3,30))
    a31 = get_ms_aum(t, date(2026,3,31))
    fmt = lambda v: f"${v/1e6:>10.2f}M" if v else f"{'—':>14}"
    pct_30_31 = (a31/a30 - 1)*100 if (a30 and a31) else None
    pct_27_31 = (a31/a27 - 1)*100 if (a27 and a31) else None
    L(f"{t:<8} {fmt(a27)} {fmt(a30)} {fmt(a31)} "
      f"{(f'{pct_30_31:>+9.2f}%' if pct_30_31 is not None else f'{'—':>10}')} "
      f"{(f'{pct_27_31:>+9.2f}%' if pct_27_31 is not None else f'{'—':>10}')}")

# ── DOWNSTREAM IMPACT ──────────────────────────────────────────────
L(f"\n{'='*100}")
L(f"DOWNSTREAM IMPACT — what's wrong in the current pipeline if data_price 3/31 is bad?")
L(f"{'='*100}")

# 1. etp_monthly_fund.price_usd for month 14
cur.execute("""
    SELECT e.ticker, mf.price_usd, mf.total_aum_usd
    FROM etp_monthly_fund mf JOIN etp e USING (etp_id)
    WHERE mf.month_id = 14 ORDER BY e.ticker
""")
db_state = {r["ticker"]: (float(r["price_usd"]) if r["price_usd"] else None,
                           float(r["total_aum_usd"]) if r["total_aum_usd"] else None) for r in cur.fetchall()}

# Compare DB price (which came from data_price) vs yf close
L(f"\nDB etp_monthly_fund.price_usd vs yfinance Mar 31 close — top discrepancies:")
db_price_issues = []
for t, etp_id in rex:
    db_price, db_aum = db_state.get(t, (None, None))
    yf_close = yf_data.get(t, {}).get(date(2026,3,31))
    if db_price and yf_close:
        gap = (db_price - yf_close) / yf_close
        if abs(gap) > 0.005:
            db_price_issues.append((t, db_price, yf_close, gap, t in ms_tickers))

L(f"  Tickers with DB price wrong by >0.5%: {len(db_price_issues)}")
for t, db, yfv, gap, is_etn in sorted(db_price_issues, key=lambda x: -abs(x[3]))[:25]:
    L(f"  {t:<8} {'ETN' if is_etn else 'ETF':<4} DB price={db:.4f} yf close={yfv:.4f} gap={gap*100:+.2f}%")

# 2. Implied DB shares (from db_aum / db_price) vs what they should be
L(f"\nImplied DB shares outstanding (db_aum / db_price) — should approximately match yfinance-implied shares")
shares_issues = 0
for t, etp_id in rex:
    db_price, db_aum = db_state.get(t, (None, None))
    yf_close = yf_data.get(t, {}).get(date(2026,3,31))
    if not (db_price and db_aum and yf_close): continue
    db_shares = db_aum / db_price
    yf_shares = db_aum / yf_close  # if AUM is right, this is the "true" shares
    gap = (db_shares - yf_shares) / yf_shares
    if abs(gap) > 0.005:
        shares_issues += 1
L(f"  Tickers with implied-shares discrepancy >0.5%: {shares_issues} of {len(rex)}")

# 3. % of global Asia for top funds — material impact?
L(f"\n% of global AUM in Asia (per fund) — recompute using yf close for ETFs")
cur.execute("""
    SELECT e.ticker, ROUND(SUM(m.exchange_aum_usd)::numeric, 2) asia_aum,
           ROUND(mf.total_aum_usd::numeric, 2) global_aum,
           ROUND(mf.price_usd::numeric, 4) price
    FROM etp_exchange_monthly_aum m JOIN etp e USING (etp_id)
    JOIN etp_monthly_fund mf ON mf.etp_id = m.etp_id AND mf.month_id = m.month_id
    WHERE m.month_id = 14 AND m.exchange_aum_usd > 0
    GROUP BY e.ticker, mf.total_aum_usd, mf.price_usd
    ORDER BY asia_aum DESC LIMIT 15
""")
top = cur.fetchall()
L(f"\n  Top 15 Asia-AUM funds — % of global recomputed with yfinance close:")
L(f"  {'Ticker':<8} {'Asia AUM':>12} {'Global (DB)':>12} {'pct DB':>8} {'pct (yf-implied)':>18}")
for r in top:
    t = r["ticker"]
    asia = float(r["asia_aum"]); glob_db = float(r["global_aum"]); price_db = float(r["price"])
    pct_db = asia / glob_db * 100 if glob_db else 0
    yf_close = yf_data.get(t, {}).get(date(2026,3,31))
    yf_glob = yf_close * (glob_db / price_db) if (yf_close and price_db) else None
    pct_yf = asia / yf_glob * 100 if yf_glob else None
    L(f"  {t:<8} ${asia/1e6:>10.2f}M ${glob_db/1e6:>10.2f}M {pct_db:>7.2f}% "
      f"{(f'{pct_yf:>17.2f}%' if pct_yf is not None else f'{'—':>18}')}")

wb.close(); conn.close(); OUT.close()
