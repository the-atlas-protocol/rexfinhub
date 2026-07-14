"""Build Seamus Harding's competitor-PAIRS Excel (REX L&I single-stock vs the
closest competitor on the SAME underlier + leverage + direction).
Read-only on the DB. Writes reports/seamus_li_pairs_<YYYY-MM-DD>.xlsx.
"""
import sqlite3
import datetime as dt
from pathlib import Path

import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO = Path("/home/jarvis/rexfinhub")
DB = REPO / "data" / "etp_tracker.db"
OUT_DIR = REPO / "reports"

conn = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
conn.row_factory = sqlite3.Row
rows = conn.execute(
    """
    SELECT ticker_clean AS tk,
           is_rex,
           COALESCE(NULLIF(issuer_display,''), NULLIF(issuer_nickname,''), issuer) AS brand,
           map_li_underlier        AS und,
           map_li_leverage_amount  AS lev,
           map_li_direction        AS dir,
           average_vol_30day       AS vol30
    FROM mkt_master_data
    WHERE map_li_subcategory = 'Single Stock'
      AND market_status = 'ACTV'
      AND map_li_underlier IS NOT NULL AND map_li_underlier != ''
      AND map_li_direction IS NOT NULL
      AND map_li_leverage_amount IS NOT NULL
    """
).fetchall()
conn.close()

products = []
for r in rows:
    tk = (r["tk"] or "").strip().upper()
    if not tk:
        continue
    products.append({
        "tk": tk,
        "is_rex": bool(r["is_rex"]),
        "brand": (r["brand"] or "").strip(),
        "und": (r["und"] or "").strip(),
        "lev": str(r["lev"]).strip(),
        "dir": (r["dir"] or "").strip(),
        "vol30": float(r["vol30"]) if r["vol30"] is not None else None,
    })

tickers = sorted({p["tk"] for p in products})
prices = {}
try:
    hist = yf.download(tickers, period="5d", progress=False,
                       auto_adjust=False, group_by="ticker")
    for t in tickers:
        try:
            closes = hist[t]["Close"].dropna() if len(tickers) > 1 else hist["Close"].dropna()
            if len(closes) > 0:
                prices[t] = float(closes.iloc[-1])
        except (KeyError, AttributeError, IndexError):
            pass
except Exception as e:
    print(f"WARN yfinance batch failed: {e}")

for p in products:
    p["price"] = prices.get(p["tk"])
    p["dvol"] = (p["price"] * p["vol30"]) if (p["price"] and p["vol30"]) else None

groups = {}
for p in products:
    groups.setdefault((p["und"], p["lev"], p["dir"]), []).append(p)

pairs = []
dropped = 0
for (und, lev, dir_), members in groups.items():
    rex = [m for m in members if m["is_rex"]]
    comp = [m for m in members if not m["is_rex"]]
    if not rex or not comp:
        continue
    for rx in rex:
        for cp in comp:
            if (rx["price"] is None or rx["dvol"] is None or
                    cp["price"] is None or cp["dvol"] is None):
                dropped += 1
                continue
            pairs.append({
                "und": und, "lev": lev, "dir": dir_,
                "rex_tk": rx["tk"], "rex_price": rx["price"], "rex_dvol": rx["dvol"],
                "cp_tk": cp["tk"], "cp_brand": cp["brand"],
                "cp_price": cp["price"], "cp_dvol": cp["dvol"],
            })

pairs.sort(key=lambda x: (x["und"], x["lev"], x["dir"], x["rex_tk"], x["cp_tk"]))

def fmt_price(v):
    return f"${v:,.2f}" if v is not None else "N/A"

def fmt_dvol(v):
    if v is None:
        return "N/A"
    if v >= 1e9:
        return f"${v/1e9:.1f}B"
    return f"${v/1e6:.1f}M"

def lev_label(lev):
    try:
        f = float(lev)
        return f"{int(f)}X" if f.is_integer() else f"{f:g}X"
    except ValueError:
        return str(lev)

def dir_label(d):
    return {"Long": "Long/Bull", "Short": "Short/Bear/Inverse"}.get(d, d)

today = dt.date.today().isoformat()
OUT_DIR.mkdir(exist_ok=True)
out_path = OUT_DIR / f"seamus_li_pairs_{today}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "L&I Competitor Pairs"

NAVY = "1F3864"
REX_BLUE = "D6E4F0"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
REX_FILL = PatternFill("solid", fgColor=REX_BLUE)
FLAG_FILL = PatternFill("solid", fgColor="FFC7CE")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

note = (
    f"REX L&I Single-Stock Competitor Pairs  |  generated {today}\n"
    "Comp basis: each pair is a REX leveraged/inverse single-stock ETF vs a competitor's "
    "L&I ETF on the SAME underlier + SAME leverage + SAME direction (e.g. MSTU vs MSTX = "
    "2X Long MSTR). 30d $Vol = 30-day average daily DOLLAR volume = average_vol_30day "
    "(Bloomberg) x latest close (yfinance). "
    "Context: leveraged ETFs trading under $3.00 face higher margin requirements; REX rows "
    "priced < $3.00 are flagged (reverse-split-under-$3 trigger; MSTU is the trigger case)."
)
ws.merge_cells("A1:L3")
c = ws["A1"]
c.value = note
c.font = Font(size=10, italic=True, color="404040")
c.alignment = Alignment(wrap_text=True, vertical="top")

headers = [
    "Underlier", "Leverage", "Direction",
    "REX Ticker", "REX Price", "REX 30d $Vol",
    "Competitor Ticker", "Competitor Issuer", "Comp Price", "Comp 30d $Vol",
    "$Vol Leader", "REX Under $3?",
]
rex_cols = {4, 5, 6}
HDR_ROW = 5
for ci, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=ci, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

flagged = set()
r = HDR_ROW + 1
for p in pairs:
    rex_under3 = p["rex_price"] < 3.00
    if rex_under3:
        flagged.add(p["rex_tk"])
    leader = "REX" if p["rex_dvol"] >= p["cp_dvol"] else p["cp_brand"]
    vals = [
        p["und"], lev_label(p["lev"]), dir_label(p["dir"]),
        p["rex_tk"], fmt_price(p["rex_price"]), fmt_dvol(p["rex_dvol"]),
        p["cp_tk"], p["cp_brand"], fmt_price(p["cp_price"]), fmt_dvol(p["cp_dvol"]),
        leader, "YES" if rex_under3 else "",
    ]
    for ci, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if ci in rex_cols:
            cell.fill = REX_FILL
            cell.font = Font(bold=True, size=10)
        else:
            cell.font = Font(size=10)
        if rex_under3 and ci in (5, 12):
            cell.fill = FLAG_FILL
            cell.font = Font(bold=True, color="9C0006", size=10)
    r += 1

widths = [11, 9, 17, 11, 11, 14, 16, 16, 11, 14, 13, 12]
for ci, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
ws.freeze_panes = "A6"

wb.save(out_path)

mstu = next((p for p in pairs if p["rex_tk"] == "MSTU" and p["cp_tk"] == "MSTX"), None)
print(f"OUT={out_path}")
print(f"PAIRS={len(pairs)}")
print(f"DROPPED={dropped}")
print(f"FLAGGED_UNDER_3={sorted(flagged)}")
if mstu:
    print(f"MSTU_MSTX: REX MSTU price={fmt_price(mstu['rex_price'])} dvol={fmt_dvol(mstu['rex_dvol'])} "
          f"| Defiance MSTX price={fmt_price(mstu['cp_price'])} dvol={fmt_dvol(mstu['cp_dvol'])} "
          f"| leader={'REX' if mstu['rex_dvol']>=mstu['cp_dvol'] else mstu['cp_brand']} "
          f"| rex_under3={mstu['rex_price']<3.0}")
else:
    print("MSTU_MSTX: NOT FOUND")
