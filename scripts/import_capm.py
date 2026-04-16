#!/usr/bin/env python3
"""Import Capital Markets Product List from Excel into capm_products table.

Reads per-suite operational sheets (T-REX, REX, REX-OSPREY, BMO) and the
ALL PRODUCTS LIST classification sheet, merges by ticker, and upserts into
the DB.

Usage:
    python scripts/import_capm.py
    python scripts/import_capm.py --file "path/to/file.xlsx"
    python scripts/import_capm.py --dry-run
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime, time
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
os.chdir(PROJECT_ROOT)

import pandas as pd
from sqlalchemy import text

from webapp.database import init_db, SessionLocal
from webapp.models import CapMProduct

DEFAULT_FILE = Path.home() / "Downloads" / "Capital Markets Product List .xlsx"

# Suite sheets and their column mappings (column index -> field name).
# T-REX, REX, REX-OSPREY share the same layout.
STANDARD_COLS = {
    0: "fund_name",
    1: "ticker",
    2: "bb_ticker",
    3: "inception_date",
    4: "trust",
    5: "exchange",
    6: "cu_size",
    7: "fixed_fee",
    8: "variable_fee",
    9: "cut_off",
    10: "custodian",
    11: "lmm",
    12: "prospectus_link",
}

# BMO has an extra first column (BMO Suites)
BMO_COLS = {
    0: "bmo_suite",
    1: "fund_name",
    2: "ticker",
    3: "bb_ticker",
    4: "inception_date",
    5: "issuer",
    6: "exchange",
    7: "cu_size",
    8: "fixed_fee",
    9: "variable_fee",
    10: "cut_off",
    11: "custodian",
    12: "lmm",
    13: "prospectus_link",
}

ALL_PRODUCTS_COLS = {
    0: "ticker",
    1: "fund_name",
    2: "inception_date",
    3: "our_category",
    4: "product_type",
    5: "category",
    6: "sub_category",
    7: "direction",
    8: "leverage",
    9: "underlying_ticker",
    10: "underlying_name",
    11: "expense_ratio",
    12: "competitor_products",
}


def _clean(v) -> str | None:
    """Clean a cell value to string or None."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return str(v)
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, (int, float)):
        # Keep numbers that aren't NaN
        if pd.isna(v):
            return None
        return str(v)
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan", "nat"):
        return None
    return s


def _to_date(v) -> date | None:
    """Parse various date formats into a date object."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan", "nat"):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_float(v) -> float | None:
    """Parse expense ratio or other float values."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if pd.isna(v):
            return None
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan", "#value!"):
        return None
    try:
        return float(s.replace("%", "").replace(",", ""))
    except ValueError:
        return None


def read_suite_sheet(wb_path: str, sheet_name: str, col_map: dict, suite_label: str) -> dict[str, dict]:
    """Read a per-suite sheet and return {ticker: field_dict}."""
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]

    products = {}
    # Row 1 is header; data starts at row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for col_idx, field_name in col_map.items():
            if col_idx < len(row):
                record[field_name] = row[col_idx]
            else:
                record[field_name] = None

        ticker = _clean(record.get("ticker"))
        fund_name = _clean(record.get("fund_name"))
        if not ticker or not fund_name:
            continue

        # Build cleaned record
        cleaned = {
            "fund_name": fund_name,
            "ticker": ticker,
            "bb_ticker": _clean(record.get("bb_ticker")),
            "inception_date": _to_date(record.get("inception_date")),
            "trust": _clean(record.get("trust")),
            "issuer": _clean(record.get("issuer")),
            "exchange": _clean(record.get("exchange")),
            "cu_size": _clean(record.get("cu_size")),
            "fixed_fee": _clean(record.get("fixed_fee")),
            "variable_fee": _clean(record.get("variable_fee")),
            "cut_off": _clean(record.get("cut_off")) if not isinstance(record.get("cut_off"), time) else record["cut_off"].strftime("%H:%M"),
            "custodian": _clean(record.get("custodian")),
            "lmm": _clean(record.get("lmm")),
            "prospectus_link": _clean(record.get("prospectus_link")),
            "suite_source": suite_label,
            "bmo_suite": _clean(record.get("bmo_suite")),
        }

        # BMO sheets carry forward the suite name from the first column
        # Handle merged cells: if bmo_suite is None but we're in BMO, carry forward
        if suite_label == "BMO" and not cleaned["bmo_suite"]:
            # Will be filled in post-processing
            pass

        products[ticker] = cleaned

    wb.close()
    return products


def read_all_products(wb_path: str) -> dict[str, dict]:
    """Read the ALL PRODUCTS LIST sheet and return {ticker: classification_dict}."""
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb["ALL PRODUCTS LIST"]

    products = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for col_idx, field_name in ALL_PRODUCTS_COLS.items():
            if col_idx < len(row):
                record[field_name] = row[col_idx]
            else:
                record[field_name] = None

        ticker = _clean(record.get("ticker"))
        if not ticker:
            continue

        products[ticker] = {
            "our_category": _clean(record.get("our_category")),
            "product_type": _clean(record.get("product_type")),
            "category": _clean(record.get("category")),
            "sub_category": _clean(record.get("sub_category")),
            "direction": _clean(record.get("direction")),
            "leverage": _clean(record.get("leverage")),
            "underlying_ticker": _clean(record.get("underlying_ticker")),
            "underlying_name": _clean(record.get("underlying_name")),
            "expense_ratio": _to_float(record.get("expense_ratio")),
            "competitor_products": _clean(record.get("competitor_products")),
            # Also capture fund_name and inception_date as fallback
            "fund_name": _clean(record.get("fund_name")),
            "inception_date": _to_date(record.get("inception_date")),
        }

    wb.close()
    return products


def fill_bmo_suites(products: dict[str, dict], wb_path: str) -> None:
    """Post-process BMO products to fill in bmo_suite from merged cells.

    The BMO sheet has the suite name (FANG+, Gold Miners, etc.) in column A,
    but only on the first row of each group (merged cells). We need to carry
    the value forward.
    """
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb["BMO"]

    current_suite = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        suite_val = _clean(row[0]) if len(row) > 0 else None
        ticker_val = _clean(row[2]) if len(row) > 2 else None

        if suite_val:
            current_suite = suite_val

        if ticker_val and ticker_val in products:
            products[ticker_val]["bmo_suite"] = current_suite

    wb.close()


def import_capm(file_path: str, dry_run: bool = False) -> dict:
    """Main import function. Returns summary stats."""
    print(f"Reading: {file_path}")

    # 1. Read per-suite operational sheets
    all_products: dict[str, dict] = {}

    suite_counts = {}
    for sheet_name, col_map, suite_label in [
        ("T-REX", STANDARD_COLS, "T-REX"),
        ("REX", STANDARD_COLS, "REX"),
        ("REX-OSPREY", STANDARD_COLS, "REX-OSPREY"),
        ("BMO", BMO_COLS, "BMO"),
    ]:
        suite_data = read_suite_sheet(file_path, sheet_name, col_map, suite_label)
        suite_counts[suite_label] = len(suite_data)
        print(f"  {suite_label}: {len(suite_data)} products")
        all_products.update(suite_data)

    # Post-process BMO suite names (merged cell carry-forward)
    fill_bmo_suites(all_products, file_path)

    # 2. Read ALL PRODUCTS LIST for classification data
    classification = read_all_products(file_path)
    print(f"  ALL PRODUCTS LIST: {len(classification)} products")

    # 3. Merge classification into operational data
    # Also add products that are ONLY in ALL PRODUCTS LIST (no suite sheet)
    for ticker, cls_data in classification.items():
        if ticker in all_products:
            # Merge classification fields into existing record
            for key in ("our_category", "product_type", "category", "sub_category",
                        "direction", "leverage", "underlying_ticker", "underlying_name",
                        "expense_ratio", "competitor_products"):
                if cls_data.get(key) is not None:
                    all_products[ticker][key] = cls_data[key]
        else:
            # Product only in ALL PRODUCTS LIST, no suite sheet data
            all_products[ticker] = {
                "fund_name": cls_data.get("fund_name", ticker),
                "ticker": ticker,
                "inception_date": cls_data.get("inception_date"),
                "suite_source": None,
                **{k: cls_data.get(k) for k in (
                    "our_category", "product_type", "category", "sub_category",
                    "direction", "leverage", "underlying_ticker", "underlying_name",
                    "expense_ratio", "competitor_products",
                )},
            }

    print(f"\n  Total unique products: {len(all_products)}")

    if dry_run:
        print("\n  [DRY RUN] No database changes made.")
        for ticker, data in sorted(all_products.items()):
            suite = data.get("suite_source") or "--"
            name = (data.get("fund_name") or "?")[:60]
            print(f"    {ticker:8s} | {suite:12s} | {name}")
        return {"total": len(all_products), "inserted": 0, "updated": 0}

    # 4. Upsert into database
    init_db()
    db = SessionLocal()
    inserted = 0
    updated = 0

    try:
        for ticker, data in all_products.items():
            existing = db.query(CapMProduct).filter(CapMProduct.ticker == ticker).first()

            if existing:
                # Update all fields
                for field, value in data.items():
                    if field == "ticker":
                        continue
                    if hasattr(existing, field) and value is not None:
                        setattr(existing, field, value)
                existing.updated_at = datetime.utcnow()
                updated += 1
            else:
                # Insert new
                product = CapMProduct(
                    fund_name=data.get("fund_name", ticker),
                    ticker=ticker,
                    bb_ticker=data.get("bb_ticker"),
                    inception_date=data.get("inception_date"),
                    trust=data.get("trust"),
                    issuer=data.get("issuer"),
                    exchange=data.get("exchange"),
                    cu_size=data.get("cu_size"),
                    fixed_fee=data.get("fixed_fee"),
                    variable_fee=data.get("variable_fee"),
                    cut_off=data.get("cut_off"),
                    custodian=data.get("custodian"),
                    lmm=data.get("lmm"),
                    prospectus_link=data.get("prospectus_link"),
                    suite_source=data.get("suite_source"),
                    our_category=data.get("our_category"),
                    product_type=data.get("product_type"),
                    category=data.get("category"),
                    sub_category=data.get("sub_category"),
                    direction=data.get("direction"),
                    leverage=data.get("leverage"),
                    underlying_ticker=data.get("underlying_ticker"),
                    underlying_name=data.get("underlying_name"),
                    expense_ratio=data.get("expense_ratio"),
                    competitor_products=data.get("competitor_products"),
                    bmo_suite=data.get("bmo_suite"),
                )
                db.add(product)
                inserted += 1

        db.commit()
        print(f"\n  Inserted: {inserted}")
        print(f"  Updated:  {updated}")
        print(f"  Total:    {inserted + updated}")
    except Exception as e:
        db.rollback()
        print(f"\n  ERROR: {e}")
        raise
    finally:
        db.close()

    return {"total": len(all_products), "inserted": inserted, "updated": updated}


def main():
    parser = argparse.ArgumentParser(description="Import Capital Markets Product List from Excel")
    parser.add_argument("--file", type=str, default=str(DEFAULT_FILE),
                        help="Path to the Excel file")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview without writing to database")
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"File not found: {file_path}")
        sys.exit(1)

    import_capm(str(file_path), dry_run=args.dry_run)


if __name__ == "__main__":
    main()
