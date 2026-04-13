"""Import REX Master Product Development Tracker from Excel into the database.

One-time migration. After this, rexfinhub DB is the source of truth.

Usage:
    python scripts/import_product_tracker.py
    python scripts/import_product_tracker.py --dry-run   # preview without writing
"""
from __future__ import annotations

import os
import sys
from datetime import date, datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
os.chdir(PROJECT_ROOT)

TRACKER_PATH = Path(
    r"C:\Users\RyuEl-Asmar\REX Financial LLC\REX Financial LLC - Rex Financial LLC"
    r"\Product Development\MasterFiles\MASTER Product Tracker"
    r"\REX Master Product Development Tracker.xlsm"
)

# Pipeline sheet: row 10 = headers, row 11+ = data
HEADER_ROW = 10
DATA_START = 11

# Column indices (0-based, from Pipeline sheet)
COL_MAP = {
    "name": 1,
    "trust": 2,
    "status": 3,
    "product_suite": 4,
    "ticker": 5,
    "initial_filing_date": 6,
    "estimated_effective_date": 7,
    "target_listing_date": 8,
    "seed_date": 9,
    "official_listed_date": 10,
    "latest_form": 11,
    "latest_prospectus_link": 12,
    "underlier": 13,
    "lmm": 14,
    "exchange": 15,
    "cik": 16,
    "series_id": 17,
    "class_contract_id": 18,
    "mgt_fee": 20,
    "tracking_index": 21,
    "fund_admin": 28,
    "cu_size": 34,
    "starting_nav": 38,
}


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s in ("None", "NaT", "0"):
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(val) -> float | None:
    if val is None:
        return None
    try:
        f = float(val)
        return f if f == f else None  # NaN check
    except (ValueError, TypeError):
        return None


def _parse_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _clean_str(val, max_len: int = 200) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == "None":
        return None
    return s[:max_len]


def load_from_excel() -> list[dict]:
    """Load product rows from the Pipeline sheet."""
    import openpyxl

    if not TRACKER_PATH.exists():
        print(f"ERROR: Tracker not found at {TRACKER_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(TRACKER_PATH, read_only=True, data_only=True)
    ws = wb["Pipeline"]

    products = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START, max_row=ws.max_row, values_only=True), DATA_START):
        # Skip empty rows
        name = _clean_str(row[COL_MAP["name"]] if len(row) > COL_MAP["name"] else None)
        status = _clean_str(row[COL_MAP["status"]] if len(row) > COL_MAP["status"] else None)
        if not name or not status:
            continue

        def _get(key):
            idx = COL_MAP.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        product = {
            "name": name,
            "trust": _clean_str(_get("trust")),
            "status": status,
            "product_suite": _clean_str(_get("product_suite")) or "Unknown",
            "ticker": _clean_str(_get("ticker"), 20),
            "underlier": _clean_str(_get("underlier"), 100),
            "initial_filing_date": _parse_date(_get("initial_filing_date")),
            "estimated_effective_date": _parse_date(_get("estimated_effective_date")),
            "target_listing_date": _parse_date(_get("target_listing_date")),
            "seed_date": _parse_date(_get("seed_date")),
            "official_listed_date": _parse_date(_get("official_listed_date")),
            "latest_form": _clean_str(_get("latest_form"), 20),
            "latest_prospectus_link": _clean_str(_get("latest_prospectus_link"), 500),
            "cik": _clean_str(_get("cik"), 20),
            "series_id": _clean_str(_get("series_id"), 20),
            "class_contract_id": _clean_str(_get("class_contract_id"), 20),
            "lmm": _clean_str(_get("lmm"), 100),
            "exchange": _clean_str(_get("exchange"), 20),
            "mgt_fee": _parse_float(_get("mgt_fee")),
            "tracking_index": _clean_str(_get("tracking_index")),
            "fund_admin": _clean_str(_get("fund_admin"), 100),
            "cu_size": _parse_int(_get("cu_size")),
            "starting_nav": _parse_float(_get("starting_nav")),
        }
        products.append(product)

    wb.close()
    return products


def import_to_db(products: list[dict], dry_run: bool = False):
    """Write products to the database."""
    from webapp.database import init_db, SessionLocal
    from webapp.models import RexProduct

    init_db()
    db = SessionLocal()

    # Check existing count
    existing = db.query(RexProduct).count()
    if existing > 0:
        print(f"WARNING: {existing} products already in DB. Skipping import (use --force to overwrite).")
        if "--force" not in sys.argv:
            db.close()
            return

        print(f"  Deleting {existing} existing products...")
        if not dry_run:
            db.query(RexProduct).delete()
            db.commit()

    # Status counts
    statuses = {}
    suites = {}
    for p in products:
        statuses[p["status"]] = statuses.get(p["status"], 0) + 1
        suites[p["product_suite"]] = suites.get(p["product_suite"], 0) + 1

    print(f"\nImporting {len(products)} products:")
    print("  Statuses:")
    for s, c in sorted(statuses.items(), key=lambda x: -x[1]):
        print(f"    {s}: {c}")
    print("  Suites:")
    for s, c in sorted(suites.items(), key=lambda x: -x[1]):
        print(f"    {s}: {c}")

    if dry_run:
        print("\n  DRY RUN -- no changes written.")
        db.close()
        return

    for p in products:
        db.add(RexProduct(**p))

    db.commit()
    final_count = db.query(RexProduct).count()
    print(f"\n  Imported {final_count} products to rex_products table.")
    db.close()


def main():
    dry_run = "--dry-run" in sys.argv
    print("=== REX Product Tracker Import ===")
    print(f"Source: {TRACKER_PATH}")
    if dry_run:
        print("MODE: Dry run (no DB writes)")

    products = load_from_excel()
    print(f"Loaded {len(products)} products from Excel.")
    import_to_db(products, dry_run=dry_run)


if __name__ == "__main__":
    main()
