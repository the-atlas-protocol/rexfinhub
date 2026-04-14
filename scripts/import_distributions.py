"""Import REX fund distributions from Excel into fund_distributions table.

Sources:
    REX_Distribution_Calendar_2026.xlsx    master schedule (Growth & Income suite)
    ATCL 2026 Distributions.xlsx           monthly autocallable
    TLDR 2026 Distributions (1).xlsx       weekly T-Bill

All 3 files share the same schema: Fund Name, Ticker, Declaration Date,
Ex Date, Record Date, Payable Date. Dedupes on (ticker, ex_date).

Also imports NYSE Holidays 2026 from the REX master file.

Usage:
    python scripts/import_distributions.py
    python scripts/import_distributions.py --dry-run
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
os.chdir(PROJECT_ROOT)

DOWNLOADS = Path.home() / "Downloads"

FILES = [
    DOWNLOADS / "REX_Distribution_Calendar_2026.xlsx",
    DOWNLOADS / "ATCL 2026 Distributions.xlsx",
    DOWNLOADS / "TLDR 2026 Distributions (1).xlsx",
]


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == "None":
        return None
    return s


def _to_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    try:
        return datetime.fromisoformat(str(v)).date()
    except Exception:
        return None


def _read_distribution_rows(wb, sheet_name: str, source_file: str) -> list[dict]:
    """Read one worksheet. Returns list of distribution dicts."""
    ws = wb[sheet_name]
    rows = []
    header = None
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
        if header is None:
            # First non-empty row with Ticker is the header
            if row and any(c and "ticker" in str(c).lower() for c in row[:8]):
                header = [str(c).strip() if c else "" for c in row[:8]]
                continue
            continue

        # Only process data rows with a ticker
        if len(row) < 6 or not row[1]:
            continue

        ticker = _clean(row[1])
        if not ticker or "ticker" in ticker.lower():
            continue

        # Normalize ticker to Bloomberg-style "ABC US" if it doesn't already have a suffix
        ticker_clean = ticker.upper().strip()
        if " US" not in ticker_clean and not any(x in ticker_clean for x in (" LN", " SW", " GR", " JP")):
            ticker_clean = f"{ticker_clean} US"

        ex_date = _to_date(row[3])
        if not ex_date:
            continue

        rows.append({
            "ticker": ticker_clean,
            "fund_name": _clean(row[0]),
            "declaration_date": _to_date(row[2]),
            "ex_date": ex_date,
            "record_date": _to_date(row[4]),
            "payable_date": _to_date(row[5]),
            "source_file": source_file,
        })
    return rows


def _read_holidays(wb, sheet_name: str) -> list[dict]:
    """Read NYSE Holidays sheet."""
    ws = wb[sheet_name]
    rows = []
    header = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not header:
            if row and row[0] and "holiday" in str(row[0]).lower():
                header = True
            continue
        if len(row) < 2 or not row[0]:
            continue
        name = _clean(row[0])
        hdate = _to_date(row[1])
        if not name or not hdate:
            continue
        rows.append({"holiday_date": hdate, "name": name, "note": _clean(row[3]) if len(row) > 3 else None})
    return rows


def main():
    parser = argparse.ArgumentParser(description="Import fund distributions from Excel")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--replace", action="store_true", help="Clear existing rows before import")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed")
        sys.exit(1)

    from webapp.database import init_db, SessionLocal
    from webapp.models import FundDistribution, NyseHoliday

    init_db()
    db = SessionLocal()

    all_rows = []
    all_holidays = []

    for path in FILES:
        if not path.exists():
            print(f"MISSING: {path}")
            continue
        print(f"Reading {path.name}...")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                lname = sheet_name.lower()
                if "holiday" in lname:
                    all_holidays.extend(_read_holidays(wb, sheet_name))
                elif "distribution" in lname or sheet_name == "GIF":
                    rows = _read_distribution_rows(wb, sheet_name, path.name)
                    print(f"  {sheet_name}: {len(rows)} distribution rows")
                    all_rows.extend(rows)
        finally:
            wb.close()

    print()
    print(f"Total distribution rows read: {len(all_rows)}")
    print(f"Total holiday rows read: {len(all_holidays)}")

    # Dedupe by (ticker, ex_date)
    seen = {}
    for r in all_rows:
        key = (r["ticker"], r["ex_date"])
        if key not in seen:
            seen[key] = r
        # keep the more-specific source file if both exist
    deduped_rows = list(seen.values())
    print(f"After dedupe: {len(deduped_rows)} unique (ticker, ex_date) rows")

    # Dedupe holidays by date
    holiday_seen = {}
    for h in all_holidays:
        if h["holiday_date"] not in holiday_seen:
            holiday_seen[h["holiday_date"]] = h
    deduped_holidays = list(holiday_seen.values())
    print(f"Unique holidays: {len(deduped_holidays)}")

    # Ticker breakdown
    by_ticker = {}
    for r in deduped_rows:
        by_ticker[r["ticker"]] = by_ticker.get(r["ticker"], 0) + 1
    print()
    print("Distributions per ticker:")
    for t in sorted(by_ticker.keys()):
        print(f"  {t}: {by_ticker[t]}")

    if args.dry_run:
        print("\nDRY RUN — no DB writes.")
        db.close()
        return

    # Write to DB (upsert by unique constraint)
    if args.replace:
        db.query(FundDistribution).delete()
        db.query(NyseHoliday).delete()
        db.commit()
        print("\nCleared existing rows.")

    dist_inserted = 0
    dist_skipped = 0
    existing_keys = {
        (r.ticker, r.ex_date)
        for r in db.query(FundDistribution.ticker, FundDistribution.ex_date).all()
    }
    for r in deduped_rows:
        if (r["ticker"], r["ex_date"]) in existing_keys:
            dist_skipped += 1
            continue
        db.add(FundDistribution(**r))
        dist_inserted += 1

    holiday_inserted = 0
    existing_hdates = {h.holiday_date for h in db.query(NyseHoliday.holiday_date).all()}
    for h in deduped_holidays:
        if h["holiday_date"] in existing_hdates:
            continue
        db.add(NyseHoliday(**h))
        holiday_inserted += 1

    db.commit()
    print(f"\nInserted {dist_inserted} new distribution rows ({dist_skipped} already existed)")
    print(f"Inserted {holiday_inserted} new holiday rows")
    db.close()


if __name__ == "__main__":
    main()
