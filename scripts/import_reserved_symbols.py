"""Import REX's reserved-symbol list from xlsx into reserved_symbols table.

Source: C:/Users/RyuEl-Asmar/Downloads/Reserved Symbols.xlsx (Master sheet,
~283 rows). Columns: Exchange, Symbol, End Date, Status, Rationale, Suite.

Usage:
    python scripts/import_reserved_symbols.py [--dry-run] [--source PATH]

Daily-safe: idempotent upsert by (exchange, symbol). Existing rows updated;
new rows inserted; rows in DB but not in xlsx are LEFT ALONE (so manually
added entries via /operations/reserved-symbols admin survive). On Render
the source file isn't present — script exits 0 with a "skipping" message.
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime, date
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

DEFAULT_SOURCE = Path("C:/Users/RyuEl-Asmar/Downloads/Reserved Symbols.xlsx")


def _coerce_date(val) -> date | None:
    """Excel often stores dates as serial integers (days since 1899-12-30)."""
    if val is None or val == "":
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    try:
        # Excel serial number
        serial = int(val)
        if 25569 <= serial <= 60000:  # plausible date range
            from datetime import timedelta
            return date(1899, 12, 30) + timedelta(days=serial)
    except (TypeError, ValueError):
        pass
    try:
        return date.fromisoformat(str(val)[:10])
    except ValueError:
        return None


def import_reserved_symbols(source: Path, dry_run: bool = False) -> dict:
    if not source.exists():
        print(f"  Source not found: {source} — skipping (OK on Render).")
        return {"inserted": 0, "updated": 0, "skipped": True}

    import openpyxl
    from webapp.database import init_db, SessionLocal
    from webapp.models import ReservedSymbol
    from sqlalchemy import select

    init_db()
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    if "Master" not in wb.sheetnames:
        print(f"  ERROR: 'Master' sheet not found. Sheets: {wb.sheetnames}")
        wb.close()
        return {"inserted": 0, "updated": 0, "error": "no Master sheet"}

    ws = wb["Master"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    # Parse into records
    records = []
    for r in rows:
        if not r or r[1] is None:  # need at least Symbol (col B)
            continue
        rec = {
            "exchange": str(r[0]).strip() if r[0] else None,
            "symbol": str(r[1]).strip().upper(),
            "end_date": _coerce_date(r[2]),
            "status": str(r[3]).strip() if r[3] else None,
            "rationale": str(r[4]).strip() if r[4] else None,
            "suite": str(r[5]).strip() if len(r) > 5 and r[5] else None,
        }
        records.append(rec)

    print(f"  Read {len(records)} rows from xlsx.")
    if dry_run:
        for r in records[:5]:
            print(f"    SAMPLE: {r}")
        return {"inserted": 0, "updated": 0, "dry_run": True, "rows": len(records)}

    # Upsert
    db = SessionLocal()
    inserted = 0
    updated = 0
    try:
        for rec in records:
            existing = db.execute(
                select(ReservedSymbol)
                .where(ReservedSymbol.exchange == rec["exchange"])
                .where(ReservedSymbol.symbol == rec["symbol"])
            ).scalar_one_or_none()
            if existing:
                # Update only if any field changed
                changed = False
                for k in ("end_date", "status", "rationale", "suite"):
                    if getattr(existing, k) != rec[k]:
                        setattr(existing, k, rec[k])
                        changed = True
                if changed:
                    existing.updated_at = datetime.utcnow()
                    updated += 1
            else:
                db.add(ReservedSymbol(**rec))
                inserted += 1
        db.commit()
    finally:
        db.close()

    print(f"  Inserted: {inserted}")
    print(f"  Updated:  {updated}")
    return {"inserted": inserted, "updated": updated}


def main():
    ap = argparse.ArgumentParser(description="Import REX reserved symbols from xlsx.")
    ap.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    print(f"Importing from: {args.source}")
    result = import_reserved_symbols(args.source, dry_run=args.dry_run)
    if result.get("error"):
        sys.exit(1)


if __name__ == "__main__":
    main()
