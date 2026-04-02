"""
Generate competitor analysis Excel for Asia team.

Reads from local SQLite DB (mkt_master_data), identifies competitors for
each REX suite, converts CUSIPs to ISINs, and produces a multi-sheet
Excel workbook with formatting.

Usage:
    python scripts/generate_competitor_excel.py
"""
from __future__ import annotations

import os
import subprocess
import sys
from datetime import date
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
os.chdir(PROJECT_ROOT)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from webapp.database import SessionLocal

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SUITE_MAPPING_CSV = PROJECT_ROOT / "config" / "rules" / "rex_suite_mapping.csv"
COMP_GROUPS_CSV = PROJECT_ROOT / "config" / "rules" / "competitor_groups.csv"
OUTPUT_DIR = PROJECT_ROOT / "reports"

# Explicit tickers always included in "Equity Premium Income" peer group
EPI_EXPLICIT_TICKERS = [
    "JEPI US", "JEPQ US", "GPIX US", "QYLD US",
    "QQQI US", "IWMI US", "SPYI US",
]

# Style constants
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
GROUP_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
GROUP_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
REX_HIGHLIGHT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)
PCT_FORMAT = "0.00%"
AUM_FORMAT = '#,##0.0'
VOL_FORMAT = '#,##0'
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")


# ---------------------------------------------------------------------------
# CUSIP to ISIN
# ---------------------------------------------------------------------------

def cusip_to_isin(cusip: str, country: str = "US") -> str:
    """Convert 9-char CUSIP to 12-char ISIN with Luhn check digit."""
    if not cusip or len(str(cusip)) < 9:
        return ""
    base = country + str(cusip)[:9]
    digits = ""
    for c in base:
        if c.isdigit():
            digits += c
        elif c.isalpha():
            digits += str(ord(c.upper()) - 55)
        else:
            return ""
    # Luhn check digit
    total = 0
    for i, d in enumerate(reversed(digits)):
        n = int(d)
        if i % 2 == 0:
            n *= 2
            if n > 9:
                n -= 9
        total += n
    check = (10 - (total % 10)) % 10
    return base + str(check)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_active_funds() -> pd.DataFrame:
    """Load all ACTV funds from mkt_master_data."""
    db = SessionLocal()
    try:
        rows = db.execute(text(
            "SELECT ticker, fund_name, issuer_display, cusip, aum, "
            "  category_display, etp_category, is_rex, rex_suite, "
            "  total_return_1day, total_return_1week, total_return_1month, "
            "  total_return_3month, total_return_ytd, total_return_1year, "
            "  average_vol_30day, open_interest, inception_date, "
            "  total_return_6month "
            "FROM mkt_master_data "
            "WHERE market_status = 'ACTV'"
        )).fetchall()
    finally:
        db.close()

    cols = [
        "ticker", "fund_name", "issuer", "cusip", "aum",
        "category_display", "etp_category", "is_rex", "rex_suite",
        "return_1d", "return_1w", "return_1m",
        "return_3m", "return_ytd", "return_1y",
        "avg_vol_30d", "open_interest", "inception_date",
        "return_6m",
    ]
    df = pd.DataFrame(rows, columns=cols)

    # Normalize types
    df["is_rex"] = df["is_rex"].astype(bool)
    df["aum"] = pd.to_numeric(df["aum"], errors="coerce").fillna(0.0)

    # CUSIP -> ISIN (zero-pad to 9 chars)
    df["cusip"] = df["cusip"].astype(str).str.strip()
    df["cusip"] = df["cusip"].apply(
        lambda x: x.zfill(9) if x and x not in ("None", "nan", "") and len(x) <= 9 else x
    )
    df["isin"] = df["cusip"].apply(
        lambda x: cusip_to_isin(x) if x and x not in ("None", "nan", "") else ""
    )

    return df


def load_suite_mapping() -> dict[str, str]:
    """Load rex_suite_mapping.csv -> {ticker: suite}."""
    df = pd.read_csv(
        SUITE_MAPPING_CSV, engine="python", on_bad_lines="skip"
    )
    df = df.dropna(subset=["ticker", "rex_suite"])
    return dict(zip(df["ticker"].str.strip(), df["rex_suite"].str.strip()))


def load_competitor_groups() -> pd.DataFrame:
    """Load competitor_groups.csv."""
    df = pd.read_csv(
        COMP_GROUPS_CSV, engine="python", on_bad_lines="skip"
    )
    df = df.dropna(subset=["group_name", "rex_ticker", "peer_ticker"])
    for col in ["group_name", "rex_ticker", "peer_ticker"]:
        df[col] = df[col].str.strip()
    return df


# ---------------------------------------------------------------------------
# Competitor identification
# ---------------------------------------------------------------------------

def assign_competitor_suites(df: pd.DataFrame) -> pd.DataFrame:
    """Assign a 'peer_suite' column indicating which REX suite each row competes with.

    A fund can compete in multiple suites, so we return an exploded DF with
    one row per (ticker, peer_suite) pair.
    """
    records: list[dict] = []

    for _, row in df.iterrows():
        suites: set[str] = set()

        if row["is_rex"]:
            # REX products get their own suite
            if row["rex_suite"]:
                suites.add(row["rex_suite"])
        else:
            cat = row["category_display"] or ""
            etp = row["etp_category"] or ""
            fname = (row["fund_name"] or "").upper()
            ticker = row["ticker"] or ""

            # T-REX: single stock L&I
            if cat == "Leverage & Inverse - Single Stock":
                suites.add("T-REX")

            # MicroSectors: index/basket L&I
            if cat == "Leverage & Inverse - Index/Basket/ETF Based":
                suites.add("MicroSectors")

            # Equity Premium Income: explicit tickers + all "Income" category
            if ticker in EPI_EXPLICIT_TICKERS:
                suites.add("Equity Premium Income")
            if "Income" in cat:
                suites.add("Equity Premium Income")

            # Growth & Income: single stock income
            if cat == "Income - Single Stock":
                suites.add("Growth & Income")

            # Autocallable: income + "autocall" in name
            if "Income" in cat and "AUTOCALL" in fname:
                suites.add("Autocallable")

            # Crypto
            if etp == "Crypto":
                suites.add("Crypto")

            # Thematic
            if etp == "Thematic":
                suites.add("Thematic")

        for s in suites:
            records.append({"ticker": row["ticker"], "peer_suite": s})

    suite_df = pd.DataFrame(records)
    if suite_df.empty:
        df["peer_suite"] = None
        return df

    # Merge back -- one row per (ticker, peer_suite)
    merged = df.merge(suite_df, on="ticker", how="inner")
    return merged


# ---------------------------------------------------------------------------
# Sheet 1: All Competitors
# ---------------------------------------------------------------------------

SHEET1_COLUMNS = [
    ("Ticker", "ticker", None),
    ("Fund Name", "fund_name", None),
    ("ISIN", "isin", None),
    ("Issuer", "issuer", None),
    ("REX Comparison", "peer_suite", None),
    ("Underlier", "_underlier", None),
    ("AUM ($M)", "aum", AUM_FORMAT),
    ("30D Avg Volume", "avg_vol_30d", VOL_FORMAT),
]


def _apply_return_font(cell, value):
    """Apply green/red font to return cells."""
    if value is not None and not pd.isna(value):
        if value > 0:
            cell.font = Font(name="Calibri", size=10, color="006100")
        elif value < 0:
            cell.font = Font(name="Calibri", size=10, color="9C0006")


def write_all_competitors(wb: Workbook, df: pd.DataFrame):
    """Write 'All Competitors' sheet grouped by suite, sorted by AUM."""
    ws = wb.active
    ws.title = "All Competitors"

    # Define suite display order
    suite_order = [
        "T-REX", "MicroSectors", "Equity Premium Income",
        "Growth & Income", "Autocallable", "Crypto", "Thematic",
        "IncomeMax", "T-Bill",
    ]

    current_row = 1

    # Add _underlier column from is_singlestock / map_li_underlier
    if "is_singlestock" not in df.columns:
        # Load it from DB
        _db = SessionLocal()
        try:
            ul_rows = _db.execute(text("SELECT ticker, is_singlestock FROM mkt_master_data")).fetchall()
            ul_map = {r[0]: str(r[1] or "").replace(" US","").replace(" Equity","").replace(" Curncy","").strip() for r in ul_rows}
        finally:
            _db.close()
        df["_underlier"] = df["ticker"].map(ul_map).fillna("")
    else:
        df["_underlier"] = df["is_singlestock"].astype(str).apply(
            lambda x: x.replace(" US","").replace(" Equity","").replace(" Curncy","").strip() if x and x != "nan" else ""
        )

    for suite in suite_order:
        group = df[df["peer_suite"] == suite].copy()
        if group.empty:
            continue

        group = group.sort_values("aum", ascending=False)

        # Suite header row (merged)
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(SHEET1_COLUMNS),
        )
        cell = ws.cell(row=current_row, column=1,
                       value=f"{suite} ({len(group)} products)")
        cell.fill = GROUP_FILL
        cell.font = GROUP_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Column headers
        for col_idx, (header, _, _) in enumerate(SHEET1_COLUMNS, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Data rows
        for i, (_, row) in enumerate(group.iterrows()):
            for col_idx, (_, field, fmt) in enumerate(SHEET1_COLUMNS, 1):
                value = row.get(field)
                if pd.isna(value):
                    value = None
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                if fmt:
                    cell.number_format = fmt
                if i % 2 == 1:
                    cell.fill = ALT_ROW_FILL
                cell.border = THIN_BORDER
                cell.font = Font(name="Calibri", size=10)
            current_row += 1

        # Blank separator row
        current_row += 1

    # Column widths
    _auto_fit_columns(ws, len(SHEET1_COLUMNS), current_row)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 2: Summary
# ---------------------------------------------------------------------------

def write_summary(wb: Workbook, df: pd.DataFrame):
    """Write pivot summary by issuer and suite."""
    ws = wb.create_sheet("Summary")

    suite_order = [
        "T-REX", "MicroSectors", "Equity Premium Income",
        "Growth & Income", "Autocallable", "Crypto", "Thematic",
    ]
    # Filter to relevant suites
    suites_present = [s for s in suite_order if s in df["peer_suite"].unique()]

    # Build pivot: issuer -> suite -> (count, aum)
    pivot: dict[str, dict[str, tuple[int, float]]] = {}
    for _, row in df.iterrows():
        issuer = row["issuer"] or "Unknown"
        suite = row["peer_suite"]
        if suite not in suites_present:
            continue
        if issuer not in pivot:
            pivot[issuer] = {}
        if suite not in pivot[issuer]:
            pivot[issuer][suite] = (0, 0.0)
        cnt, aum = pivot[issuer][suite]
        pivot[issuer][suite] = (cnt + 1, aum + (row["aum"] or 0.0))

    # Sort issuers by total AUM descending
    issuer_totals = {}
    for issuer, suites in pivot.items():
        issuer_totals[issuer] = sum(aum for _, aum in suites.values())
    sorted_issuers = sorted(issuer_totals, key=issuer_totals.get, reverse=True)

    # Suite totals for market share
    suite_totals: dict[str, float] = {}
    for suites in pivot.values():
        for suite, (_, aum) in suites.items():
            suite_totals[suite] = suite_totals.get(suite, 0.0) + aum
    grand_total = sum(suite_totals.values())

    # Headers
    headers = ["Issuer"]
    for s in suites_present:
        headers.extend([f"{s} (#)", f"{s} AUM ($M)", f"{s} Share (%)"])
    headers.extend(["Total (#)", "Total AUM ($M)", "Total Share (%)"])

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Category totals row
    row_num = 2
    ws.cell(row=row_num, column=1, value="CATEGORY TOTAL").font = Font(
        name="Calibri", bold=True, size=10
    )
    col = 2
    for s in suites_present:
        total_cnt = sum(
            pivot.get(iss, {}).get(s, (0, 0.0))[0] for iss in sorted_issuers
        )
        total_aum = suite_totals.get(s, 0.0)
        ws.cell(row=row_num, column=col, value=total_cnt).number_format = VOL_FORMAT
        ws.cell(row=row_num, column=col + 1, value=total_aum).number_format = AUM_FORMAT
        ws.cell(row=row_num, column=col + 2, value=1.0).number_format = PCT_FORMAT
        col += 3
    grand_cnt = sum(
        sum(cnt for cnt, _ in suites.values()) for suites in pivot.values()
    )
    ws.cell(row=row_num, column=col, value=grand_cnt).number_format = VOL_FORMAT
    ws.cell(row=row_num, column=col + 1, value=grand_total).number_format = AUM_FORMAT
    ws.cell(row=row_num, column=col + 2, value=1.0).number_format = PCT_FORMAT
    for c in range(1, len(headers) + 1):
        ws.cell(row=row_num, column=c).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row_num, column=c).border = Border(
            bottom=Side(style="medium", color="1F3864")
        )
    row_num += 1

    # Issuer rows
    for i, issuer in enumerate(sorted_issuers):
        ws.cell(row=row_num, column=1, value=issuer)
        col = 2
        total_cnt = 0
        total_aum = 0.0
        for s in suites_present:
            cnt, aum = pivot.get(issuer, {}).get(s, (0, 0.0))
            total_cnt += cnt
            total_aum += aum
            share = aum / suite_totals[s] if suite_totals.get(s, 0) > 0 else 0.0
            ws.cell(row=row_num, column=col, value=cnt if cnt else None).number_format = VOL_FORMAT
            ws.cell(row=row_num, column=col + 1, value=aum if aum else None).number_format = AUM_FORMAT
            ws.cell(row=row_num, column=col + 2, value=share if share else None).number_format = PCT_FORMAT
            col += 3
        grand_share = total_aum / grand_total if grand_total > 0 else 0.0
        ws.cell(row=row_num, column=col, value=total_cnt).number_format = VOL_FORMAT
        ws.cell(row=row_num, column=col + 1, value=total_aum).number_format = AUM_FORMAT
        ws.cell(row=row_num, column=col + 2, value=grand_share).number_format = PCT_FORMAT

        # Alternating row color
        if i % 2 == 1:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).fill = ALT_ROW_FILL
        # Highlight REX row
        if issuer == "REX Financial":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).fill = REX_HIGHLIGHT_FILL

        for c in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=c).border = THIN_BORDER
        row_num += 1

    _auto_fit_columns(ws, len(headers), row_num)
    ws.freeze_panes = "B3"


# ---------------------------------------------------------------------------
# Sheet 3: Key Comps
# ---------------------------------------------------------------------------

def write_key_comps(wb: Workbook, df_all: pd.DataFrame, comp_groups: pd.DataFrame):
    """Write Key Comps as ONE flat table. No section separators. REX Comparison + Underlier columns."""
    ws = wb.create_sheet("Key Comps")

    db = SessionLocal()
    try:
        rows = db.execute(text(
            "SELECT ticker, fund_name, issuer_display, cusip, aum, "
            "  category_display, etp_category, is_rex, rex_suite, "
            "  total_return_1day, total_return_1week, total_return_1month, "
            "  total_return_3month, total_return_ytd, total_return_1year, "
            "  average_vol_30day, open_interest, is_singlestock, map_li_underlier, "
            "  map_li_direction, map_crypto_underlier "
            "FROM mkt_master_data WHERE market_status = 'ACTV'"
        )).fetchall()
    finally:
        db.close()

    comp_df = pd.DataFrame(rows, columns=[
        "ticker", "fund_name", "issuer", "cusip", "aum",
        "category_display", "etp_category", "is_rex", "rex_suite",
        "return_1d", "return_1w", "return_1m", "return_3m", "return_ytd", "return_1y",
        "avg_vol_30d", "open_interest", "is_singlestock", "map_li_underlier",
        "map_li_direction", "map_crypto_underlier",
    ])
    comp_df["aum"] = pd.to_numeric(comp_df["aum"], errors="coerce").fillna(0)
    comp_df["is_rex"] = comp_df["is_rex"].astype(bool)
    comp_df["cusip"] = comp_df["cusip"].astype(str).str.strip()
    comp_df["cusip"] = comp_df["cusip"].apply(lambda x: x.zfill(9) if x and x not in ("None","nan","") and len(x)<=9 else x)
    comp_df["isin"] = comp_df["cusip"].apply(lambda x: cusip_to_isin(x) if x and x not in ("None","nan","") else "")

    rex_df = comp_df[comp_df["is_rex"]].copy()
    non_rex = comp_df[~comp_df["is_rex"]].copy()

    ncols = len(SHEET1_COLUMNS)

    # Write headers
    for col_idx, (header, _, _) in enumerate(SHEET1_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    current_row = 2

    def _add(r, comparison, underlier):
        nonlocal current_row
        row_idx = current_row - 2  # for alternating
        for col_idx, (_, field, fmt) in enumerate(SHEET1_COLUMNS, 1):
            if field == "peer_suite":
                value = comparison
            elif field == "_underlier":
                value = underlier
            else:
                value = r.get(field)
                if value is not None and pd.notna(value):
                    pass
                else:
                    value = None
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if fmt:
                cell.number_format = fmt
            if row_idx % 2 == 1:
                cell.fill = ALT_ROW_FILL
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)
        current_row += 1

    # ── T-REX: 1 Long + 1 Short per underlier ──
    trex_rex = rex_df[rex_df["rex_suite"] == "T-REX"].sort_values("aum", ascending=False)
    li_ss = non_rex[non_rex["category_display"] == "Leverage & Inverse - Single Stock"].copy()
    seen_ul = set()
    for _, rr in trex_rex.iterrows():
        ul = str(rr.get("is_singlestock") or "")
        if not ul or ul == "nan":
            continue
        ul_clean = ul.replace(" US","").replace(" Equity","").replace(" Curncy","").strip()
        if ul_clean in seen_ul:
            continue
        seen_ul.add(ul_clean)
        matches = li_ss[li_ss["is_singlestock"].astype(str).str.contains(ul_clean, na=False)]
        if matches.empty:
            continue
        rex_t = rr["ticker"].replace(" US","")
        longs = matches[matches["map_li_direction"].astype(str).str.lower().str.contains("long", na=False)]
        if not longs.empty:
            _add(longs.nlargest(1,"aum").iloc[0], f"T-REX ({rex_t})", ul_clean)
        shorts = matches[matches["map_li_direction"].astype(str).str.lower().str.contains("short", na=False)]
        if not shorts.empty:
            _add(shorts.nlargest(1,"aum").iloc[0], f"T-REX ({rex_t})", ul_clean)
        if longs.empty and shorts.empty:
            _add(matches.nlargest(1,"aum").iloc[0], f"T-REX ({rex_t})", ul_clean)

    # ── L&I Index/Basket: Top 20 ──
    ms_comps = non_rex[non_rex["category_display"] == "Leverage & Inverse - Index/Basket/ETF Based"]
    for _, r in ms_comps.nlargest(20, "aum").iterrows():
        _add(r, "MicroSectors", str(r.get("is_singlestock") or r.get("map_li_underlier") or ""))

    # ── EPI: Top 20 ──
    epi_comps = non_rex[non_rex["category_display"].str.contains("Income", na=False)]
    for _, r in epi_comps.nlargest(20, "aum").iterrows():
        ul = str(r.get("is_singlestock") or "").replace(" US","").replace(" Equity","").strip()
        _add(r, "EPI", ul or "Broad")

    # ── G&I: Top 5 per underlier ──
    gi_rex = rex_df[rex_df["rex_suite"] == "Growth & Income"].sort_values("aum", ascending=False)
    gi_comps = non_rex[non_rex["category_display"] == "Income - Single Stock"]
    for _, rr in gi_rex.iterrows():
        ul = str(rr.get("is_singlestock") or "")
        if not ul or ul == "nan":
            continue
        ul_clean = ul.replace(" US","").replace(" Equity","").strip()
        rex_t = rr["ticker"].replace(" US","")
        matches = gi_comps[gi_comps["is_singlestock"].astype(str).str.contains(ul_clean, na=False)]
        for _, r in matches.nlargest(5, "aum").iterrows():
            _add(r, f"G&I ({rex_t})", ul_clean)

    # ── Autocallable: Top 5 ──
    auto_comps = non_rex[non_rex["fund_name"].str.upper().str.contains("AUTOCALL", na=False)]
    for _, r in auto_comps.nlargest(5, "aum").iterrows():
        _add(r, "Autocallable (ATCL)", "")

    # ── DRNZ: Defense/Aerospace + JEDI ──
    defense_kw = "DEFENSE|DEFENCE|AEROSPACE|DRONE|MILITARY|WEAPON|ITA |PPA |XAR "
    drnz_comps = non_rex[
        non_rex["fund_name"].str.upper().str.contains(defense_kw, na=False, regex=True) |
        (non_rex["ticker"] == "JEDI US")
    ]
    for _, r in drnz_comps.nlargest(10, "aum").iterrows():
        _add(r, "Thematic (DRNZ)", "Defense/Aerospace")

    # ── Crypto: by underlier (SOL, DOGE, XRP, BTC, ETH) ──
    crypto_rex = {"SSK": "Solana", "DOJE": "Dogecoin", "XRPR": "XRP", "OBTC": "Bitcoin", "ESK": "Ethereum"}
    crypto_comps = non_rex[non_rex["etp_category"] == "Crypto"]
    for rex_t, coin in crypto_rex.items():
        matches = crypto_comps[
            crypto_comps["fund_name"].str.upper().str.contains(coin.upper(), na=False) |
            crypto_comps["map_crypto_underlier"].astype(str).str.contains(coin, na=False, case=False)
        ]
        for _, r in matches.nlargest(5, "aum").iterrows():
            _add(r, f"Crypto ({rex_t})", coin)

    _auto_fit_columns(ws, ncols, current_row)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 4: REX Products
# ---------------------------------------------------------------------------

SHEET4_COLUMNS = [
    ("Ticker", "ticker", None),
    ("Fund Name", "fund_name", None),
    ("ISIN", "isin", None),
    ("Suite", "rex_suite", None),
    ("AUM ($M)", "aum", AUM_FORMAT),
    ("1D Return (%)", "return_1d", PCT_FORMAT),
    ("1W Return (%)", "return_1w", PCT_FORMAT),
    ("1M Return (%)", "return_1m", PCT_FORMAT),
    ("3M Return (%)", "return_3m", PCT_FORMAT),
    ("YTD Return (%)", "return_ytd", PCT_FORMAT),
    ("1Y Return (%)", "return_1y", PCT_FORMAT),
    ("30D Avg Volume", "avg_vol_30d", VOL_FORMAT),
    ("Open Interest", "open_interest", VOL_FORMAT),
    ("Inception Date", "inception_date", None),
]

RETURN_FIELDS_SHEET4 = {
    "return_1d", "return_1w", "return_1m", "return_3m", "return_ytd", "return_1y",
}


def write_rex_products(wb: Workbook, df: pd.DataFrame):
    """Write REX Products sheet."""
    ws = wb.create_sheet("REX Products")

    rex = df[df["is_rex"]].copy()
    rex = rex.sort_values(["rex_suite", "aum"], ascending=[True, False])
    # Deduplicate by ticker (same ticker can appear multiple times from suite explosion)
    rex = rex.drop_duplicates(subset=["ticker"])

    # Headers
    for col_idx, (header, _, _) in enumerate(SHEET4_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for i, (_, row) in enumerate(rex.iterrows()):
        for col_idx, (_, field, fmt) in enumerate(SHEET4_COLUMNS, 1):
            value = row.get(field)
            if pd.isna(value):
                value = None

            if field in RETURN_FIELDS_SHEET4 and value is not None:
                value = value / 100.0

            cell = ws.cell(row=row_num, column=col_idx, value=value)
            if fmt:
                cell.number_format = fmt
            if field in RETURN_FIELDS_SHEET4:
                _apply_return_font(cell, row.get(field))

            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL
            cell.border = THIN_BORDER

        row_num += 1

    _auto_fit_columns(ws, len(SHEET4_COLUMNS), row_num)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _auto_fit_columns(ws, num_cols: int, num_rows: int):
    """Auto-fit column widths based on content (capped at 40)."""
    for col_idx in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(num_rows + 1, 200)):  # sample first 200 rows
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        # Add padding, cap at 40
        width = min(max_len + 3, 40)
        width = max(width, 8)  # minimum width
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading data from database...")
    df = load_active_funds()
    print(f"  Loaded {len(df)} active funds ({df['is_rex'].sum()} REX products)")

    print("Loading suite mapping...")
    suite_map = load_suite_mapping()
    print(f"  {len(suite_map)} REX tickers mapped to suites")

    print("Loading competitor groups...")
    comp_groups = load_competitor_groups()
    print(f"  {len(comp_groups)} peer mappings loaded")

    print("Assigning competitor suites...")
    df_comp = assign_competitor_suites(df)
    df_comp = df_comp[df_comp["peer_suite"].notna()].copy()

    # COMPETITORS ONLY — exclude REX products
    df_comps_only = df_comp[~df_comp["is_rex"]].copy()
    print(f"  {len(df_comps_only)} competitors across suites:")
    for suite, cnt in df_comps_only["peer_suite"].value_counts().items():
        print(f"    {suite}: {cnt}")

    print("Building Excel workbook (competitors only, 2 sheets)...")
    wb = Workbook()

    # Sheet 1: Full competitor list (no REX products)
    write_all_competitors(wb, df_comps_only)

    # Sheet 2: Key comps per REX fund
    write_key_comps(wb, df, comp_groups)

    # Save
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"competitor_analysis_{date.today()}.xlsx"
    wb.save(str(output_path))
    print(f"Saved to {output_path}")

    # Open the file
    subprocess.Popen(["start", "", str(output_path)], shell=True)


if __name__ == "__main__":
    main()
