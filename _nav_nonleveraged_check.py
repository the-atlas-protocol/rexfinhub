"""
1) Check NAV vs Price for NON-leveraged products around Mar 31.
2) Pull yfinance close prices as external truth for select tickers.
3) Try to identify what day's price each Mar 31 NAV value actually corresponds to.
"""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
from datetime import date, timedelta

import yfinance as yf
import psycopg2
import psycopg2.extras

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")
OUT = open("_nav_nonlev.txt", "w", encoding="utf-8")
def L(*a): print(*a); print(*a, file=OUT); OUT.flush()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

def index(sheet):
    ws = wb[sheet]
    h = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cols = {c: i for i, c in enumerate(h) if c and " Equity" in str(c)}
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d and hasattr(d, 'year'):
            rows[d.date() if hasattr(d, 'date') else d] = row
    return cols, rows

p_cols, p_rows = index("data_price")
n_cols, n_rows = index("data_nav")

def get(rows, cols, ticker, dt):
    col = f"{ticker} US Equity"
    if col not in cols: return None
    row = rows.get(dt)
    if not row: return None
    v = row[cols[col]]
    return float(v) if isinstance(v, (int, float)) and v > 0 else None

# Identify non-leveraged products from DB
conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia",
                        cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
cur.execute("""
    SELECT e.ticker, pf.name family
    FROM etp e JOIN product_family pf USING (family_id)
    WHERE pf.name IN ('Income', 'REX Osprey', 'Other')
    ORDER BY pf.name, e.ticker
""")
nonlev = [r["ticker"] for r in cur.fetchall()]
L(f"Non-leveraged tickers ({len(nonlev)}): {nonlev}")

# ── Step 1: Mar 31 NAV/Price ratio for non-leveraged products ──────
WINDOW_DATES = [date(2026,3,27), date(2026,3,30), date(2026,3,31), date(2026,4,1), date(2026,4,2)]
L(f"\n{'='*100}")
L(f"NON-LEVERAGED PRODUCTS — NAV vs Price around Mar 31")
L(f"{'='*100}")
L(f"{'Ticker':<8}  " + "  ".join(f"{d.strftime('%a %m/%d'):<11}" for d in WINDOW_DATES))

for t in nonlev:
    ratios = []
    for d in WINDOW_DATES:
        n = get(n_rows, n_cols, t, d)
        p = get(p_rows, p_cols, t, d)
        if n and p:
            ratios.append(f"{n/p:>10.4f}")
        else:
            ratios.append(f"{'—':>10}")
    L(f"  {t:<6}  " + "  ".join(f"{x:<11}" for x in ratios))

# Also show absolute NAVs for non-leveraged on Mar 31 vs nearby days
L(f"\nNon-leveraged NAVs (raw values) around Mar 31:")
L(f"{'Ticker':<8}  {'Mar 27':<10}  {'Mar 30':<10}  {'Mar 31':<10}  {'Apr 1':<10}  {'Apr 2':<10}")
for t in nonlev[:30]:
    row = []
    for d in WINDOW_DATES:
        v = get(n_rows, n_cols, t, d)
        row.append(f"{v:>10.4f}" if v else f"{'—':>10}")
    L(f"  {t:<6}  " + "  ".join(row))

# ── Step 2: yfinance external truth for select tickers ─────────────
L(f"\n{'='*100}")
L(f"YFINANCE EXTERNAL TRUTH — close prices around Mar 31 2026")
L(f"{'='*100}")

# Pick 5 representative tickers spanning leveraged + non-leveraged
SAMPLE = ["TSLT", "MSTU", "FEPI", "AIPI", "BMNU"]
L(f"Pulling yfinance for: {SAMPLE}")
L("")

for t in SAMPLE:
    try:
        hist = yf.Ticker(t).history(start="2026-03-25", end="2026-04-04", auto_adjust=False)
        if hist.empty:
            L(f"  {t}: no yfinance data")
            continue
        L(f"  {t}:")
        L(f"    {'Date':<11} {'yfinance Close':>14} {'BBG Price':>12} {'BBG NAV':>12} {'NAV/Close':>11}")
        for ts, row in hist.iterrows():
            d = ts.date()
            yclose = row["Close"]
            bbg_price = get(p_rows, p_cols, t, d)
            bbg_nav = get(n_rows, n_cols, t, d)
            nav_to_close = (bbg_nav / yclose) if (bbg_nav and yclose) else None
            L(f"    {str(d):<11} {yclose:>14.4f} "
              f"{bbg_price if bbg_price else '—':>12} "
              f"{bbg_nav if bbg_nav else '—':>12} "
              f"{nav_to_close if nav_to_close else '—':>11}")
    except Exception as e:
        L(f"  {t}: yfinance error {e}")

# ── Step 3: Try to identify which day's value Mar 31 NAV matches ───
L(f"\n{'='*100}")
L(f"DAY-MATCH HUNT — for each ticker's Mar 31 NAV, which day's PRICE does it most closely match?")
L(f"{'='*100}")
SEARCH_DATES = [date(2026,3,25)+timedelta(days=i) for i in range(15)]

L(f"{'Ticker':<8}  {'Mar31 NAV':>10}  best-match day(s)")
for t in nonlev[:15] + ["TSLT", "MSTU", "BMNU", "BULZ", "GDXU"]:
    n31 = get(n_rows, n_cols, t, date(2026,3,31))
    if not n31: continue
    candidates = []
    for d in SEARCH_DATES:
        p_val = get(p_rows, p_cols, t, d)
        if p_val:
            ratio = n31 / p_val
            if 0.998 <= ratio <= 1.002:  # within 0.2%
                candidates.append((d, p_val, ratio))
    cand_str = ", ".join(f"{d}=${p:.4f} (×{r:.4f})" for d,p,r in candidates) if candidates else "no day matches within 0.2%"
    L(f"  {t:<6}  {n31:>10.4f}  {cand_str}")

wb.close(); conn.close(); OUT.close()
