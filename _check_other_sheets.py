"""Check other Bloomberg sheets (data_notional, flowcalcs, data_nav, bbg_pull, repull_168) for Mar 31 total."""
import openpyxl
from pathlib import Path
from datetime import date
import psycopg2

BB = Path(r"C:/Users/RyuEl-Asmar/REX Financial LLC/REX Financial LLC - MasterFiles/MASTER Data/bloomberg_daily_file.xlsm")
OUT = open("_sheets.txt", "w", encoding="utf-8")
def P(*a, **k): print(*a, **k, file=OUT); OUT.flush()

conn = psycopg2.connect(host="localhost", port=5433, user="postgres", dbname="rex_asia")
cur = conn.cursor()
cur.execute("SELECT ticker FROM etp")
rex_tickers = set(r[0] for r in cur.fetchall())
conn.close()

wb = openpyxl.load_workbook(BB, data_only=True, read_only=True)

TARGET = date(2026, 3, 31)

def sum_rex_on_sheet(sheet_name, header_row=1, date_col=1, data_start_row=2):
    """Try to sum REX tickers from a sheet at the target date."""
    if sheet_name not in wb.sheetnames:
        P(f"  {sheet_name}: NOT FOUND")
        return
    ws = wb[sheet_name]
    try:
        header = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    except StopIteration:
        P(f"  {sheet_name}: empty")
        return
    ticker_cols = {}
    for i, col in enumerate(header):
        if col and isinstance(col, str) and " Equity" in col:
            parts = col.split()
            base = parts[0]
            listing = parts[1] if len(parts) > 1 else "US"
            key = f"{base}_LN" if listing == "LN" else base
            ticker_cols[key] = i
    if not ticker_cols:
        P(f"  {sheet_name}: no ticker columns in header row {header_row}")
        return
    # Find target date row
    found_row = None
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        d = row[date_col - 1]
        if d and hasattr(d, 'year') and d.year == TARGET.year and d.month == TARGET.month and d.day == TARGET.day:
            found_row = row
            break
    if not found_row:
        P(f"  {sheet_name}: no row for {TARGET}")
        return
    total = 0; n = 0
    for t in rex_tickers:
        if t in ticker_cols:
            v = found_row[ticker_cols[t]]
            if isinstance(v, (int, float)) and v > 0:
                total += v
                n += 1
    P(f"  {sheet_name}: {n} tickers matched, raw sum = {total:,.0f}  (if $M, that's ${total/1e3:,.1f}B)")

P(f"=== Target date: {TARGET} ===")
P()
P("Trying each candidate sheet (header row 1, data from row 2):")
for s in ["data_aum", "data_notional", "data_flow", "data_nav", "data_price", "bbg_pull"]:
    sum_rex_on_sheet(s, header_row=1, date_col=1, data_start_row=2)

P()
P("bbg_pull (header row 2, dates from row 3 — AUM preface on row 1):")
sum_rex_on_sheet("bbg_pull", header_row=2, date_col=1, data_start_row=3)

P()
P("repull_168:")
sum_rex_on_sheet("repull_168", header_row=1, date_col=1, data_start_row=2)

P()
P("flowcalcs — show structure (first 15 rows, all cols):")
if "flowcalcs" in wb.sheetnames:
    ws = wb["flowcalcs"]
    for r in range(1, min(16, ws.max_row + 1)):
        row_vals = []
        for c in range(1, min(16, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            s = str(v) if v is not None else ""
            if len(s) > 15: s = s[:12] + "..."
            row_vals.append(s)
        P(f"  r{r}: " + " | ".join(row_vals))

OUT.close()
wb.close()
